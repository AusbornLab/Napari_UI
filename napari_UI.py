import numpy as np
import json
import csv
import sys
import os
import time
import napari
import torch
from datetime import datetime
from pathlib import Path
from qtpy.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QCheckBox,
    QSlider, QGroupBox, QFileDialog, QApplication, QMessageBox, QProgressBar,
    QSpinBox, QDialog, QColorDialog, QSplitter, QListWidget, QAbstractItemView,
    QListWidgetItem, QDialogButtonBox, QDoubleSpinBox, QInputDialog, QLineEdit, QComboBox, QScrollArea
)
from qtpy.QtCore import Qt
from napari.qt.threading import thread_worker
from skimage.filters import gaussian, threshold_otsu
from skimage.morphology import remove_small_objects, closing, disk, dilation, ball, binary_opening, binary_closing
from aicsimageio import AICSImage
import tifffile
import imageio
from skimage.transform import rotate
from skimage.measure import label, regionprops
from skimage.segmentation import watershed  
import matplotlib.pyplot as plt
from qtpy.QtGui import QIntValidator
from skimage.draw import polygon as sk_polygon  
from scipy import ndimage as ndi
from skimage.draw import polygon2mask
import re
import xml.etree.ElementTree as ET
from skimage.feature import peak_local_max
from cellpose import models, train
from cellpose import io as cellpose_io
import traceback
from scipy.spatial import cKDTree
from skimage.measure import marching_cubes








###############################################################################
# ----------------------------- FILE LOADING ---------------------------------
###############################################################################

def load_lif_memmap(path, scene_index=0, max_ch=4):
    """Load LIF using memory mapping for instant slice access without RAM overhead.
    Returns (list_of_memmaps, contrast_limits_dict, voxel_size(Z,Y,X), paths).
    """
    print("Creating memory-mapped arrays from LIF (memmap)...")
    img = AICSImage(str(path))

    # Determine number of channels
    try:
        c_total = int(img.dims.C)
    except Exception:
        c_total = 1

    n_ch = min(max_ch, max(1, c_total))

    # Get the underlying dask array in CZYX order if possible
    try:
        arr = img.get_image_dask_data("CZYX", S=scene_index)
    except Exception:
        arr = img.get_image_dask_data("CYX", S=scene_index)

    print(f"Image array shape (expected C,Z,Y,X or C,Y,X): {arr.shape}")

    voxel_size = (1.0, 1.0, 1.0)
    try:
        p = getattr(img, "physical_pixel_sizes", None)
        if p is not None:
            # NamedTuple: PhysicalPixelSizes(Z=..., Y=..., X=...) 
            if hasattr(p, "Z") and hasattr(p, "Y") and hasattr(p, "X"):
                z = p.Z
                y = p.Y
                x = p.X
            else:
                # tuple-like; docs indicate order Z,Y,X 
                z, y, x = tuple(p)[:3]

            if z is not None and y is not None and x is not None:
                voxel_size = (float(z), float(y), float(x))
        else:
            # Older AICSImageIO API: returns X,Y,Z 
            if hasattr(img, "get_physical_pixel_size"):
                pxyz = img.get_physical_pixel_size(scene_index)
                if isinstance(pxyz, (list, tuple)) and len(pxyz) >= 3:
                    x, y, z = pxyz[:3]
                    voxel_size = (float(z), float(y), float(x))
    except Exception:
        pass

    # Create temporary memory-mapped files
    temp_dir = Path.cwd() / "napari_mmap_files"
    temp_dir.mkdir(parents=True, exist_ok=True)

    # Normalize arr to shape (C,Z,Y,X) where possible
    if arr.ndim == 4:
        c, z, y, x = arr.shape
    elif arr.ndim == 3:
        # could be (C,Y,X) -> treat Z=1
        c = arr.shape[0]
        z = 1
        y = arr.shape[1]
        x = arr.shape[2]
    else:
        raise ValueError(f"Unsupported LIF array shape: {arr.shape}")

    n_z = z

    # create memmaps for up to n_ch channels
    memmaps = []
    paths = []
    for ch_idx in range(min(n_ch, c)):
        pth = temp_dir / f"ch{ch_idx+1}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.dat"
        mm = np.memmap(pth, dtype=np.float32, mode='w+', shape=(n_z, y, x))
        memmaps.append(mm)
        paths.append(pth)

    print("Writing data to memory-mapped files (this may take a few minutes)...")
    chunk_size = 30
    start_time = time.time()

    for idx, start_z in enumerate(range(0, n_z, chunk_size)):
        end_z = min(start_z + chunk_size, n_z)
        for ch_i in range(len(memmaps)):
            if arr.ndim == 4:
                chunk = arr[ch_i, start_z:end_z].compute().astype(np.float32)
            else:
                chunk = arr[ch_i][np.newaxis, ...].astype(np.float32)

            memmaps[ch_i][start_z:end_z] = chunk
            memmaps[ch_i].flush()

        elapsed = time.time() - start_time
        chunks_done = idx + 1
        total_chunks = (n_z + chunk_size - 1) // chunk_size
        remaining_chunks = total_chunks - chunks_done
        time_per_chunk = elapsed / chunks_done
        est_remaining = remaining_chunks * time_per_chunk
        print(f"  Processed slices {start_z}-{end_z-1} of {n_z-1} ({elapsed:.1f}s elapsed, ~{est_remaining:.1f}s remaining)")

    print("Memory-mapped files created successfully!")

    # Calculate contrast limits per channel
    contrast_limits = {}
    mid_z = max(0, n_z // 2)
    for i, mm in enumerate(memmaps, start=1):
        try:
            sample = mm[mid_z]
            lo, hi = np.percentile(sample, [1, 99.5])
        except Exception:
            lo, hi = 0, 65535
        contrast_limits[f'ch{i}'] = (float(lo), float(hi))

    print(f"Voxel size (Z, Y, X): {voxel_size}")
    return memmaps, contrast_limits, voxel_size, paths


def _read_estimated_voxel_size(path):
    with tifffile.TiffFile(str(path)) as tif:
        voxel_z = None
        voxel_y = None
        voxel_x = None
        unit = None

        # ImageJ metadata
        try:
            ij = tif.imagej_metadata
        except Exception:
            ij = None

        if ij is not None:
            voxel_z = ij.get("spacing", None)
            unit = ij.get("unit", None)

        # TIFF tags
        page = tif.pages[0]
        tags = page.tags

        if "XResolution" in tags:
            try:
                num, den = tags["XResolution"].value
                if num != 0:
                    voxel_x = float(den) / float(num)
            except Exception:
                pass

        if "YResolution" in tags:
            try:
                num, den = tags["YResolution"].value
                if num != 0:
                    voxel_y = float(den) / float(num)
            except Exception:
                pass

        if "ResolutionUnit" in tags:
            try:
                ru = tags["ResolutionUnit"].value
                print("Raw ResolutionUnit:", ru)
            except Exception:
                ru = None
        else:
            ru = None

        z = 1.0 if voxel_z is None else float(voxel_z)
        y = 1.0 if voxel_y is None else float(voxel_y)
        x = 1.0 if voxel_x is None else float(voxel_x)

        print("=== Estimated voxel size ===")
        print("Z:", z, unit)
        print("Y:", y, unit)
        print("X:", x, unit)

        return (z, y, x)

def load_tiff_memmap(path, max_ch=4):
    """Load TIFF using memory-mapped arrays via tifffile, write float32 memmaps.
    Returns (list_of_memmaps, contrast_limits_dict, voxel_size(Z,Y,X), paths).
    Tries to extract voxel size from:
      1) OME-TIFF OME-XML
      2) ImageJ metadata
      3) TIFF resolution tags
    """


    print("Creating memory-mapped arrays from TIFF (memmap) using tifffile...")
    # ---------- read pixels (prefer memmap) ----------
    try:
        data = tifffile.memmap(str(path))
        arr = np.asarray(data)
    except Exception:
        try:
            arr = tifffile.imread(str(path))
        except Exception as e:
            # last resort: AICSImage -> delegate
            img = AICSImage(str(path))
            try:
                arr = img.get_image_dask_data("CZYX")
            except Exception:
                arr = img.get_image_dask_data("CYX")
            try:
                c_total = int(img.dims.C)
            except Exception:
                c_total = arr.shape[0] if arr.ndim == 4 else 1
            return load_lif_memmap(path, scene_index=0, max_ch=min(max_ch, max(1, c_total)))


    # ---------- extract voxel size ----------
    voxel_size = _read_estimated_voxel_size(path)
    # ---------- normalize array into (C, Z, Y, X) ----------
    arr = np.asarray(arr)
    if arr.ndim == 4:
        a0, a1, a2, a3 = arr.shape
        if a0 <= 4 and a1 > 4:
            arr_czyx = arr  # (C,Z,Y,X)
        elif a1 <= 4 and a0 > 4:
            arr_czyx = np.moveaxis(arr, [0, 1], [1, 0])  # (Z,C,Y,X)->(C,Z,Y,X)
        else:
            arr_czyx = arr
    elif arr.ndim == 3:
        arr_czyx = np.expand_dims(arr, axis=0)  # (1,Z,Y,X)
    elif arr.ndim == 2:
        arr_czyx = arr.reshape((1, 1) + arr.shape)  # (1,1,Y,X)
    else:
        arr_czyx = np.reshape(arr, (1, 1) + arr.shape[-2:])


    c_total = arr_czyx.shape[0]
    n_ch = min(max_ch, max(1, c_total))


    # ---------- write float32 memmaps ----------
    temp_dir = Path.cwd() / "napari_mmap_files"
    temp_dir.mkdir(parents=True, exist_ok=True)


    n_z, y, x = arr_czyx.shape[1], arr_czyx.shape[2], arr_czyx.shape[3]


    memmaps = []
    paths = []
    for ch_idx in range(n_ch):
        p = temp_dir / f"ch{ch_idx+1}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.dat"
        mm = np.memmap(p, dtype=np.float32, mode='w+', shape=(n_z, y, x))
        memmaps.append(mm)
        paths.append(p)


    print("Writing data to memory-mapped files (this may take a few minutes)...")
    chunk_size = 30
    start_time = time.time()
    for idx, start_z in enumerate(range(0, n_z, chunk_size)):
        end_z = min(start_z + chunk_size, n_z)
        for ch_i in range(n_ch):
            chunk = arr_czyx[ch_i, start_z:end_z].astype(np.float32, copy=False)
            memmaps[ch_i][start_z:end_z] = chunk
            memmaps[ch_i].flush()


        elapsed = time.time() - start_time
        chunks_done = idx + 1
        total_chunks = (n_z + chunk_size - 1) // chunk_size
        remaining_chunks = total_chunks - chunks_done
        time_per_chunk = elapsed / chunks_done
        est_remaining = remaining_chunks * time_per_chunk
        print(f"  Processed slices {start_z}-{end_z-1} of {n_z-1} ({elapsed:.1f}s elapsed, ~{est_remaining:.1f}s remaining)")


    # ---------- contrast limits ----------
    contrast_limits = {}
    mid_z = max(0, n_z // 2)
    for i, mm in enumerate(memmaps, start=1):
        try:
            sample = mm[mid_z]
            lo, hi = np.percentile(sample, [1, 99.5])
        except Exception:
            lo, hi = 0, 65535
        contrast_limits[f'ch{i}'] = (float(lo), float(hi))


    print(f"Voxel size (Z, Y, X): {voxel_size}")
    print("Memory-mapped TIFF files created successfully!")
    return memmaps, contrast_limits, voxel_size, paths

###############################################################################
# ----------------------------- MASK COMPUTATION-----------------------------
###############################################################################

def compute_mask(img2d, threshold, blur_enabled=False, blur_sigma=1.0,
                 min_size=0, do_close=False, close_radius=3):
    data = np.asarray(img2d).astype(np.float32)
    if blur_enabled:
        data = gaussian(data, sigma=blur_sigma, preserve_range=True)

    mask = data > float(threshold)

    if min_size and min_size > 0:
        mask = remove_small_objects(mask, min_size=min_size)  # can delete everything 
    if do_close:
        mask = closing(mask, footprint=disk(int(close_radius)))

    return mask.astype(np.uint8)
###############################################################################
# ----------------------- ANALYSIS POPUP DIALOG ----------------------------
###############################################################################

class AnalysisPopupDialog(QDialog):
    """Separate popup for analysis buttons."""
    
    def __init__(self, parent_widget):
        super().__init__()
        self.parent_widget = parent_widget
        self.setWindowTitle("Analysis Tools")
        self.setGeometry(100, 100, 400, 300)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()

        # Keep only the pared-down buttons requested in implementation.txt
        save_btn = QPushButton("Save Metadata")
        save_btn.clicked.connect(self.parent_widget.save_output)
        layout.addWidget(save_btn)

        load_meta_btn = QPushButton("Load Metadata")
        load_meta_btn.clicked.connect(self.parent_widget.load_metadata)
        layout.addWidget(load_meta_btn)

        load_channels_btn = QPushButton("Load Channels...")
        load_channels_btn.clicked.connect(self.parent_widget.load_channels)
        layout.addWidget(load_channels_btn)

        export_btn = QPushButton("Export Selected Channels")
        export_btn.clicked.connect(self.parent_widget.export_selected_layers)
        layout.addWidget(export_btn)

        coloc_btn = QPushButton("Colocalize Z Stack")
        coloc_btn.clicked.connect(self.parent_widget.generate_coloc_entire_stack)
        layout.addWidget(coloc_btn)

        maxproj_btn = QPushButton("Generate Max Projection")
        maxproj_btn.clicked.connect(self.parent_widget.generate_max_projection)
        layout.addWidget(maxproj_btn)

        # Keep screenshot/video buttons associated with projections
        self.video_btn = QPushButton("Save Video of Current View")
        self.video_btn.clicked.connect(self.parent_widget.save_video)
        layout.addWidget(self.video_btn)

        self.screenshot_btn = QPushButton("Save Screenshot of Current View")
        self.screenshot_btn.clicked.connect(
            lambda _=False: self.parent_widget.save_screenshot(
                viewer=self.parent_widget.viewer,
                parent=self,
                default_name="main_viewer.png",
                canvas_only=True,
            )
        )
        layout.addWidget(self.screenshot_btn)

        # 3D surfaces from selected layers
        self.surfaces3d_btn = QPushButton("Generate 3D Surfaces (Marching Cubes)")
        self.surfaces3d_btn.clicked.connect(lambda _=False: self.parent_widget.generate_3d_surfaces())
        layout.addWidget(self.surfaces3d_btn)

        self.cellcount_btn = QPushButton("Cell Counter (ROI examples → detect)")
        self.cellcount_btn.clicked.connect(lambda _=False: self.parent_widget.open_cell_counter())
        layout.addWidget(self.cellcount_btn)

        self.punctacount_btn = QPushButton("Puncta Counter (Count puncta in with Mask ROI)")
        self.punctacount_btn.clicked.connect(lambda _=False: self.parent_widget.open_puncta_counter())
        layout.addWidget(self.punctacount_btn)

        layout.addStretch()
        self.setLayout(layout)

###############################################################################
#Mask tuner for individual channel mask genearation and preview before writing to main viewer
class MaskTunerDialog(QDialog):
    """
    Interactive per-channel threshold/mask tuner.

    - Embedded napari viewer (2D slice view driven by this dialog's Z slider)
    - Right-side per-layer controls for channel + mask (visible/opacity/color/contrast)
    - Save/Load metadata that includes ALL widget + layer control settings

    Updated behavior:
    - Always refreshes source data from the parent widget's CURRENT channel layer/data.
    - Picks up cropped/replaced channel stacks from the main viewer.
    - Keeps dialog Z synchronized with the parent viewer's Z when possible.
    """

    def __init__(self, parent_widget, channel_index, channel_data, file_path=None):
        super().__init__(parent_widget)

        # ---- basic state ----
        self.parent_widget = parent_widget
        self.channel_index = int(channel_index)  # 0-based
        self.file_path = str(file_path) if file_path is not None else str(
            getattr(parent_widget, "file_path", "unknown")
        )

        self._syncing_parent_z = False
        self._syncing_local_z = False

        # Pull CURRENT source data from parent if available; fall back to provided data
        src = self._get_current_source_array(fallback=channel_data)
        self.arr = np.asarray(src)
        self.arr3 = self.arr[np.newaxis, ...] if self.arr.ndim == 2 else self.arr
        self.nz = int(self.arr3.shape[0]) if self.arr3.ndim == 3 else 1

        self.z = self._get_parent_current_z()
        self.z = max(0, min(self.z, max(self.nz - 1, 0)))

        # thresholding state (drives mask computation)
        self.blur_enabled = False
        self.blur_sigma = 1.0
        self.threshold = 500

        # full 3D mask storage; displayed mask is a 2D slice
        self.preview_masks = np.zeros(self.arr3.shape, dtype=np.uint8)

        self.setWindowTitle(f"Mask Tuner — Channel {self.channel_index + 1}")
        self.setMinimumSize(950, 700)
        self.setAttribute(Qt.WA_DeleteOnClose, True)

        outer = QVBoxLayout(self)

        # ---- embedded viewer + right panel ----
        self.viewer = napari.Viewer()

        split = QSplitter(Qt.Horizontal)
        outer.addWidget(split)

        # Left: napari viewer widget
        left = QWidget()
        left_lay = QVBoxLayout(left)
        left_lay.setContentsMargins(0, 0, 0, 0)
        left_lay.addWidget(self.viewer.window._qt_viewer)
        split.addWidget(left)

        # Create layers BEFORE creating controls that reference them
        self.img_layer = self.viewer.add_image(
            self.arr3[self.z],
            name=f"Channel {self.channel_index + 1}",
            blending="additive",
            opacity=1.0,
        )
        self.mask_layer = self.viewer.add_image(
            self.preview_masks[self.z],
            name="Mask (preview)",
            colormap="yellow",
            blending="additive",
            opacity=0.6,
            contrast_limits=(0, 1),
        )

        # Right: controls panel
        right = QWidget()
        right_lay = QVBoxLayout(right)
        right_lay.setContentsMargins(6, 6, 6, 6)

        self.layer_ctrl_widgets = {}

        gb1, ctrl1 = self._layer_controls_group(
            title=f"Channel {self.channel_index + 1} Controls",
            layer=self.img_layer,
            is_mask=False,
        )
        self.layer_ctrl_widgets["channel"] = ctrl1
        right_lay.addWidget(gb1)

        gb2, ctrl2 = self._layer_controls_group(
            title="Mask Controls",
            layer=self.mask_layer,
            is_mask=True,
        )
        self.layer_ctrl_widgets["mask"] = ctrl2
        right_lay.addWidget(gb2)

        right_lay.addStretch(1)
        split.addWidget(right)
        split.setSizes([700, 250])

        # ---- bottom control row ----
        ctrl = QHBoxLayout()

        ctrl.addWidget(QLabel("Z"))
        self.z_slider = QSlider(Qt.Horizontal)
        self.z_slider.setRange(0, max(self.nz - 1, 0))
        self.z_slider.setValue(self.z)
        ctrl.addWidget(self.z_slider)

        self.z_label = QLabel(f"{self.z}/{max(self.nz - 1, 0)}")
        self.z_label.setFixedWidth(90)
        ctrl.addWidget(self.z_label)

        self.otsu_btn = QPushButton("Otsu (this slice)")
        ctrl.addWidget(self.otsu_btn)

        ctrl.addWidget(QLabel("Threshold"))
        self.th_slider = QSlider(Qt.Horizontal)
        self._update_threshold_slider_range()
        self.th_slider.setValue(int(max(0, min(self.threshold, self.th_slider.maximum()))))
        ctrl.addWidget(self.th_slider)

        self.th_val = QLabel(str(int(self.threshold)))
        self.th_val.setFixedWidth(70)
        ctrl.addWidget(self.th_val)

        self.blur_cb = QCheckBox("Blur")
        ctrl.addWidget(self.blur_cb)

        ctrl.addWidget(QLabel("σ"))
        self.sigma_slider = QSlider(Qt.Horizontal)
        self.sigma_slider.setRange(0, 50)  # 0..5.0
        self.sigma_slider.setValue(int(round(self.blur_sigma * 10)))
        ctrl.addWidget(self.sigma_slider)

        self.sigma_val = QLabel(f"{self.blur_sigma:.1f}")
        self.sigma_val.setFixedWidth(60)
        ctrl.addWidget(self.sigma_val)

        self.preview_all_btn = QPushButton("Preview full stack")
        ctrl.addWidget(self.preview_all_btn)

        self.accept_btn = QPushButton("Accept (write to main)")
        ctrl.addWidget(self.accept_btn)

        outer.addLayout(ctrl)

        # ---- meta row ----
        meta_row = QHBoxLayout()
        self.save_meta_btn = QPushButton("Save Channel Mask Meta")
        self.load_meta_btn = QPushButton("Load Channel Mask Meta")
        meta_row.addWidget(self.save_meta_btn)
        meta_row.addWidget(self.load_meta_btn)
        meta_row.addStretch(1)
        outer.addLayout(meta_row)

        # ---- signals ----
        self.z_slider.valueChanged.connect(self._on_z_changed)
        self.th_slider.valueChanged.connect(self._on_threshold_changed)
        self.blur_cb.stateChanged.connect(self._on_blur_toggle)
        self.sigma_slider.valueChanged.connect(self._on_sigma_changed)

        self.otsu_btn.clicked.connect(self._otsu_current_slice)
        self.preview_all_btn.clicked.connect(self._preview_full_stack)
        self.accept_btn.clicked.connect(self._accept_mask)

        self.save_meta_btn.clicked.connect(self._save_channel_meta)
        self.load_meta_btn.clicked.connect(self._load_channel_meta)

        # Sync with parent viewer Z if available
        try:
            if hasattr(self.parent_widget, "viewer") and hasattr(self.parent_widget.viewer, "dims"):
                self.parent_widget.viewer.dims.events.current_step.connect(self._on_parent_dims_changed)
        except Exception:
            pass

        # ---- initialize preview from latest parent data ----
        self._refresh_source_from_parent()
        self._preview_current_slice()

    # ---------------- parent/source sync ----------------
    def _get_parent_current_z(self):
        try:
            if hasattr(self.parent_widget, "current_z"):
                return int(getattr(self.parent_widget, "current_z", 0))
        except Exception:
            pass

        try:
            if hasattr(self.parent_widget, "viewer") and hasattr(self.parent_widget.viewer, "dims"):
                return int(self.parent_widget.viewer.dims.current_step[0])
        except Exception:
            pass

        return 0

    def _get_current_source_array(self, fallback=None):
        candidates = []

        # Prefer current data from the parent's live napari layer
        try:
            layers = getattr(self.parent_widget, "channel_layers", None)
            if layers is not None and self.channel_index < len(layers):
                lyr = layers[self.channel_index]
                arr = np.asarray(lyr.data)
                if arr.size > 0:
                    candidates.append(arr)
        except Exception:
            pass

        # Fall back to parent widget's channel list storage
        try:
            ch_list = getattr(self.parent_widget, "channel_list", None)
            if ch_list is not None and self.channel_index < len(ch_list):
                arr = np.asarray(ch_list[self.channel_index])
                if arr.size > 0:
                    candidates.append(arr)
        except Exception:
            pass

        # Finally use constructor data
        try:
            if fallback is not None:
                arr = np.asarray(fallback)
                if arr.size > 0:
                    candidates.append(arr)
        except Exception:
            pass

        if not candidates:
            return np.zeros((1, 1), dtype=np.float32)

        # Prefer a 3D stack if one exists, otherwise first valid candidate
        for arr in candidates:
            if arr.ndim >= 3:
                return arr
        return candidates[0]

    def _update_threshold_slider_range(self):
        try:
            th_max = int(np.nanmax(self.arr3)) if self.arr3.size else 1
        except Exception:
            th_max = 1
        th_max = max(int(th_max), 1)

        if hasattr(self, "th_slider") and self.th_slider is not None:
            cur = int(getattr(self, "threshold", 0))
            self.th_slider.blockSignals(True)
            self.th_slider.setRange(0, th_max)
            self.th_slider.setValue(max(0, min(cur, th_max)))
            self.th_slider.blockSignals(False)

        self.threshold = max(0, min(int(getattr(self, "threshold", 0)), th_max))
        if hasattr(self, "th_val") and self.th_val is not None:
            self.th_val.setText(str(int(self.threshold)))

    def _refresh_source_from_parent(self):
        old_shape = tuple(self.arr3.shape) if hasattr(self, "arr3") else None

        src = self._get_current_source_array(fallback=getattr(self, "arr3", None))
        self.arr = np.asarray(src)
        self.arr3 = self.arr[np.newaxis, ...] if self.arr.ndim == 2 else self.arr

        if self.arr3.ndim != 3:
            self.arr3 = np.asarray(self.arr3).reshape((1,) + tuple(np.asarray(self.arr3).shape[-2:]))

        self.nz = int(self.arr3.shape[0]) if self.arr3.ndim == 3 else 1

        parent_z = self._get_parent_current_z()
        self.z = max(0, min(int(parent_z), max(self.nz - 1, 0)))

        new_shape = tuple(self.arr3.shape)

        # If cropping changed shape, reset preview masks to match new data
        if not hasattr(self, "preview_masks") or old_shape != new_shape:
            self.preview_masks = np.zeros(new_shape, dtype=np.uint8)

        # Keep dialog widgets aligned with current source
        if hasattr(self, "z_slider") and self.z_slider is not None:
            self.z_slider.blockSignals(True)
            self.z_slider.setRange(0, max(self.nz - 1, 0))
            self.z_slider.setValue(self.z)
            self.z_slider.blockSignals(False)

        if hasattr(self, "z_label") and self.z_label is not None:
            self.z_label.setText(f"{self.z}/{max(self.nz - 1, 0)}")

        self._update_threshold_slider_range()

        # Refresh displayed image slice
        if hasattr(self, "img_layer") and self.img_layer is not None:
            try:
                self.img_layer.data = self.arr3[self.z]
            except Exception:
                pass

        if hasattr(self, "mask_layer") and self.mask_layer is not None:
            try:
                self.mask_layer.data = self.preview_masks[self.z]
                self.mask_layer.contrast_limits = (0, 1)
            except Exception:
                pass

        try:
            self.viewer.reset_view()
        except Exception:
            pass

    def _push_z_to_parent(self, z):
        try:
            if hasattr(self.parent_widget, "z_changed"):
                self.parent_widget.z_changed(int(z))
                return
        except Exception:
            pass

        try:
            if hasattr(self.parent_widget, "viewer") and hasattr(self.parent_widget.viewer, "dims"):
                self.parent_widget.viewer.dims.set_current_step(0, int(z))
        except Exception:
            pass

    def _on_parent_dims_changed(self, event=None):
        if self._syncing_parent_z:
            return

        self._syncing_parent_z = True
        try:
            self._refresh_source_from_parent()

            z = self._get_parent_current_z()
            z = max(0, min(int(z), max(self.nz - 1, 0)))
            self.z = z

            self.z_slider.blockSignals(True)
            self.z_slider.setRange(0, max(self.nz - 1, 0))
            self.z_slider.setValue(self.z)
            self.z_slider.blockSignals(False)

            self.z_label.setText(f"{self.z}/{max(self.nz - 1, 0)}")
            self._preview_current_slice()
        finally:
            self._syncing_parent_z = False

    # ---------------- helpers ----------------
    def _set_slider_from_lineedit(self, lineedit, slider, lo, hi):
        try:
            val = int(lineedit.text())
        except Exception:
            return
        val = max(int(lo), min(int(hi), val))
        slider.setValue(val)

    def _safe_get_colormap_str(self, layer):
        """Return something JSON-serializable for colormap."""
        try:
            cm = getattr(layer, "colormap", None)
            if cm is None:
                return None
            name = getattr(cm, "name", None)
            if isinstance(name, str) and name:
                return name
            if isinstance(cm, str) and cm:
                return cm
            return str(cm)
        except Exception:
            return None

    def _apply_layer_settings(self, layer, settings):
        """Apply layer settings dict safely."""
        if not isinstance(settings, dict):
            return
        try:
            if "visible" in settings:
                layer.visible = bool(settings["visible"])
        except Exception:
            pass
        try:
            if "opacity" in settings:
                layer.opacity = float(settings["opacity"])
        except Exception:
            pass
        try:
            if "contrast_limits" in settings and settings["contrast_limits"] is not None:
                a, b = settings["contrast_limits"]
                layer.contrast_limits = (float(a), float(b))
        except Exception:
            pass
        try:
            if "colormap" in settings and settings["colormap"]:
                layer.colormap = settings["colormap"]
        except Exception:
            pass

    # ---------------- UI handlers ----------------
    def _on_z_changed(self, v):
        if self._syncing_local_z:
            return

        self._syncing_local_z = True
        try:
            self._refresh_source_from_parent()

            self.z = int(v)
            nz = int(getattr(self, "nz", 0) or 0)
            if nz <= 0:
                return

            self.z = max(0, min(self.z, nz - 1))
            self.z_label.setText(f"{self.z}/{max(nz - 1, 0)}")

            # Update local 2D display
            self.img_layer.data = self.arr3[self.z]

            if hasattr(self, "preview_masks"):
                m = np.asarray(self.preview_masks)
                if m.ndim == 3 and m.shape[0] == nz:
                    self.mask_layer.data = m[self.z]
                    try:
                        self.mask_layer.contrast_limits = (0, 1)
                    except Exception:
                        pass

            # Sync main UI / parent viewer z
            if not self._syncing_parent_z:
                self._push_z_to_parent(self.z)

            # Optional dynamically loaded channels
            if hasattr(self, "channel_full_stacks") and hasattr(self, "channel_layers"):
                for stack, lay in zip(self.channel_full_stacks, self.channel_layers):
                    a = np.asarray(stack)
                    if a.ndim == 3:
                        z = max(0, min(self.z, a.shape[0] - 1))
                        lay.data = a[z]
                    else:
                        lay.data = a

            self._preview_current_slice()

        finally:
            self._syncing_local_z = False

    def _dtype_max_value(self, dtype):
        dt = np.dtype(dtype)
        if np.issubdtype(dt, np.integer):
            return int(np.iinfo(dt).max)
        if np.issubdtype(dt, np.floating):
            return 1.0
        return 65535

    def _on_threshold_changed(self, v):
        self._refresh_source_from_parent()
        self.threshold = int(v)
        self.th_val.setText(str(int(self.threshold)))
        self._preview_current_slice()

    def _on_blur_toggle(self, s):
        self._refresh_source_from_parent()
        self.blur_enabled = bool(s)
        self._preview_current_slice()

    def _on_sigma_changed(self, v):
        self._refresh_source_from_parent()
        self.blur_sigma = float(v) / 10.0
        self.sigma_val.setText(f"{self.blur_sigma:.1f}")
        if self.blur_enabled:
            self._preview_current_slice()

    def _otsu_current_slice(self):
        self._refresh_source_from_parent()
        sl = self.arr3[self.z]
        try:
            th = float(threshold_otsu(sl))
        except Exception:
            th = float(np.mean(sl))
        self.threshold = int(th)

        self.th_slider.blockSignals(True)
        try:
            self.th_slider.setValue(max(0, min(self.threshold, self.th_slider.maximum())))
        finally:
            self.th_slider.blockSignals(False)

        self.th_val.setText(str(int(self.threshold)))
        self._preview_current_slice()

    # ---------------- mask preview ----------------
    def _preview_current_slice(self):
        self._refresh_source_from_parent()

        m = compute_mask(
            self.arr3[self.z],
            threshold=self.threshold,
            blur_enabled=self.blur_enabled,
            blur_sigma=self.blur_sigma,
            min_size=0,
            do_close=False,
        )
        self.preview_masks[self.z] = m

        try:
            self.img_layer.data = self.arr3[self.z]
        except Exception:
            pass

        self.mask_layer.data = self.preview_masks[self.z]
        try:
            self.mask_layer.contrast_limits = (0, 1)
        except Exception:
            pass

    def _preview_full_stack(self):
        self._refresh_source_from_parent()

        for z in range(self.nz):
            self.preview_masks[z] = compute_mask(
                self.arr3[z],
                threshold=self.threshold,
                blur_enabled=self.blur_enabled,
                blur_sigma=self.blur_sigma,
                min_size=0,
                do_close=False,
            )

        self.img_layer.data = self.arr3[self.z]
        self.mask_layer.data = self.preview_masks[self.z]
        try:
            self.mask_layer.contrast_limits = (0, 1)
        except Exception:
            pass

    # ---------------- accept into main ----------------
    def _accept_mask(self):
        try:
            self._refresh_source_from_parent()

            # Always build a complete 3D mask using the *current* parameters
            self._preview_full_stack()

            masks_out = (self.preview_masks > 0).astype(np.uint8)

            self.parent_widget.create_mask_layer(self.channel_index, masks_out)
            try:
                self.parent_widget.z_changed(int(getattr(self.parent_widget, "current_z", self.z)))
            except Exception:
                pass

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to write mask to main UI: {e}")
            return

        self.accept()

    # ---------------- save/load per-channel meta ----------------
    def _meta_path_default(self):
        meta_dir = Path("meta_data")
        meta_dir.mkdir(exist_ok=True)
        stem = Path(self.file_path).stem
        return meta_dir / f"{stem}.channel{self.channel_index + 1}.maskmeta.json"

    def _collect_ui_state(self):
        md = {
            "type": "mask_tuner_channel_meta_v2",
            "source_file": str(self.file_path),
            "timestamp": datetime.now().isoformat(),
            "channel_index_0based": int(self.channel_index),
            "z": int(self.z_slider.value()),
            "threshold": int(self.th_slider.value()),
            "blur_enabled": bool(self.blur_cb.isChecked()),
            "blur_sigma": float(self.sigma_slider.value()) / 10.0,
            "layers": {
                "channel": {
                    "visible": bool(getattr(self.img_layer, "visible", True)),
                    "opacity": float(getattr(self.img_layer, "opacity", 1.0)),
                    "contrast_limits": list(getattr(self.img_layer, "contrast_limits", (0, 1))),
                    "colormap": self._safe_get_colormap_str(self.img_layer),
                    "ui": self.layer_ctrl_widgets.get("channel", {}).get("ui_state", {}),
                },
                "mask": {
                    "visible": bool(getattr(self.mask_layer, "visible", True)),
                    "opacity": float(getattr(self.mask_layer, "opacity", 0.6)),
                    "contrast_limits": list(getattr(self.mask_layer, "contrast_limits", (0, 1))),
                    "colormap": self._safe_get_colormap_str(self.mask_layer),
                    "ui": self.layer_ctrl_widgets.get("mask", {}).get("ui_state", {}),
                }
            }
        }
        return md

    def _save_channel_meta(self):
        try:
            path = self._meta_path_default()

            for key in ("channel", "mask"):
                ctrl = self.layer_ctrl_widgets.get(key, {})
                if not ctrl:
                    continue
                ctrl["ui_state"] = {
                    "visible": bool(ctrl["visible_cb"].isChecked()),
                    "opacity_slider": int(ctrl["opacity_slider"].value()),
                    "min_slider": int(ctrl["min_slider"].value()),
                    "max_slider": int(ctrl["max_slider"].value()),
                }

            md = self._collect_ui_state()

            with open(path, "w") as f:
                json.dump(md, f, indent=4)
            QMessageBox.information(self, "Saved", f"Saved:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save channel meta: {e}")

    def _load_channel_meta(self):
        try:
            self._refresh_source_from_parent()

            path = self._meta_path_default()
            if not path.exists():
                QMessageBox.warning(self, "Not found", f"No meta file found:\n{path}")
                return

            with open(path, "r") as f:
                md = json.load(f)

            self.z_slider.blockSignals(True)
            self.th_slider.blockSignals(True)
            self.blur_cb.blockSignals(True)
            self.sigma_slider.blockSignals(True)
            try:
                z = int(md.get("z", self.z))
                z = max(0, min(z, self.nz - 1))
                th = int(md.get("threshold", self.threshold))
                be = bool(md.get("blur_enabled", self.blur_enabled))
                bs = float(md.get("blur_sigma", self.blur_sigma))

                self.z_slider.setValue(z)
                self.th_slider.setValue(max(0, min(th, self.th_slider.maximum())))
                self.blur_cb.setChecked(be)
                self.sigma_slider.setValue(int(round(bs * 10)))
            finally:
                self.z_slider.blockSignals(False)
                self.th_slider.blockSignals(False)
                self.blur_cb.blockSignals(False)
                self.sigma_slider.blockSignals(False)

            self.z = int(self.z_slider.value())
            self.threshold = int(self.th_slider.value())
            self.blur_enabled = bool(self.blur_cb.isChecked())
            self.blur_sigma = float(self.sigma_slider.value()) / 10.0
            self.z_label.setText(f"{self.z}/{max(self.nz - 1, 0)}")
            self.th_val.setText(str(int(self.threshold)))
            self.sigma_val.setText(f"{self.blur_sigma:.1f}")

            layers_md = md.get("layers", {})
            self._apply_layer_settings(self.img_layer, layers_md.get("channel", {}))
            self._apply_layer_settings(self.mask_layer, layers_md.get("mask", {}))

            def _apply_ctrl(key):
                ctrl = self.layer_ctrl_widgets.get(key, {})
                ui = (layers_md.get(key, {}) or {}).get("ui", {}) or (layers_md.get(key, {}) or {}).get("ui_state", {})
                if not ctrl or not ui:
                    return

                for wname, setter in [
                    ("visible_cb", "setChecked"),
                    ("opacity_slider", "setValue"),
                    ("min_slider", "setValue"),
                    ("max_slider", "setValue"),
                ]:
                    if wname not in ctrl:
                        continue
                    widget = ctrl[wname]
                    widget.blockSignals(True)
                    try:
                        if wname == "visible_cb" and "visible" in ui:
                            getattr(widget, setter)(bool(ui["visible"]))
                        elif wname == "opacity_slider" and "opacity_slider" in ui:
                            getattr(widget, setter)(int(ui["opacity_slider"]))
                        elif wname == "min_slider" and "min_slider" in ui:
                            getattr(widget, setter)(int(ui["min_slider"]))
                        elif wname == "max_slider" and "max_slider" in ui:
                            getattr(widget, setter)(int(ui["max_slider"]))
                    finally:
                        widget.blockSignals(False)

                try:
                    ctrl["apply_fn"]()
                except Exception:
                    pass

            _apply_ctrl("channel")
            _apply_ctrl("mask")

            self._refresh_source_from_parent()
            self.img_layer.data = self.arr3[self.z]
            self.mask_layer.data = self.preview_masks[self.z]
            self._preview_current_slice()

            QMessageBox.information(self, "Loaded", f"Loaded:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load channel meta: {e}")

    # ---------------- right-side layer controls ----------------
    def _layer_controls_group(self, title, layer, is_mask=False):
        """
        Returns (groupbox, ctrl_dict) where ctrl_dict stores widget handles
        and an apply_fn() that writes widget values to the napari layer.
        """
        gb = QGroupBox(title)
        vlay = QVBoxLayout(gb)

        top = QHBoxLayout()
        vis_cb = QCheckBox("Visible")
        vis_cb.setChecked(True)
        top.addWidget(vis_cb)

        top.addWidget(QLabel("Opacity"))
        opacity_slider = QSlider(Qt.Horizontal)
        opacity_slider.setRange(0, 100)
        opacity_slider.setValue(int(getattr(layer, "opacity", 1.0) * 100))
        top.addWidget(opacity_slider)

        color_btn = QPushButton("Color")
        top.addWidget(color_btn)
        vlay.addLayout(top)

        if is_mask:
            lo, hi = 0, 1
            default_min, default_max = 0, 1
        else:
            try:
                sl = np.asarray(layer.data)
                default_min = int(np.nanmin(sl))
                default_max = int(np.nanmax(sl))
                if default_max <= default_min:
                    default_max = default_min + 1
                lo, hi = 0, max(default_max, 1)
            except Exception:
                lo, hi = 0, 65535
                default_min, default_max = 0, 65535

        def _minmax_row(label, init):
            row = QHBoxLayout()
            row.addWidget(QLabel(label))
            sld = QSlider(Qt.Horizontal)
            sld.setRange(int(lo), int(hi))
            sld.setValue(int(init))
            row.addWidget(sld)
            le = QLineEdit(str(int(init)))
            le.setFixedWidth(80)
            le.setValidator(QIntValidator(int(lo), int(hi)))
            le.editingFinished.connect(
                lambda le=le, sld=sld: self._set_slider_from_lineedit(le, sld, int(lo), int(hi))
            )
            sld.valueChanged.connect(lambda v, le=le: le.setText(str(int(v))))
            row.addWidget(le)
            vlay.addLayout(row)
            return sld

        min_slider = _minmax_row("Min:", default_min)
        max_slider = _minmax_row("Max:", default_max)

        def apply_fn():
            try:
                layer.visible = bool(vis_cb.isChecked())
            except Exception:
                pass
            try:
                layer.opacity = float(opacity_slider.value()) / 100.0
            except Exception:
                pass
            try:
                mn = int(min_slider.value())
                mx = int(max_slider.value())
                if mx <= mn:
                    mx = mn + 1
                layer.contrast_limits = (mn, mx)
            except Exception:
                pass
            if is_mask:
                try:
                    layer.contrast_limits = (0, 1)
                except Exception:
                    pass

        vis_cb.stateChanged.connect(lambda _=None: apply_fn())
        opacity_slider.valueChanged.connect(lambda _=None: apply_fn())
        min_slider.valueChanged.connect(lambda _=None: apply_fn())
        max_slider.valueChanged.connect(lambda _=None: apply_fn())

        def pick_color():
            col = QColorDialog.getColor(parent=self)
            if not col.isValid():
                return
            try:
                layer.colormap = col.name()
            except Exception:
                pass

        color_btn.clicked.connect(pick_color)

        apply_fn()

        ctrl = {
            "visible_cb": vis_cb,
            "opacity_slider": opacity_slider,
            "min_slider": min_slider,
            "max_slider": max_slider,
            "apply_fn": apply_fn,
            "ui_state": {},
        }
        return gb, ctrl

    def closeEvent(self, event):
        try:
            if hasattr(self.parent_widget, "viewer") and hasattr(self.parent_widget.viewer, "dims"):
                self.parent_widget.viewer.dims.events.current_step.disconnect(self._on_parent_dims_changed)
        except Exception:
            pass
        super().closeEvent(event)

###############################################################################

# Building of the main UI
class UIWidget(QWidget):
    def __init__(self, viewer, channel_list, channel_layers, contrast_limits, file_path):
        super().__init__()
        self.viewer = viewer
        self.channel_list = channel_list
        self.channel_layers = channel_layers
        self.contrast_limits = contrast_limits
        self.file_path = file_path

        # determine Z size from first channel
        first = np.asarray(channel_list[0])
        self.n_z = first.shape[0] if first.ndim == 3 else 1
        self.current_z = 0
        self._syncing_z = False

        self.setLayout(QVBoxLayout())
        self.build_ui()
        self.setAttribute(Qt.WA_DeleteOnClose, True)

        self.coloc_controls = getattr(self, "coloc_controls", [])
        self.coloc_control_boxes = getattr(self, "coloc_control_boxes", [])
        self.layer_control_widgets = getattr(self, "layer_control_widgets", {})
        self.mask_layers = getattr(self, "mask_layers", {})
        self.derived_layers = getattr(self, "derived_layers", {})
        self.masks_by_channel = getattr(self, "masks_by_channel", {})
        self.derived_stacks = getattr(self, "derived_stacks", {})

        arr = np.asarray(channel_list[0])
        print("dtype:", arr.dtype, "shape:", arr.shape, "min/max:", float(np.min(arr)), float(np.max(arr)))

        # Sync custom slider from napari dims
        try:
            self.viewer.dims.events.current_step.connect(self._sync_from_napari_dims)
        except Exception:
            pass

        # Initialize napari dims to a valid z if possible
        try:
            self._set_napari_z(self.current_z)
        except Exception:
            pass

        # Initialize slider from napari state
        try:
            self._sync_from_napari_dims()
        except Exception:
            pass

    def _get_z_axis(self):
        try:
            first = np.asarray(self.channel_list[0])
            if first.ndim >= 3:
                return 0
        except Exception:
            pass
        return 0

    def _set_napari_z(self, z):
        z = int(max(0, min(self.n_z - 1, int(z))))
        z_axis = self._get_z_axis()

        try:
            self.viewer.dims.set_current_step(z_axis, z)
        except Exception:
            try:
                steps = list(self.viewer.dims.current_step)
                if z_axis < len(steps):
                    steps[z_axis] = z
                    self.viewer.dims.current_step = tuple(steps)
            except Exception:
                pass

    def _sync_from_napari_dims(self, event=None):
        if self._syncing_z:
            return

        self._syncing_z = True
        try:
            z_axis = self._get_z_axis()

            try:
                z = int(self.viewer.dims.current_step[z_axis])
            except Exception:
                z = self.current_z

            z = int(max(0, min(self.n_z - 1, z)))
            self.current_z = z

            if hasattr(self, "z_slider") and self.z_slider is not None:
                self.z_slider.blockSignals(True)
                self.z_slider.setRange(0, max(self.n_z - 1, 0))
                self.z_slider.setValue(z)
                self.z_slider.blockSignals(False)

            self._update_aux_layers_for_z(z)

        finally:
            self._syncing_z = False

    def z_changed(self, v):
        """
        Keep this function name unchanged because other code may call it.
        Now it drives napari's dims Z position instead of replacing 3D layer data.
        """
        if self._syncing_z:
            return

        self._syncing_z = True
        try:
            v = int(max(0, min(self.n_z - 1, int(v))))
            self.current_z = v

            # Update napari's actual Z dimension
            self._set_napari_z(v)

            # Keep any UI-managed 2D auxiliary layers in sync
            self._update_aux_layers_for_z(v)

        finally:
            self._syncing_z = False

    def _update_aux_layers_for_z(self, v):
        v = int(v)

        # ---------------------------------------------------------
        # DO NOT reslice the main channel image layers here.
        # Leave them as full 3D stacks so napari controls Z.
        # ---------------------------------------------------------

        # ---- masks ----
        if hasattr(self, "masks_by_channel"):
            for ch_idx, m3 in list(self.masks_by_channel.items()):
                if m3 is None:
                    continue

                m3 = np.asarray(m3)
                if m3.ndim == 2:
                    m2 = (m3 > 0).astype(np.uint8)
                else:
                    # ---- FIX: skip stale masks whose Z count doesn't match current n_z ----
                    if m3.shape[0] != self.n_z:
                        continue
                    z = max(0, min(v, m3.shape[0] - 1))
                    m2 = (m3[z] > 0).astype(np.uint8)

                lay = None
                if hasattr(self, "mask_layers"):
                    lay = self.mask_layers.get(ch_idx, None)

                if lay is None:
                    name = f"Channel {ch_idx+1} mask"
                    try:
                        lay = self.viewer.layers[name]
                    except Exception:
                        lay = None

                if lay is not None:
                    try:
                        lay.data = m2
                        lay.contrast_limits = (0, 1)
                        # ---- FIX: keep mask scale/translate in sync with channel layer ----
                        ch_layers = getattr(self, "channel_layers", [])
                        if ch_idx < len(ch_layers):
                            ref = ch_layers[ch_idx]
                            raw_s = getattr(ref, "scale", None)
                            raw_t = getattr(ref, "translate", None)
                            if raw_s is not None:
                                s = tuple(float(x) for x in raw_s)
                                yx_s = s[-2:] if len(s) >= 2 else (1.0, 1.0)
                                try:
                                    lay.scale = yx_s
                                except Exception:
                                    pass
                            if raw_t is not None:
                                t = tuple(float(x) for x in raw_t)
                                yx_t = t[-2:] if len(t) >= 2 else (0.0, 0.0)
                                try:
                                    lay.translate = yx_t
                                except Exception:
                                    pass
                    except Exception:
                        pass

        # ---- derived (coloc/multiply) ----
        if hasattr(self, "derived_stacks"):
            self.derived_layers = getattr(self, "derived_layers", {})
            for k, arr3 in list(self.derived_stacks.items()):
                lay = self.derived_layers.get(k, None)
                if lay is None:
                    continue

                arr3 = np.asarray(arr3)
                try:
                    if arr3.ndim == 3:
                        # ---- FIX: guard against stale derived stack with wrong n_z ----
                        if arr3.shape[0] != self.n_z:
                            continue
                        z = max(0, min(v, arr3.shape[0] - 1))
                        lay.data = arr3[z]
                    else:
                        lay.data = arr3
                except Exception:
                    pass

        # prune deleted coloc controls
        if hasattr(self, "coloc_controls"):
            alive = []
            alive_boxes = []
            for ctrl, box in zip(self.coloc_controls, getattr(self, "coloc_control_boxes", [])):
                lay = ctrl.get("layer", None)
                if lay is not None and lay in self.viewer.layers:
                    alive.append(ctrl)
                    alive_boxes.append(box)
                else:
                    try:
                        box.setParent(None)
                    except Exception:
                        pass
            self.coloc_controls = alive
            self.coloc_control_boxes = alive_boxes

    def build_ui(self):
        layout = self.layout()

        # Z slider
        z_box = QGroupBox("Z Slice Navigation")
        z_layout = QVBoxLayout()
        self.z_slider = QSlider(Qt.Horizontal)
        self.z_slider.setRange(0, max(self.n_z - 1, 0))
        self.z_slider.setValue(self.current_z)
        self.z_slider.valueChanged.connect(self.z_changed)
        z_layout.addWidget(self.z_slider)
        z_box.setLayout(z_layout)
        layout.addWidget(z_box)

        def infer_max_range_from_channels(channel_list, fallback=65535):
            try:
                best_dtype = None
                best_p99 = 0.0
                best_max = 0.0

                for ch in channel_list:
                    arr = np.asarray(ch)
                    if best_dtype is None:
                        best_dtype = arr.dtype

                    if arr.ndim == 3 and arr.shape[0] > 1:
                        z_idx = [0, arr.shape[0] // 2, arr.shape[0] - 1]
                        sample = np.asarray(arr[z_idx, ...]).ravel()
                    else:
                        sample = arr.ravel()

                    if np.issubdtype(sample.dtype, np.floating):
                        sample = sample[np.isfinite(sample)]

                    if sample.size == 0:
                        continue

                    best_max = max(best_max, float(np.max(sample)))
                    best_p99 = max(best_p99, float(np.percentile(sample, 99)))

                if best_dtype is not None and np.issubdtype(best_dtype, np.integer):
                    return int(min(np.iinfo(best_dtype).max, fallback))

                v = best_p99 if best_p99 > 0 else best_max
                if v <= 0:
                    return fallback

                if v <= 255:
                    return 255
                if v <= 4095:
                    return 4095
                if v <= 65535:
                    return 65535
                return int(min(v, fallback))

            except Exception:
                return fallback

        max_range = infer_max_range_from_channels(self.channel_list, fallback=65535)

        # Per-channel controls
        for idx, layer in enumerate(self.channel_layers, start=1):
            gb = QGroupBox(f"Channel {idx} Controls")
            vlay = QVBoxLayout()

            # Visibility + opacity
            top_h = QHBoxLayout()
            vis_cb = QCheckBox("Visible")
            vis_cb.setChecked(True)
            vis_cb.stateChanged.connect(lambda s, lyr=layer: setattr(lyr, "visible", bool(s)))
            top_h.addWidget(vis_cb)

            top_h.addWidget(QLabel("Opacity"))
            op = QSlider(Qt.Horizontal)
            op.setRange(0, 100)
            op.setValue(int(layer.opacity * 100))
            op.valueChanged.connect(lambda v, lyr=layer: setattr(lyr, "opacity", v / 100.0))
            top_h.addWidget(op)

            # Color picker
            color_btn = QPushButton("Color")

            def make_color_picker(lyr):
                def pick():
                    col = QColorDialog.getColor(parent=self)
                    if col.isValid():
                        try:
                            lyr.colormap = col.name()
                        except Exception:
                            pass
                return pick

            color_btn.clicked.connect(make_color_picker(layer))
            top_h.addWidget(color_btn)
            vlay.addLayout(top_h)

            # Min slider
            ch_min_layout = QHBoxLayout()
            ch_min_layout.addWidget(QLabel("Min:"))
            ch_min = QSlider(Qt.Horizontal)
            ch_min.setRange(0, max_range)

            try:
                lim = self.contrast_limits.get(f"ch{idx}", None)
            except Exception:
                lim = None

            if lim is None:
                try:
                    lim = tuple(layer.contrast_limits)
                except Exception:
                    lim = (0, max_range)

            ch_min.setValue(int(lim[0]))
            ch_min_layout.addWidget(ch_min)

            ch_min_val = QLineEdit(str(int(lim[0])))
            ch_min_val.setFixedWidth(80)
            ch_min_val.setValidator(QIntValidator(0, max_range))
            ch_min_val.editingFinished.connect(
                lambda le=ch_min_val, sl=ch_min: self._set_slider_from_lineedit(le, sl, 0, max_range)
            )
            ch_min_layout.addWidget(ch_min_val)
            vlay.addLayout(ch_min_layout)

            # Max slider
            ch_max_layout = QHBoxLayout()
            ch_max_layout.addWidget(QLabel("Max:"))
            ch_max = QSlider(Qt.Horizontal)
            ch_max.setRange(0, max_range)
            ch_max.setValue(int(lim[1]))
            ch_max_layout.addWidget(ch_max)

            ch_max_val = QLineEdit(str(int(lim[1])))
            ch_max_val.setFixedWidth(80)
            ch_max_val.setValidator(QIntValidator(0, max_range))
            ch_max_val.editingFinished.connect(
                lambda le=ch_max_val, sl=ch_max: self._set_slider_from_lineedit(le, sl, 0, max_range)
            )
            ch_max_layout.addWidget(ch_max_val)
            vlay.addLayout(ch_max_layout)

            def _apply_contrast(_=None, lyr=layer, smin=ch_min, smax=ch_max, le_min=ch_min_val, le_max=ch_max_val, i=idx):
                lo = int(smin.value())
                hi = int(smax.value())
                if hi < lo:
                    hi = lo
                    smax.blockSignals(True)
                    smax.setValue(hi)
                    smax.blockSignals(False)

                le_min.setText(str(lo))
                le_max.setText(str(hi))

                try:
                    lyr.contrast_limits = (lo, hi)
                except Exception:
                    pass

                try:
                    self.contrast_limits[f"ch{i}"] = (lo, hi)
                except Exception:
                    pass

            ch_min.valueChanged.connect(_apply_contrast)
            ch_max.valueChanged.connect(_apply_contrast)

            ch_min_val.setText(str(int(ch_min.value())))
            ch_max_val.setText(str(int(ch_max.value())))

            gb.setLayout(vlay)
            layout.addWidget(gb)

        # Container for dynamic mask controls
        self.mask_container = QVBoxLayout()
        layout.addLayout(self.mask_container)

        # Analysis buttons
        btns = QHBoxLayout()

        analysis_btn = QPushButton("Data Analysis")
        analysis_btn.clicked.connect(self.open_analysis)
        btns.addWidget(analysis_btn)

        preprocess_btn = QPushButton("Pre-process Analysis")
        preprocess_btn.clicked.connect(self.open_preprocess)
        btns.addWidget(preprocess_btn)

        layout.addLayout(btns)

    def _suggest_export_dtype_and_scale(self, arr):
        """
        Export rules (no forced uint16 fallback):
        - Bool masks -> uint8 with values {0,1} (binary).
        - Float masks in [0,1] -> uint8 with values {0,1} (binary), preserving mask semantics.
        - Integer images -> preserve dtype exactly (uint8/uint16/int16/etc).
        - Float intensity (not mask-like) -> keep float32 (no scaling).

        Returns: (out_arr, out_dtype, suggested_contrast_limits)
        """
        a = np.asarray(arr)

        if a.dtype == np.bool_:
            out = a.astype(np.uint8)
            return out, out.dtype, (0, 1)

        if np.issubdtype(a.dtype, np.integer):
            out = a
            info = np.iinfo(out.dtype)
            return out, out.dtype, (int(info.min), int(info.max))

        if np.issubdtype(a.dtype, np.floating):
            finite = a[np.isfinite(a)]
            if finite.size == 0:
                out = a.astype(np.float32, copy=False)
                return out, out.dtype, (0.0, 1.0)

            amin = float(finite.min())
            amax = float(finite.max())

            if amin >= -1e-6 and amax <= 1.0 + 1e-6:
                out = (a > 0).astype(np.uint8)
                return out, out.dtype, (0, 1)

            out = a.astype(np.float32, copy=False)
            return out, out.dtype, (amin, amax)

        out = np.asarray(a)
        return out, out.dtype, (float(np.nanmin(out)), float(np.nanmax(out)))

    def _set_slider_from_lineedit(self, lineedit, slider, lo, hi):
        try:
            val = int(lineedit.text())
        except Exception:
            return

        val = max(int(lo), min(int(hi), val))
        try:
            lineedit.setText(str(int(val)))
        except Exception:
            pass
        slider.setValue(int(val))

    def _get_layer_by_name(self, name):
        for lyr in self.viewer.layers:
            if getattr(lyr, "name", None) == name:
                return lyr
        return None

    def _add_coloc_controls(self, layer, title: str, max_range: int = 65535, is_binary: bool = False):
        gb = QGroupBox(title)
        v = QVBoxLayout()

        # Opacity
        op_row = QHBoxLayout()
        op_row.addWidget(QLabel("Opacity"))
        op = QSlider(Qt.Horizontal)
        op.setRange(0, 100)
        op.setValue(int(layer.opacity * 100) if layer.opacity is not None else 70)
        op_val = QLineEdit(str(op.value()))
        op_val.setFixedWidth(60)
        op_val.setValidator(QIntValidator(0, 100))

        def _set_opacity(val):
            layer.opacity = int(val) / 100.0
            op_val.setText(str(int(val)))

        op.valueChanged.connect(_set_opacity)
        op_row.addWidget(op)
        op_row.addWidget(op_val)
        v.addLayout(op_row)

        # Contrast limits
        cl_row1 = QHBoxLayout()
        cl_row1.addWidget(QLabel("Min"))
        smin = QSlider(Qt.Horizontal)
        smin.setRange(0, 1 if is_binary else max_range)
        cl_row1.addWidget(smin)
        v.addLayout(cl_row1)

        cl_row2 = QHBoxLayout()
        cl_row2.addWidget(QLabel("Max"))
        smax = QSlider(Qt.Horizontal)
        smax.setRange(0, 1 if is_binary else max_range)
        cl_row2.addWidget(smax)
        v.addLayout(cl_row2)

        try:
            c0, c1 = layer.contrast_limits
        except Exception:
            c0, c1 = (0, 1) if is_binary else (0, max_range)

        c0 = int(max(0, min(c0, smin.maximum())))
        c1 = int(max(0, min(c1, smax.maximum())))
        if c1 < c0:
            c1 = c0

        smin.setValue(c0)
        smax.setValue(c1)

        def _apply_clims(_=None):
            lo = int(smin.value())
            hi = int(smax.value())
            if hi < lo:
                hi = lo
                smax.setValue(hi)
            try:
                layer.contrast_limits = (lo, hi)
            except Exception:
                pass

        smin.valueChanged.connect(_apply_clims)
        smax.valueChanged.connect(_apply_clims)

        gb.setLayout(v)
        ctrl = {"opacity": op, "cmin": smin, "cmax": smax, "layer": layer}
        return gb, ctrl

    def ensure_layer_controls(self, layer, key, title=None, max_range=65535):
        """
        Create controls once per 'key'. key can be ('mask', ch_idx) or ('derived', name).
        """
        if not hasattr(self, "layer_control_widgets"):
            self.layer_control_widgets = {}

        if key in self.layer_control_widgets:
            return

        gb = QGroupBox(title or layer.name)
        vlay = QVBoxLayout(gb)

        row = QHBoxLayout()
        vis = QCheckBox("Visible")
        vis.setChecked(True)
        vis.stateChanged.connect(lambda s, lyr=layer: setattr(lyr, "visible", bool(s)))
        row.addWidget(vis)

        row.addWidget(QLabel("Opacity"))
        op = QSlider(Qt.Horizontal)
        op.setRange(0, 100)
        op.setValue(int(getattr(layer, "opacity", 1.0) * 100))
        op.valueChanged.connect(lambda v, lyr=layer: setattr(lyr, "opacity", v / 100.0))
        row.addWidget(op)
        vlay.addLayout(row)

        try:
            mn, mx = layer.contrast_limits
        except Exception:
            mn, mx = 0, max_range

        min_sl = QSlider(Qt.Horizontal)
        min_sl.setRange(0, max_range)
        min_sl.setValue(int(mn))

        max_sl = QSlider(Qt.Horizontal)
        max_sl.setRange(0, max_range)
        max_sl.setValue(int(mx))

        def upd():
            try:
                a = int(min_sl.value())
                b = int(max_sl.value())
                if b <= a:
                    b = a + 1
                layer.contrast_limits = (a, b)
            except Exception:
                pass

        min_sl.valueChanged.connect(lambda _=None: upd())
        max_sl.valueChanged.connect(lambda _=None: upd())

        vlay.addWidget(QLabel("Min"))
        vlay.addWidget(min_sl)
        vlay.addWidget(QLabel("Max"))
        vlay.addWidget(max_sl)

        self.layer_control_widgets[key] = gb
        self.mask_container.addWidget(gb)

    def _add_channel_controls(self, layer, idx, max_range=65535):
        """Create UI controls for one channel and return (groupbox, ctrl_dict)."""
        gb = QGroupBox(f"Channel {idx} Controls")
        vlay = QVBoxLayout(gb)

        top_h = QHBoxLayout()
        vis_cb = QCheckBox("Visible")
        vis_cb.setChecked(True)
        vis_cb.stateChanged.connect(lambda s, lyr=layer: setattr(lyr, "visible", bool(s)))
        top_h.addWidget(vis_cb)

        top_h.addWidget(QLabel("Opacity"))
        op = QSlider(Qt.Horizontal)
        op.setRange(0, 100)
        op.setValue(int(getattr(layer, "opacity", 1.0) * 100))
        op.valueChanged.connect(lambda v, lyr=layer: setattr(lyr, "opacity", v / 100.0))
        top_h.addWidget(op)

        color_btn = QPushButton("Color")

        def make_color_picker(lyr):
            def pick():
                col = QColorDialog.getColor(parent=self)
                if col.isValid():
                    try:
                        lyr.colormap = col.name()
                    except Exception:
                        pass
            return pick

        color_btn.clicked.connect(make_color_picker(layer))
        top_h.addWidget(color_btn)

        vlay.addLayout(top_h)

        lim = self.contrast_limits.get(f"ch{idx}", (0, max_range))

        def to_slider(val):
            if max_range <= 1:
                return int(round(float(val)))
            return int(round(float(val)))

        def from_slider(v):
            if max_range <= 1:
                return float(v)
            return float(v)

        ch_min_layout = QHBoxLayout()
        ch_min_layout.addWidget(QLabel("Min:"))
        ch_min = QSlider(Qt.Horizontal)
        ch_min.setRange(0, int(max_range))
        ch_min.setValue(to_slider(lim[0]))
        ch_min_layout.addWidget(ch_min)

        ch_min_val = QLineEdit(str(to_slider(lim[0])))
        ch_min_val.setFixedWidth(80)
        ch_min_val.setValidator(QIntValidator(0, int(max_range)))
        ch_min_val.editingFinished.connect(
            lambda le=ch_min_val, sl=ch_min: self._set_slider_from_lineedit(le, sl, 0, int(max_range))
        )
        ch_min_layout.addWidget(ch_min_val)
        vlay.addLayout(ch_min_layout)

        ch_max_layout = QHBoxLayout()
        ch_max_layout.addWidget(QLabel("Max:"))
        ch_max = QSlider(Qt.Horizontal)
        ch_max.setRange(0, int(max_range))
        ch_max.setValue(to_slider(lim[1]))
        ch_max_layout.addWidget(ch_max)

        ch_max_val = QLineEdit(str(to_slider(lim[1])))
        ch_max_val.setFixedWidth(80)
        ch_max_val.setValidator(QIntValidator(0, int(max_range)))
        ch_max_val.editingFinished.connect(
            lambda le=ch_max_val, sl=ch_max: self._set_slider_from_lineedit(le, sl, 0, int(max_range))
        )
        ch_max_layout.addWidget(ch_max_val)
        vlay.addLayout(ch_max_layout)

        def _apply_contrast_from_sliders(
            _=None,
            lyr=layer,
            smin=ch_min,
            smax=ch_max,
            le_min=ch_min_val,
            le_max=ch_max_val,
            i=idx,
        ):
            lo_i = int(smin.value())
            hi_i = int(smax.value())

            if hi_i < lo_i:
                hi_i = lo_i
                smax.blockSignals(True)
                smax.setValue(hi_i)
                smax.blockSignals(False)

            le_min.setText(str(lo_i))
            le_max.setText(str(hi_i))

            lo = from_slider(lo_i)
            hi = from_slider(hi_i)

            try:
                lyr.contrast_limits = (lo, hi)
            except Exception:
                pass

            self.contrast_limits[f"ch{i}"] = (lo, hi)

        ch_min.valueChanged.connect(_apply_contrast_from_sliders)
        ch_max.valueChanged.connect(_apply_contrast_from_sliders)

        ch_min_val.setText(str(int(ch_min.value())))
        ch_max_val.setText(str(int(ch_max.value())))

        ctrl = {
            "visible_cb": vis_cb,
            "opacity_slider": op,
            "min_slider": ch_min,
            "max_slider": ch_max,
            "min_val": ch_min_val,
            "max_val": ch_max_val,
            "layer": layer,
            "groupbox": gb,
        }
        return gb, ctrl


    def open_analysis(self):
        # Keep a reference so dialog doesn't get garbage-collected
        self.analysis_dialog = AnalysisPopupDialog(self)
        self.analysis_dialog.show()

    def save_output(self):
        """Save metadata (UI + per-channel display settings) to JSON."""
        try:
            metadata_dir = Path("meta_data")
            metadata_dir.mkdir(exist_ok=True)

            original_filename = Path(self.file_path)  # <-- use file_path
            json_filename = original_filename.stem + ".json"
            metadata_path = metadata_dir / json_filename

            channels_meta = []
            for i, layer in enumerate(getattr(self, "channel_layers", []), start=1):
                if layer is None:
                    continue

                ch = {
                    "channel_index": i,
                    "layer_name": getattr(layer, "name", f"ch{i}"),
                    "visible": bool(getattr(layer, "visible", True)),
                    "opacity": float(getattr(layer, "opacity", 1.0)),
                }

                # napari Image layer properties (safe if present) 
                try:
                    cl = getattr(layer, "contrast_limits", None)
                    if cl is not None:
                        ch["contrast_limits"] = [float(cl[0]), float(cl[1])]
                except Exception:
                    pass

                try:
                    cm = getattr(layer, "colormap", None)
                    # colormap can be an object; str() is at least JSON-serializable
                    if cm is not None:
                        ch["colormap"] = str(cm)
                except Exception:
                    pass

                # If you stored widgets per channel, save the actual slider values too
                if hasattr(self, "channel_controls") and i - 1 < len(self.channel_controls):
                    ctrl = self.channel_controls[i - 1]
                    try:
                        ch["ui"] = {
                            "opacity_slider": ctrl["opacity_slider"].value() / 100.0,
                            "min": ctrl["min_slider"].value(),
                            "max": ctrl["max_slider"].value(),
                            "brightness": ctrl["brightness_slider"].value(),
                            "contrast_percent": ctrl["contrast_slider"].value(),
                            "visible_checkbox": bool(ctrl["visible_cb"].isChecked()),
                        }
                    except Exception:
                        pass

                channels_meta.append(ch)

            metadata = {
                "original_file": str(original_filename),
                "timestamp": datetime.now().isoformat(),  # JSON-friendly 
                "current_z_slice": int(getattr(self, "current_z", 0)),
                "total_z_slices": int(getattr(self, "n_z", 1)),
                "channels": channels_meta,
            }

            with open(metadata_path, "w") as f:
                json.dump(metadata, f, indent=4)

            QMessageBox.information(self, "Success", f"Metadata saved to:\n{metadata_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save metadata: {str(e)}")
    
    def load_metadata(self):
        """Load metadata JSON from meta_data/<thisfile>.json and apply to UI/layers."""
        try:
            # Match save_output(): metadata filename is based on self.file_path
            srcp = Path(self.file_path)
            meta_dir = Path("meta_data")
            metadata_path = meta_dir / (srcp.stem + ".json")  # same stem logic as save_output 

            if not metadata_path.exists():
                QMessageBox.critical(self, "Not Found", f"No metadata found.\nExpected: {metadata_path}")
                return

            with open(metadata_path, "r") as f:
                md = json.load(f)

            # ---------- Restore per-channel ----------
            channels = md.get("channels", [])
            for ch in channels:
                try:
                    idx1 = int(ch.get("channel_index", -1))   # 1-based in your JSON
                    i = idx1 - 1                              # 0-based
                    if i < 0:
                        continue

                    lyr = None
                    if hasattr(self, "channel_layers") and i < len(self.channel_layers):
                        lyr = self.channel_layers[i]

                    # Layer properties
                    if lyr is not None:
                        if "visible" in ch:
                            try:
                                lyr.visible = bool(ch["visible"])
                            except Exception:
                                pass

                        if "opacity" in ch:
                            try:
                                lyr.opacity = float(ch["opacity"])
                            except Exception:
                                pass

                        if "contrast_limits" in ch:
                            try:
                                cl0, cl1 = ch["contrast_limits"]
                                lyr.contrast_limits = (float(cl0), float(cl1))
                            except Exception:
                                pass

                        # Only set if it's a simple usable string (your saver uses str(cm))
                        if "colormap" in ch:
                            cm = ch.get("colormap", None)
                            if isinstance(cm, str) and cm:
                                try:
                                    lyr.colormap = cm
                                except Exception:
                                    pass

                    # UI controls (only if you created/stored self.channel_controls)
                    ui = ch.get("ui", None)
                    if ui and hasattr(self, "channel_controls") and i < len(self.channel_controls):
                        ctrl = self.channel_controls[i]

                        def _set(widget, value):
                            widget.blockSignals(True)
                            try:
                                # checkbox vs slider
                                if hasattr(widget, "setChecked"):
                                    widget.setChecked(bool(value))
                                else:
                                    widget.setValue(int(value))
                            finally:
                                widget.blockSignals(False)

                        if "visible_cb" in ctrl and "visible_checkbox" in ui:
                            _set(ctrl["visible_cb"], ui["visible_checkbox"])

                        if "opacity_slider" in ctrl and "opacity_slider" in ui:
                            # saved as 0..1, slider expects 0..100
                            _set(ctrl["opacity_slider"], int(round(float(ui["opacity_slider"]) * 100.0)))

                        if "min_slider" in ctrl and "min" in ui:
                            _set(ctrl["min_slider"], ui["min"])
                        if "max_slider" in ctrl and "max" in ui:
                            _set(ctrl["max_slider"], ui["max"])
                        if "brightness_slider" in ctrl and "brightness" in ui:
                            _set(ctrl["brightness_slider"], ui["brightness"])
                        if "contrast_slider" in ctrl and "contrast_percent" in ui:
                            _set(ctrl["contrast_slider"], ui["contrast_percent"])

                except Exception:
                    pass

            # ---------- Restore Z ----------
            if "current_z_slice" in md and hasattr(self, "z_slider"):
                try:
                    z = int(md.get("current_z_slice", 0))
                    z = max(0, min(z, max(self.n_z - 1, 0)))
                    self.z_slider.setValue(z)  # should call z_changed
                except Exception:
                    pass

            # Force refresh (helpful if some signals were blocked)
            try:
                self.z_changed(self.z_slider.value())
            except Exception:
                pass

            QMessageBox.information(self, "Loaded", f"Metadata loaded successfully from:\n{metadata_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load metadata:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def _cell_counter_safe_stem(self, text):
        text = str(text).strip() if text is not None else "unnamed"
        text = Path(text).stem
        text = re.sub(r"[^\w\-.]+", "_", text)
        return text or "unnamed"


    def _cell_counter_extract_channel_number(self, src_layer):
        try:
            md = getattr(src_layer, "metadata", {}) or {}
            ch_num = md.get("channel_number", None)
            if ch_num is not None:
                return int(ch_num)
        except Exception:
            pass

        try:
            nm = str(getattr(src_layer, "name", "") or "")
            m = re.search(r"Channel\s*(\d+)", nm, flags=re.IGNORECASE)
            if m:
                return int(m.group(1))
        except Exception:
            pass

        return 1


    def _cell_counter_get_source_file_path(self):
        """
        Prefer the same file source used by your working channel metadata save/load.
        """
        try:
            fp = getattr(self, "file_path", None)
            if fp:
                return Path(fp)
        except Exception:
            pass

        try:
            fp = getattr(self, "current_file_path", None)
            if fp:
                return Path(fp)
        except Exception:
            pass

        try:
            st = getattr(self, "_cell_counter_state", {}) or {}
            src_layer = st.get("source_main_layer", None)
            if src_layer is not None:
                md = getattr(src_layer, "metadata", {}) or {}
                fp = (
                    md.get("path")
                    or md.get("file_path")
                    or md.get("source_file")
                    or md.get("filename")
                )
                if fp:
                    return Path(fp)
        except Exception:
            pass

        return None


    def _cell_counter_roi_json_path(self):
        st = getattr(self, "_cell_counter_state", {}) or {}
        src_layer = st.get("source_main_layer", None)

        channel_num = self._cell_counter_extract_channel_number(src_layer)
        src_path = self._cell_counter_get_source_file_path()

        if src_path is None:
            raise FileNotFoundError(
                "Could not resolve source file path for ROI metadata. "
                "Make sure self.file_path is set."
            )

        metadata_dir = Path("meta_data")
        metadata_dir.mkdir(exist_ok=True)

        json_filename = f"{src_path.stem}_Channel_{channel_num}_ROIs.json"
        return metadata_dir / json_filename


    def _cell_counter_shapes_to_list(self, layer):
        out = []

        try:
            data_list = list(layer.data) if getattr(layer, "data", None) is not None else []
        except Exception:
            data_list = []

        try:
            shape_types = list(getattr(layer, "shape_type", []))
        except Exception:
            try:
                shape_types = list(getattr(layer, "shape_types", []))
            except Exception:
                shape_types = []

        for i, verts in enumerate(data_list):
            arr = np.asarray(verts, dtype=float)
            if arr.ndim != 2 or arr.shape[0] < 2:
                continue

            shape_type = shape_types[i] if i < len(shape_types) else "polygon"

            out.append({
                "shape_type": str(shape_type),
                "vertices": arr.tolist(),
            })

        return out


    def _cell_counter_clear_shapes_layer(self, layer):
        try:
            layer.selected_data = set(range(len(layer.data)))
        except Exception:
            pass

        try:
            layer.remove_selected()
            return
        except Exception:
            pass

        try:
            layer.data = []
        except Exception:
            pass


    def _cell_counter_load_shapes_from_list(self, layer, items):
        self._cell_counter_clear_shapes_layer(layer)

        if not items:
            try:
                layer.visible = True
                layer.refresh()
            except Exception:
                pass
            return []

        loaded_z = []

        for item in items:
            try:
                verts = np.asarray(item.get("vertices", []), dtype=float)
                shape_type = str(item.get("shape_type", "polygon"))

                if verts.ndim != 2 or verts.shape[0] < 2:
                    continue

                if verts.shape[1] == 2:
                    z_now = 0
                    try:
                        viewer = getattr(self, "_cell_counter_state", {}).get("viewer", None)
                        if viewer is not None and hasattr(viewer, "dims") and len(viewer.dims.point) > 0:
                            z_now = int(viewer.dims.point[0])
                    except Exception:
                        z_now = 0
                    verts = np.c_[np.full((verts.shape[0], 1), z_now), verts]

                layer.add(verts, shape_type=shape_type)

                if verts.shape[1] >= 3:
                    try:
                        loaded_z.append(int(round(float(verts[0, 0]))))
                    except Exception:
                        pass

            except Exception:
                continue

        try:
            layer.visible = True
        except Exception:
            pass

        try:
            layer.refresh()
        except Exception:
            pass

        return loaded_z


    def _cell_counter_save_roi_metadata(self):
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            QMessageBox.warning(self, "ROI metadata", "Cell counter state not found.")
            return

        try:
            out_path = self._cell_counter_roi_json_path()
            src_path = self._cell_counter_get_source_file_path()

            payload = {
                "original_file": str(src_path) if src_path is not None else "",
                "channel_number": self._cell_counter_extract_channel_number(
                    st.get("source_main_layer", None)
                ),
                "positive_rois": self._cell_counter_shapes_to_list(st["pos_shapes"]),
                "negative_rois": self._cell_counter_shapes_to_list(st["neg_shapes"]),
            }

            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=4)

            QMessageBox.information(
                self,
                "ROI metadata saved",
                f"Saved ROI metadata to:\n{out_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "ROI metadata", f"Failed to save ROI metadata:\n{e}")


    def _cell_counter_load_roi_metadata(self):
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            QMessageBox.warning(self, "ROI metadata", "Cell counter state not found.")
            return

        try:
            in_path = self._cell_counter_roi_json_path()
            if not in_path.exists():
                QMessageBox.information(
                    self,
                    "ROI metadata",
                    f"No ROI metadata file found for this image/channel:\n{in_path}"
                )
                return

            with open(in_path, "r", encoding="utf-8") as f:
                payload = json.load(f)

            pos_layer = st["pos_shapes"]
            neg_layer = st["neg_shapes"]

            pos_z = self._cell_counter_load_shapes_from_list(
                pos_layer,
                payload.get("positive_rois", []),
            )
            neg_z = self._cell_counter_load_shapes_from_list(
                neg_layer,
                payload.get("negative_rois", []),
            )

            try:
                pos_layer.edge_color = "lime"
                pos_layer.face_color = "transparent"
                pos_layer.visible = True
            except Exception:
                pass

            try:
                neg_layer.edge_color = "red"
                neg_layer.face_color = "transparent"
                neg_layer.visible = True
            except Exception:
                pass

            all_z = []
            all_z.extend(pos_z or [])
            all_z.extend(neg_z or [])

            try:
                viewer = st.get("viewer", None)
                vol = st.get("vol", None)

                if viewer is not None and all_z:
                    z_target = int(all_z[0])

                    if vol is not None:
                        zmax = max(0, int(np.asarray(vol).shape[0]) - 1)
                        z_target = max(0, min(z_target, zmax))

                    viewer.dims.set_point(0, z_target)

                if viewer is not None:
                    try:
                        viewer.layers.selection.active = pos_layer if len(pos_layer.data) else neg_layer
                    except Exception:
                        pass
                    try:
                        viewer.reset_view()
                    except Exception:
                        pass
            except Exception:
                pass

            QMessageBox.information(
                self,
                "ROI metadata loaded",
                f"Loaded ROI metadata from:\n{in_path}\n\n"
                f"Positive ROIs: {len(payload.get('positive_rois', []))}\n"
                f"Negative ROIs: {len(payload.get('negative_rois', []))}"
            )

        except Exception as e:
            QMessageBox.critical(self, "ROI metadata", f"Failed to load ROI metadata:\n{e}")

    def open_cell_counter(self):
        """
        Single-image-layer Cell Counter with two detection backends:
        - Classic (threshold/cleanup/watershed) for mask generation
        - ML (Cellpose train/load/infer), optional

        Intended workflow:
        A) Manual detection + mask generation (Classic + label editing)
        B) ML (train/infer) using the curated labels mask
        C) Analysis (measure intensities, screenshots)
        """
        try:
            layer_names = [lyr.name for lyr in self.viewer.layers]
            if not layer_names:
                QMessageBox.warning(self, "No layers", "No layers found in viewer.")
                return

            # --- dialog: select exactly ONE layer ---
            dlg = QDialog(self)
            dlg.setWindowTitle("Cell counter: select layer")
            dlg.setMinimumWidth(420)
            v = QVBoxLayout(dlg)
            v.addWidget(QLabel("Select ONE layer to use for cell detection + intensity measurement:"))

            lst = QListWidget()
            lst.setSelectionMode(QAbstractItemView.SingleSelection)
            for nm in layer_names:
                it = QListWidgetItem(nm)
                lst.addItem(it)
                if lst.currentItem() is None and nm.startswith("Channel ") and nm[8:].isdigit():
                    lst.setCurrentItem(it)
            v.addWidget(lst)

            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            v.addWidget(buttons)
            buttons.accepted.connect(dlg.accept)
            buttons.rejected.connect(dlg.reject)

            if dlg.exec() != QDialog.Accepted:
                return

            item = lst.currentItem()
            if item is None:
                QMessageBox.information(self, "Cell Counter", "No layer selected.")
                return

            selected_name = item.text()

            # Grab the actual layer object
            try:
                src_layer = self.viewer.layers[selected_name]
            except Exception:
                src_layer = None

            # Resolve 3D stack
            s3 = self._resolve_stack3d(selected_name)
            if s3 is None:
                QMessageBox.warning(self, "Missing data", f"Could not resolve stack for: {selected_name}")
                return

            vol = np.asarray(s3, dtype=np.float32)
            if vol.ndim != 3:
                QMessageBox.warning(self, "Unsupported", f"{selected_name} is not 3D (shape {vol.shape}).")
                return

            # Pull current colormap from main layer
            cmap_arg = "gray"
            if src_layer is not None:
                cm = getattr(src_layer, "colormap", None)
                if isinstance(cm, str) and cm:
                    cmap_arg = cm
                elif isinstance(cm, tuple) and len(cm) >= 2 and isinstance(cm[0], str):
                    cmap_arg = (cm[0], cm[1])

            # --- create cell-counter viewer ---
            cc = napari.Viewer()
            cc.title = "Cell Counter"

            detection_layer_name = f"Detection input - {selected_name}"
            det_layer = cc.add_image(
                vol,
                name=detection_layer_name,
                colormap=cmap_arg,
                blending="additive",
                opacity=1.0,
            )

            # Scale bar
            cc.scale_bar.visible = True
            cc.scale_bar.unit = "um"
            cc.scale_bar.font_size = 12
            cc.scale_bar.color = "white"
            cc.scale_bar.box = True
            cc.scale_bar.position = "bottom_right"

            # Shapes (3D)
            pos_shapes = cc.add_shapes(
                name="Positive ROIs (seed somas)",
                opacity=0.7,
                ndim=3,
                edge_color="lime",
                face_color="transparent",
            )
            neg_shapes = cc.add_shapes(
                name="Negative ROIs (exclude junk)",
                opacity=0.7,
                ndim=3,
                edge_color="red",
                face_color="transparent",
            )

            # Labels
            labels_layer = cc.add_labels(np.zeros(vol.shape, dtype=np.int32), name="Detected cells (labels mask)")
            labels_layer.visible = True

            # Store refs (add ML fields)
            self._cell_counter_state = dict(
                viewer=cc,
                vol=vol,
                detection_layer=det_layer,
                source_main_layer=src_layer,
                pos_shapes=pos_shapes,
                neg_shapes=neg_shapes,
                labels=labels_layer,
                # ML state:
                ml_backend="none",          # "none" | "cellpose" | "stardist" | "custom"
                ml_model_path="",           # where weights/config live
                ml_device=self._torch_best_device(),
                status_label=None,
            )

            # Promote accidentally-2D shapes to 3D (slice-locked)
            def _promote_shapes_to_3d(layer, event=None):
                try:
                    z_now = int(cc.dims.point[0]) if hasattr(cc, "dims") and len(cc.dims.point) > 0 else 0
                except Exception:
                    z_now = 0

                changed = False
                new_data = []
                for poly in layer.data:
                    pts = np.asarray(poly) if poly is not None else None
                    if pts is None or pts.ndim != 2 or pts.shape[0] < 3:
                        new_data.append(poly)
                        continue
                    if pts.shape[1] == 2:
                        pts3 = np.c_[np.full(len(pts), z_now), pts]
                        new_data.append(pts3)
                        changed = True
                    else:
                        new_data.append(pts)

                if changed:
                    layer.data = new_data

            try:
                pos_shapes.events.data.connect(lambda e: _promote_shapes_to_3d(pos_shapes, e))
            except Exception:
                pass
            try:
                neg_shapes.events.data.connect(lambda e: _promote_shapes_to_3d(neg_shapes, e))
            except Exception:
                pass

            # ---------------- Dock controls (reorganized) ----------------
            container = QWidget()
            layout = QVBoxLayout(container)
            layout.setContentsMargins(6, 6, 6, 6)
            layout.setSpacing(6)


            # Status line (top)
            status = QLabel(f"Device: {self._cell_counter_state['ml_device']} | Backend: (none)")
            layout.addWidget(status)
            self._cell_counter_state["status_label"] = status

            # ===== Section A: Manual detection / mask generation =====
            layout.addWidget(QLabel("Manual detection + mask generation (Classic)"))

            layout.addWidget(QLabel("Step 1: Draw GREEN ROIs on somas (seeds) and RED ROIs on junk to exclude."))
            layout.addWidget(QLabel("Step 2: Click 'Generate/Update labels mask (Classic)'."))
            layout.addWidget(QLabel("Step 3: Manually edit 'Detected cells (labels mask)' if needed (this mask trains ML)."))

            # Threshold
            th_row = QHBoxLayout()
            th_row.addWidget(QLabel("Threshold (0–1)"))
            th = QSlider(Qt.Horizontal)
            th.setRange(0, 100)
            th.setValue(50)
            th_row.addWidget(th)
            th_val = QLineEdit("0.50")
            th_val.setFixedWidth(70)
            th_row.addWidget(th_val)
            layout.addLayout(th_row)

            def _sync_th(v=None):
                if v is None:
                    v = th.value()
                th_val.setText(f"{float(v) / 100.0:.2f}")

            th.valueChanged.connect(_sync_th)

            # Min voxels
            minvox_row = QHBoxLayout()
            minvox_row.addWidget(QLabel("Min voxels"))
            minvox = QSpinBox()
            minvox.setRange(1, 10_000_000)
            minvox.setValue(200)
            minvox_row.addWidget(minvox)
            layout.addLayout(minvox_row)

            # Max voxels
            maxvox_row = QHBoxLayout()
            maxvox_row.addWidget(QLabel("Max voxels (0=off)"))
            maxvox = QSpinBox()
            maxvox.setRange(0, 10_000_000)
            maxvox.setValue(0)
            maxvox_row.addWidget(maxvox)
            layout.addLayout(maxvox_row)

            # Compactness
            comp_row = QHBoxLayout()
            comp_row.addWidget(QLabel("Watershed compactness"))
            comp = QDoubleSpinBox()
            comp.setDecimals(2)
            comp.setRange(0.0, 50.0)
            comp.setValue(0.0)
            comp_row.addWidget(comp)
            layout.addLayout(comp_row)

            # Neg ROI grow
            neggrow_row = QHBoxLayout()
            neggrow_row.addWidget(QLabel("Neg ROI grow (px)"))
            neg_grow = QSpinBox()
            neg_grow.setRange(0, 50)
            neg_grow.setValue(2)
            neggrow_row.addWidget(neg_grow)
            layout.addLayout(neggrow_row)

            # Smoothing
            gs_row = QHBoxLayout()
            gs_row.addWidget(QLabel("Gaussian sigma (0=off)"))
            gs = QDoubleSpinBox()
            gs.setDecimals(2)
            gs.setRange(0.0, 10.0)
            gs.setValue(1.0)
            gs_row.addWidget(gs)
            layout.addLayout(gs_row)

            # Soma opening
            open_row = QHBoxLayout()
            open_row.addWidget(QLabel("Soma opening radius (px)"))
            open_r = QSpinBox()
            open_r.setRange(0, 10)
            open_r.setValue(1)
            open_row.addWidget(open_r)
            layout.addLayout(open_row)

            # Auto-split
            auto_split = QCheckBox("Auto-split touching cells (distance peaks)")
            auto_split.setChecked(True)
            layout.addWidget(auto_split)

            seed_row = QHBoxLayout()
            seed_row.addWidget(QLabel("Min seed distance (px)"))
            seed_dist = QSpinBox()
            seed_dist.setRange(1, 100)
            seed_dist.setValue(8)
            seed_row.addWidget(seed_dist)
            layout.addLayout(seed_row)

            # Manual/classic run button (renamed to clarify purpose)
            classic_btn = QPushButton("Generate/Update labels mask (Classic)")
            layout.addWidget(classic_btn)

            roi_meta_row = QHBoxLayout()
            save_roi_meta_btn = QPushButton("Save ROI metadata")
            load_roi_meta_btn = QPushButton("Load ROI metadata")
            roi_meta_row.addWidget(save_roi_meta_btn)
            roi_meta_row.addWidget(load_roi_meta_btn)
            layout.addLayout(roi_meta_row)

            save_roi_meta_btn.clicked.connect(lambda _=False: self._cell_counter_save_roi_metadata())
            load_roi_meta_btn.clicked.connect(lambda _=False: self._cell_counter_load_roi_metadata())

            def _run_classic():
                self._cell_counter_set_status("Backend: classic | Generating labels...")
                self._cell_counter_run_detection(
                    threshold=float(th.value()) / 100.0,
                    min_voxels=int(minvox.value()),
                    max_voxels=int(maxvox.value()),
                    compactness=float(comp.value()),
                    neg_grow=int(neg_grow.value()),
                    gaussian_sigma=float(gs.value()),
                    open_radius=int(open_r.value()),
                    auto_split=bool(auto_split.isChecked()),
                    min_seed_distance=int(seed_dist.value()),
                )
                # _cell_counter_run_detection already shows a QMessageBox; keep status updated anyway
                self._cell_counter_set_status("Backend: classic | Labels updated")

            classic_btn.clicked.connect(lambda _=False: _run_classic())

            # ===== Section B: ML (Cellpose) =====
            layout.addWidget(QLabel("ML model (Cellpose): train / load / infer (uses labels mask)"))
            layout.addWidget(QLabel("Train uses the current 'Detected cells (labels mask)' (0=bg, 1..N instances)."))

            ml_backend_row = QHBoxLayout()
            ml_backend_row.addWidget(QLabel("ML backend"))
            ml_backend = QComboBox()
            ml_backend.addItems(["Cellpose (recommended)"])  # keep focused until other backends implemented
            ml_backend_row.addWidget(ml_backend)
            layout.addLayout(ml_backend_row)

            device_row = QHBoxLayout()
            device_row.addWidget(QLabel("Compute device"))
            device = QComboBox()
            device.addItems(["auto", "cuda", "mps", "cpu"])
            device.setCurrentIndex(0)
            device_row.addWidget(device)
            layout.addLayout(device_row)

            model_path_row = QHBoxLayout()
            model_path_row.addWidget(QLabel("Model path"))
            model_path = QLineEdit("")
            model_path_row.addWidget(model_path)
            browse_btn = QPushButton("Browse")
            model_path_row.addWidget(browse_btn)
            layout.addLayout(model_path_row)

            def _browse_model():
                # NOTE: using qt_viewer here matches your existing code; you can swap to cc.window._qt_window later.
                pth = QFileDialog.getExistingDirectory(cc.window.qt_viewer, "Select model folder")
                if pth:
                    model_path.setText(pth)

            browse_btn.clicked.connect(_browse_model)

            # ---- NEW: Model mode (pretrained cpsam vs custom weights) ----
            ml_mode_row = QHBoxLayout()
            ml_mode_row.addWidget(QLabel("Model mode"))

            ml_mode = QComboBox()
            ml_mode.addItems([
                "Pretrained (cpsam) – no training data",
                "Custom weights (from folder/file) – train or load",
            ])
            ml_mode.setCurrentIndex(0)

            ml_mode_row.addWidget(ml_mode)
            layout.addLayout(ml_mode_row)

            self._cell_counter_state["ml_mode_combo"] = ml_mode
            # -------------------------------------------------------------

            load_btn = QPushButton("Load model / weights")
            train_btn = QPushButton("Train Cellpose from labels mask (manual-first)")
            infer_btn = QPushButton("Run Cellpose inference → write labels mask")
            save_btn = QPushButton("Save model / weights")

            layout.addWidget(load_btn)
            layout.addWidget(train_btn)
            layout.addWidget(infer_btn)
            layout.addWidget(save_btn)

            # --- ML progress UI (dock-local, in addition to napari's global progress) ---
            ml_prog_label = QLabel("ML progress: idle")
            ml_prog_bar = QProgressBar()
            ml_prog_bar.setRange(0, 100)
            ml_prog_bar.setValue(0)
            ml_prog_bar.setTextVisible(True)

            layout.addWidget(ml_prog_label)
            layout.addWidget(ml_prog_bar)

            cancel_btn = QPushButton("Cancel training")
            layout.addWidget(cancel_btn)
            cancel_btn.clicked.connect(lambda _=False: self._cell_counter_ml_cancel())

            self._cell_counter_state["ml_progress_label"] = ml_prog_label
            self._cell_counter_state["ml_progress_bar"] = ml_prog_bar
            self._cell_counter_state["ml_worker"] = None
            self._cell_counter_state["ml_cancel_flag"] = False

            # --- UPDATED: Load should support both modes ---
            # If mode=Pretrained -> just set cpsam and clear any custom file
            # If mode=Custom -> try to resolve weights file now (prompt if needed) and store it
            def _ml_load_clicked():
                st = getattr(self, "_cell_counter_state", None)
                if not st:
                    return
                mode_txt = ""
                try:
                    mode_txt = st["ml_mode_combo"].currentText()
                except Exception:
                    mode_txt = ""

                mp = str(model_path.text()).strip()

                if "Pretrained (cpsam)" in (mode_txt or ""):
                    st["ml_last_trained_model_file"] = ""
                    st["ml_model_path"] = mp
                    self._cell_counter_set_status(f"Device: {self._torch_best_device()} | Backend: cellpose | Weights: cpsam")
                    self._ml_ui_set_progress("ML progress: using pretrained cpsam", 0)
                    return

                # Custom mode: prompt/resolve a weights file now
                w = self._cell_counter_resolve_infer_weights(model_path=mp)
                st["ml_model_path"] = mp
                st["ml_last_trained_model_file"] = w if (w and w != "cpsam") else st.get("ml_last_trained_model_file", "")
                self._cell_counter_set_status(f"Device: {self._torch_best_device()} | Backend: cellpose | Weights: {os.path.basename(w) if os.path.exists(w) else w}")
                self._ml_ui_set_progress("ML progress: custom weights loaded", 0)

            load_btn.clicked.connect(lambda _=False: _ml_load_clicked())

            train_btn.clicked.connect(lambda _=False: self._cell_counter_ml_train_threaded(
                backend=str(ml_backend.currentText()),
                device=str(device.currentText()),
                model_path=str(model_path.text()).strip(),
            ))

            infer_btn.clicked.connect(lambda _=False: self._cell_counter_ml_infer(
                backend=str(ml_backend.currentText()),
                device=str(device.currentText()),
                model_path=str(model_path.text()).strip(),
            ))

            save_btn.clicked.connect(lambda _=False: self._cell_counter_ml_save(
                backend=str(ml_backend.currentText()),
                device=str(device.currentText()),
                model_path=str(model_path.text()).strip(),
            ))


            ml_mode_row = QHBoxLayout()
            ml_mode_row.addWidget(QLabel("Model mode"))

            ml_mode = QComboBox()
            ml_mode.addItems([
                "Pretrained (cpsam) – no training data",
                "Custom weights (from folder/file) – train or load",
            ])
            ml_mode.setCurrentIndex(0)

            ml_mode_row.addWidget(ml_mode)
            layout.addLayout(ml_mode_row)

            self._cell_counter_state["ml_mode_combo"] = ml_mode

            # ===== Section C: Analysis / export =====
            layout.addWidget(QLabel("Analysis / export"))

            measure_btn = QPushButton("Measure intensities per cell (min/max/mean; exclude zeros)")
            layout.addWidget(measure_btn)

            ss_btn = QPushButton("Save Screenshot")
            ss_btn.clicked.connect(
                lambda _=False: self.save_screenshot(
                    viewer=cc, parent=self, default_name="cell_counter.png", canvas_only=True
                )
            )
            layout.addWidget(ss_btn)

            measure_btn.clicked.connect(lambda _=False: self.calculate_intensities_per_cell())

            layout.addStretch(1)
            container.setLayout(layout)

            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setWidget(container)

            cc.window.add_dock_widget(scroll, area="right")
            return

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open cell counter:\n{e}")

    def open_puncta_counter(self):
        try:
            import csv
            import gc
            import json
            import traceback
            from pathlib import Path

            import numpy as np
            import napari

            from qtpy.QtCore import Qt, QEventLoop
            from qtpy.QtWidgets import (
                QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout, QMessageBox,
                QDialog, QListWidget, QListWidgetItem, QDialogButtonBox,
                QAbstractItemView, QComboBox, QCheckBox, QSpinBox, QDoubleSpinBox,
                QFileDialog, QScrollArea, QGroupBox, QInputDialog
            )

            from scipy import ndimage as ndi
            from scipy.spatial import cKDTree

            from skimage.measure import marching_cubes
            from skimage.morphology import (
                ball, binary_closing, binary_opening, binary_dilation, remove_small_objects
            )

            # ============================================================
            # Lightweight helpers
            # ============================================================
            def _safe_stem(text):
                text = str(text).strip() if text is not None else "unnamed"
                text = Path(text).stem
                text = "".join(ch if (ch.isalnum() or ch in ("_", "-", ".")) else "_" for ch in text)
                return text or "unnamed"

            def _round_int(x, default=0):
                try:
                    if x is None:
                        return int(default)
                    if isinstance(x, str) and not x.strip():
                        return int(default)
                    return int(round(float(x)))
                except Exception:
                    return int(default)

            def _suggest_voxel_size():
                suggested = None
                try:
                    vs = getattr(self, "voxel_size", None)
                    if isinstance(vs, (list, tuple)) and len(vs) == 3:
                        suggested = (float(vs[0]), float(vs[1]), float(vs[2]))
                except Exception:
                    pass

                if suggested is None:
                    suggested = (1.0, 1.0, 1.0)

                z_um, ok1 = QInputDialog.getDouble(
                    self, "Voxel size", "Z voxel size (µm):", float(suggested[0]), 1e-6, 1e9, 4
                )
                if not ok1:
                    return None
                y_um, ok2 = QInputDialog.getDouble(
                    self, "Voxel size", "Y voxel size (µm):", float(suggested[1]), 1e-6, 1e9, 4
                )
                if not ok2:
                    return None
                x_um, ok3 = QInputDialog.getDouble(
                    self, "Voxel size", "X voxel size (µm):", float(suggested[2]), 1e-6, 1e9, 4
                )
                if not ok3:
                    return None

                return (float(z_um), float(y_um), float(x_um))

            def _warn_if_spacing_looks_swapped(spacing_zyx):
                try:
                    z, y, x = map(float, spacing_zyx)
                    if z < min(y, x):
                        QMessageBox.information(
                            self,
                            "Voxel size check",
                            "You entered voxel size as (Z, Y, X).\n\n"
                            f"Current values: ({z}, {y}, {x})\n\n"
                            "Z is smaller than X/Y, which will compress the object along Z.\n"
                            "If your microscope has larger spacing between slices than in-plane pixels,\n"
                            "you may want something like (z_step_um, y_pixel_um, x_pixel_um)."
                        )
                except Exception:
                    pass

            def _fallback_color(name):
                palette = [
                    (1.0, 0.2, 0.2, 0.90),
                    (0.2, 1.0, 0.2, 0.90),
                    (0.2, 0.5, 1.0, 0.90),
                    (1.0, 1.0, 0.2, 0.90),
                    (1.0, 0.2, 1.0, 0.90),
                    (0.2, 1.0, 1.0, 0.90),
                ]
                return palette[abs(hash(name)) % len(palette)]

            def _hex_to_rgba(hex_str, alpha=0.90):
                s = str(hex_str).strip()
                if not s.startswith("#"):
                    return None
                h = s[1:]
                if len(h) == 6:
                    r = int(h[0:2], 16) / 255.0
                    g = int(h[2:4], 16) / 255.0
                    b = int(h[4:6], 16) / 255.0
                    return (r, g, b, float(alpha))
                if len(h) == 8:
                    a = int(h[0:2], 16) / 255.0
                    r = int(h[2:4], 16) / 255.0
                    g = int(h[4:6], 16) / 255.0
                    b = int(h[6:8], 16) / 255.0
                    return (r, g, b, a)
                return None

            NAMED_RGBA = {
                "red":     (1.0, 0.0, 0.0, 0.90),
                "green":   (0.0, 1.0, 0.0, 0.90),
                "blue":    (0.0, 0.4, 1.0, 0.90),
                "magenta": (1.0, 0.0, 1.0, 0.90),
                "cyan":    (0.0, 1.0, 1.0, 0.90),
                "yellow":  (1.0, 1.0, 0.0, 0.90),
                "gray":    (0.85, 0.85, 0.85, 0.90),
                "grey":    (0.85, 0.85, 0.85, 0.90),
                "orange":  (1.0, 0.55, 0.1, 0.90),
                "white":   (1.0, 1.0, 1.0, 0.90),
            }

            def _rgba_from_layer_current(layer_obj, fallback_name):
                if layer_obj is None:
                    return _fallback_color(fallback_name)

                cm = getattr(layer_obj, "colormap", None)
                if cm is None:
                    return _fallback_color(fallback_name)

                if isinstance(cm, str):
                    rgba = _hex_to_rgba(cm)
                    if rgba is not None:
                        return rgba
                    if cm.lower() in NAMED_RGBA:
                        return NAMED_RGBA[cm.lower()]
                    return _fallback_color(fallback_name)

                if isinstance(cm, tuple) and len(cm) >= 2:
                    name = cm[0]
                    cmap_obj = cm[1]

                    if isinstance(name, str):
                        rgba = _hex_to_rgba(name)
                        if rgba is not None:
                            return rgba
                        if name.lower() in NAMED_RGBA:
                            return NAMED_RGBA[name.lower()]

                    try:
                        if hasattr(cmap_obj, "map"):
                            rgba = tuple(float(x) for x in cmap_obj.map([0.95])[0])
                            if len(rgba) == 4:
                                return (rgba[0], rgba[1], rgba[2], 0.90)
                    except Exception:
                        pass

                try:
                    if hasattr(cm, "map"):
                        rgba = tuple(float(x) for x in cm.map([0.95])[0])
                        if len(rgba) == 4:
                            return (rgba[0], rgba[1], rgba[2], 0.90)
                except Exception:
                    pass

                return _fallback_color(fallback_name)

            def _configure_viewer_3d(vw, title):
                vw.title = title
                try:
                    vw.scale_bar.visible = True
                    vw.scale_bar.unit = "um"
                    vw.scale_bar.font_size = 12
                    vw.scale_bar.color = "white"
                    vw.scale_bar.box = True
                    vw.scale_bar.position = "bottom_right"
                except Exception:
                    pass
                try:
                    vw.dims.axis_labels = ["Z", "Y", "X"]
                except Exception:
                    pass

            def _update_status(msg):
                try:
                    status_label.setText(str(msg))
                except Exception:
                    pass

            def _current_spacing():
                st = getattr(self, "_puncta_counter_state", None) or {}
                sp = st.get("spacing_um", (1.0, 1.0, 1.0))
                return (float(sp[0]), float(sp[1]), float(sp[2]))

            def _set_spacing(sp):
                st = self._puncta_counter_state
                st["spacing_um"] = (float(sp[0]), float(sp[1]), float(sp[2]))
                for lyr in st.get("surface_layers", {}).values():
                    try:
                        pass
                    except Exception:
                        pass
                for lyr in st.get("point_layers", {}).values():
                    try:
                        lyr.scale = st["spacing_um"]
                    except Exception:
                        pass

            def _ask_role_assignments(layer_names):
                dlg = QDialog(self)
                dlg.setWindowTitle("Puncta Counter: assign channel roles")
                dlg.setMinimumWidth(520)
                v = QVBoxLayout(dlg)

                v.addWidget(QLabel("Assign channels to roles. Use 'None / N/A' for unused roles."))

                none_label = "None / N/A"
                role_combos = {}

                for role in ["Cell of interest", "Protein 1", "Protein 2", "Protein 3"]:
                    row = QHBoxLayout()
                    row.addWidget(QLabel(role))
                    combo = QComboBox()
                    combo.addItem(none_label)
                    for nm in layer_names:
                        combo.addItem(nm)
                    row.addWidget(combo)
                    v.addLayout(row)
                    role_combos[role] = combo

                buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                buttons.accepted.connect(dlg.accept)
                buttons.rejected.connect(dlg.reject)
                v.addWidget(buttons)

                if dlg.exec() != QDialog.Accepted:
                    return None

                role_map = {}
                for role, combo in role_combos.items():
                    val = str(combo.currentText())
                    role_map[role] = None if val == none_label else val

                used = [x for x in role_map.values() if x is not None]
                if not used:
                    QMessageBox.information(self, "Puncta Counter", "No channels were assigned.")
                    return None

                if len(used) != len(set(used)):
                    QMessageBox.warning(self, "Puncta Counter", "Each role must use a different channel.")
                    return None

                return role_map

            def _resolve_source_layer(name):
                try:
                    return self.viewer.layers[name]
                except Exception:
                    return None

            def _get_3d_array(name):
                src = _resolve_source_layer(name)
                if src is not None:
                    try:
                        arr = np.asarray(src.data)
                        if arr.ndim == 3:
                            return arr
                    except Exception:
                        pass
                try:
                    arr = np.asarray(self._resolve_stack3d(name))
                    if arr.ndim == 3:
                        return arr
                except Exception:
                    pass
                return None

            def _get_channel_list_from_role_map(role_map):
                out = []
                for role in ["Cell of interest", "Protein 1", "Protein 2", "Protein 3"]:
                    ch = role_map.get(role, None)
                    if ch is not None and ch not in out:
                        out.append(ch)
                return out

            def _default_cell_params(channel_name, spacing):
                return {
                    "source_layer": str(channel_name),
                    "voxel_size_um": [float(spacing[0]), float(spacing[1]), float(spacing[2])],
                    "keep_largest_component": True,
                    "fill_holes": True,
                    "remove_small_objects_min_voxels": 50,
                    "binary_closing_radius_px": 1,
                    "binary_opening_radius_px": 0,
                    "surface_smoothing_sigma_px": 0.0,
                    "marching_cubes_step_size": 1,
                }

            def _default_protein_params(channel_name, spacing):
                return {
                    "source_layer": str(channel_name),
                    "voxel_size_um": [float(spacing[0]), float(spacing[1]), float(spacing[2])],
                    "surface_smoothing_sigma_px": 0.0,
                    "marching_cubes_step_size": 1,
                }

            def _load_cell_metadata_file():
                json_path, _ = QFileDialog.getOpenFileName(
                    self,
                    "Load cell-of-interest metadata",
                    "",
                    "JSON files (*.json)"
                )
                if not json_path:
                    return None
                try:
                    with open(json_path, "r", encoding="utf-8") as f:
                        payload = json.load(f)
                    if not isinstance(payload, dict):
                        raise ValueError("Metadata JSON does not contain a dictionary.")
                    return payload
                except Exception as e:
                    QMessageBox.critical(self, "Load metadata", f"Failed to load metadata:\n{e}")
                    return None

            def _save_cell_metadata_file(payload, default_name):
                save_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Save meta data from cell of interest",
                    default_name,
                    "JSON files (*.json)"
                )
                if not save_path:
                    return None
                if not str(save_path).lower().endswith(".json"):
                    save_path = str(save_path) + ".json"
                try:
                    with open(save_path, "w", encoding="utf-8") as f:
                        json.dump(payload, f, indent=2)
                    return save_path
                except Exception as e:
                    QMessageBox.critical(self, "Save metadata", f"Failed to save metadata:\n{e}")
                    return None

            # ============================================================
            # Bounding-box helpers
            # ============================================================
            def _mask_bbox(mask):
                if mask is None or mask.ndim != 3 or not np.any(mask):
                    return None

                z_any = np.any(mask, axis=(1, 2))
                y_any = np.any(mask, axis=(0, 2))
                x_any = np.any(mask, axis=(0, 1))

                z_idx = np.where(z_any)[0]
                y_idx = np.where(y_any)[0]
                x_idx = np.where(x_any)[0]

                if len(z_idx) == 0 or len(y_idx) == 0 or len(x_idx) == 0:
                    return None

                return (
                    int(z_idx[0]), int(z_idx[-1]),
                    int(y_idx[0]), int(y_idx[-1]),
                    int(x_idx[0]), int(x_idx[-1]),
                )

            def _expand_bbox(bbox, shape, margin_zyx):
                z0, z1, y0, y1, x0, x1 = bbox
                mz, my, mx = margin_zyx
                return (
                    max(0, int(z0 - mz)),
                    min(shape[0] - 1, int(z1 + mz)),
                    max(0, int(y0 - my)),
                    min(shape[1] - 1, int(y1 + my)),
                    max(0, int(x0 - mx)),
                    min(shape[2] - 1, int(x1 + mx)),
                )

            def _bbox_to_slices(bbox):
                return (
                    slice(int(bbox[0]), int(bbox[1]) + 1),
                    slice(int(bbox[2]), int(bbox[3]) + 1),
                    slice(int(bbox[4]), int(bbox[5]) + 1),
                )

            def _bbox_from_coords(coords_zyx):
                arr = np.asarray(coords_zyx, dtype=np.int32)
                mn = arr.min(axis=0)
                mx = arr.max(axis=0)
                return (
                    int(mn[0]), int(mx[0]),
                    int(mn[1]), int(mx[1]),
                    int(mn[2]), int(mx[2]),
                )

            def _bboxes_intersect(a, b):
                return not (
                    a[1] < b[0] or b[1] < a[0] or
                    a[3] < b[2] or b[3] < a[2] or
                    a[5] < b[4] or b[5] < a[4]
                )

            # ============================================================
            # Surface helpers
            # ============================================================
            def _surface_from_binary(mask, spacing, src_layer, fallback_name, sigma_px=0.0, step_size=1):
                arr = np.asarray(mask, dtype=np.uint8)
                if arr.ndim != 3:
                    raise ValueError("Input mask is not 3D.")
                if not np.any(arr):
                    raise ValueError("Input mask is empty.")

                if float(sigma_px) > 0:
                    work = ndi.gaussian_filter(arr.astype(np.float32), sigma=float(sigma_px))
                else:
                    work = arr.astype(np.float32)

                verts, faces, normals, values = marching_cubes(
                    work,
                    level=0.5,
                    spacing=spacing,
                    step_size=max(1, int(step_size)),
                )
                rgba = _rgba_from_layer_current(src_layer, fallback_name)
                vertex_colors = np.tile(np.array(rgba, dtype=np.float32), (verts.shape[0], 1))

                out = {
                    "verts": np.asarray(verts, dtype=np.float32),
                    "faces": np.asarray(faces, dtype=np.int32),
                    "values": np.asarray(values, dtype=np.float32),
                    "vertex_colors": np.asarray(vertex_colors, dtype=np.float32),
                }

                del arr, work, normals
                gc.collect()
                return out

            def _write_obj(obj_path, verts_zyx_um, faces):
                verts_zyx_um = np.asarray(verts_zyx_um, dtype=float)
                faces = np.asarray(faces, dtype=np.int32)
                with open(obj_path, "w", encoding="utf-8") as f:
                    f.write("# Exported from puncta counter\n")
                    f.write("# Coordinates stored in microns\n")
                    for v in verts_zyx_um:
                        z, y, x = float(v[0]), float(v[1]), float(v[2])
                        f.write(f"v {x:.4f} {y:.4f} {z:.4f}\n")
                    for tri in faces:
                        a, b, c = int(tri[0]) + 1, int(tri[1]) + 1, int(tri[2]) + 1
                        f.write(f"f {a} {b} {c}\n")

            # ============================================================
            # Preview parameter dialogs
            # ============================================================
            def _edit_cell_params(initial_params, channel_name):
                params = dict(initial_params)

                dlg = QDialog(self)
                dlg.setWindowTitle(f"Cell-of-interest mesh parameters: {channel_name}")
                dlg.setMinimumWidth(460)
                v = QVBoxLayout(dlg)

                keep_largest_ck = QCheckBox("Keep largest connected component only")
                keep_largest_ck.setChecked(bool(params.get("keep_largest_component", True)))
                v.addWidget(keep_largest_ck)

                fill_holes_ck = QCheckBox("Fill internal holes for watertight mesh")
                fill_holes_ck.setChecked(bool(params.get("fill_holes", True)))
                v.addWidget(fill_holes_ck)

                row1 = QHBoxLayout()
                row1.addWidget(QLabel("Remove objects smaller than (voxels)"))
                min_obj = QSpinBox()
                min_obj.setRange(0, 100000000)
                min_obj.setValue(int(params.get("remove_small_objects_min_voxels", 50)))
                row1.addWidget(min_obj)
                v.addLayout(row1)

                row2 = QHBoxLayout()
                row2.addWidget(QLabel("Binary closing radius (px)"))
                close_r = QSpinBox()
                close_r.setRange(0, 20)
                close_r.setValue(int(params.get("binary_closing_radius_px", 1)))
                row2.addWidget(close_r)
                v.addLayout(row2)

                row3 = QHBoxLayout()
                row3.addWidget(QLabel("Binary opening radius (px)"))
                open_r = QSpinBox()
                open_r.setRange(0, 20)
                open_r.setValue(int(params.get("binary_opening_radius_px", 0)))
                row3.addWidget(open_r)
                v.addLayout(row3)

                row4 = QHBoxLayout()
                row4.addWidget(QLabel("Surface smoothing sigma (px)"))
                sigma = QDoubleSpinBox()
                sigma.setDecimals(2)
                sigma.setRange(0.0, 10.0)
                sigma.setValue(float(params.get("surface_smoothing_sigma_px", 0.0)))
                row4.addWidget(sigma)
                v.addLayout(row4)

                row5 = QHBoxLayout()
                row5.addWidget(QLabel("Marching cubes step size"))
                mc_step = QSpinBox()
                mc_step.setRange(1, 8)
                mc_step.setValue(int(params.get("marching_cubes_step_size", 1)))
                row5.addWidget(mc_step)
                v.addLayout(row5)

                buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                buttons.accepted.connect(dlg.accept)
                buttons.rejected.connect(dlg.reject)
                v.addWidget(buttons)

                if dlg.exec() != QDialog.Accepted:
                    return None

                params["keep_largest_component"] = bool(keep_largest_ck.isChecked())
                params["fill_holes"] = bool(fill_holes_ck.isChecked())
                params["remove_small_objects_min_voxels"] = int(min_obj.value())
                params["binary_closing_radius_px"] = int(close_r.value())
                params["binary_opening_radius_px"] = int(open_r.value())
                params["surface_smoothing_sigma_px"] = float(sigma.value())
                params["marching_cubes_step_size"] = int(mc_step.value())
                return params

            def _edit_protein_params(initial_params, role_label, channel_name):
                params = dict(initial_params)

                dlg = QDialog(self)
                dlg.setWindowTitle(f"{role_label} mesh parameters: {channel_name}")
                dlg.setMinimumWidth(420)
                v = QVBoxLayout(dlg)

                row1 = QHBoxLayout()
                row1.addWidget(QLabel("Surface smoothing sigma (px)"))
                sigma = QDoubleSpinBox()
                sigma.setDecimals(2)
                sigma.setRange(0.0, 10.0)
                sigma.setValue(float(params.get("surface_smoothing_sigma_px", 0.0)))
                row1.addWidget(sigma)
                v.addLayout(row1)

                row2 = QHBoxLayout()
                row2.addWidget(QLabel("Marching cubes step size"))
                mc_step = QSpinBox()
                mc_step.setRange(1, 8)
                mc_step.setValue(int(params.get("marching_cubes_step_size", 1)))
                row2.addWidget(mc_step)
                v.addLayout(row2)

                buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                buttons.accepted.connect(dlg.accept)
                buttons.rejected.connect(dlg.reject)
                v.addWidget(buttons)

                if dlg.exec() != QDialog.Accepted:
                    return None

                params["surface_smoothing_sigma_px"] = float(sigma.value())
                params["marching_cubes_step_size"] = int(mc_step.value())
                return params

            # ============================================================
            # Preview builders
            # ============================================================
            def _build_clean_cell_mask(cell_vol, params):
                mask = np.asarray(cell_vol) > 0
                if mask.ndim != 3:
                    raise ValueError("Cell-of-interest channel must be a 3D binary mask.")

                if int(params.get("binary_closing_radius_px", 0)) > 0:
                    mask = binary_closing(mask, footprint=ball(int(params["binary_closing_radius_px"])))

                if int(params.get("binary_opening_radius_px", 0)) > 0:
                    mask = binary_opening(mask, footprint=ball(int(params["binary_opening_radius_px"])))

                if bool(params.get("fill_holes", True)):
                    mask = ndi.binary_fill_holes(mask)

                min_vox = int(params.get("remove_small_objects_min_voxels", 0))
                if min_vox > 0:
                    mask = remove_small_objects(mask.astype(bool), min_size=min_vox)
                    mask = np.asarray(mask, dtype=bool)

                if bool(params.get("keep_largest_component", True)):
                    lab, n = ndi.label(mask)
                    if n > 0:
                        counts = np.bincount(lab.ravel())
                        counts[0] = 0
                        keep_id = int(np.argmax(counts))
                        mask = (lab == keep_id)
                    del lab, counts
                    gc.collect()

                if bool(params.get("fill_holes", True)):
                    mask = ndi.binary_fill_holes(mask)

                return np.asarray(mask, dtype=bool)

            def _prepare_cell_role(channel_name, spacing, params):
                vol = _get_3d_array(channel_name)
                if vol is None:
                    raise RuntimeError(f"Could not resolve 3D data for {channel_name}")

                clean_mask = _build_clean_cell_mask(vol, params)
                if not np.any(clean_mask):
                    raise ValueError("Cell mask is empty after cleanup.")

                src_layer = _resolve_source_layer(channel_name)
                surf = _surface_from_binary(
                    clean_mask,
                    spacing=spacing,
                    src_layer=src_layer,
                    fallback_name=channel_name,
                    sigma_px=float(params.get("surface_smoothing_sigma_px", 0.0)),
                    step_size=int(params.get("marching_cubes_step_size", 1)),
                )

                try:
                    tree = cKDTree(surf["verts"])
                except Exception:
                    tree = None

                metadata = dict(params)
                metadata["source_layer"] = str(channel_name)
                metadata["voxel_size_um"] = [float(spacing[0]), float(spacing[1]), float(spacing[2])]

                result = {
                    "role": "Cell of interest",
                    "kind": "cell",
                    "channel": channel_name,
                    "surface": surf,
                    "clean_mask": clean_mask,
                    "surface_tree": tree,
                    "metadata": metadata,
                }

                gc.collect()
                return result

            def _prepare_protein_role(role_label, channel_name, spacing, params):
                vol = _get_3d_array(channel_name)
                if vol is None:
                    raise RuntimeError(f"Could not resolve 3D data for {channel_name}")

                protein_mask = np.asarray(vol) > 0
                if not np.any(protein_mask):
                    raise ValueError(f"{role_label}: no positive voxels found.")

                src_layer = _resolve_source_layer(channel_name)
                surf = _surface_from_binary(
                    protein_mask,
                    spacing=spacing,
                    src_layer=src_layer,
                    fallback_name=channel_name,
                    sigma_px=float(params.get("surface_smoothing_sigma_px", 0.0)),
                    step_size=int(params.get("marching_cubes_step_size", 1)),
                )

                result = {
                    "role": role_label,
                    "kind": "protein",
                    "channel": channel_name,
                    "surface": surf,
                    "metadata": dict(params),
                }

                del protein_mask
                gc.collect()
                return result

            def _preview_result_loop(role_label, result_obj, spacing):
                pv = napari.Viewer(ndisplay=3)
                _configure_viewer_3d(pv, f"Preview: {role_label}")

                surf = result_obj["surface"]
                pv.add_surface(
                    (surf["verts"], surf["faces"], surf["values"]),
                    name=f"{role_label} surface",
                    shading="smooth",
                    vertex_colors=surf["vertex_colors"],
                )

                ctrl = QDialog(self)
                ctrl.setWindowTitle(f"{role_label} preview decision")
                ctrl.setMinimumWidth(420)
                cv = QVBoxLayout(ctrl)
                cv.addWidget(QLabel(
                    f"Inspect the preview viewer for {role_label}.\n"
                    f"Accept = use this mesh in the final analysis viewer.\n"
                    f"Rerun = close preview and choose new parameters."
                ))

                btn_row = QHBoxLayout()
                accept_btn = QPushButton("Accept")
                rerun_btn = QPushButton("Rerun")
                cancel_btn = QPushButton("Cancel / Skip")
                btn_row.addWidget(accept_btn)
                btn_row.addWidget(rerun_btn)
                btn_row.addWidget(cancel_btn)
                cv.addLayout(btn_row)

                decision = {"value": "cancel"}

                def _finish(val):
                    decision["value"] = val
                    try:
                        ctrl.done(0)
                    except Exception:
                        try:
                            ctrl.close()
                        except Exception:
                            pass

                accept_btn.clicked.connect(lambda _=False: _finish("accept"))
                rerun_btn.clicked.connect(lambda _=False: _finish("rerun"))
                cancel_btn.clicked.connect(lambda _=False: _finish("cancel"))

                loop = QEventLoop()
                ctrl.finished.connect(loop.quit)
                ctrl.show()
                loop.exec()

                try:
                    ctrl.close()
                except Exception:
                    pass
                try:
                    pv.close()
                except Exception:
                    pass

                gc.collect()
                return decision["value"]

            def _run_role_preview(role_label, channel_name, spacing, is_cell=False):
                if is_cell:
                    params = _default_cell_params(channel_name, spacing)
                    ask_load = QMessageBox.question(
                        self,
                        "Cell mesh metadata",
                        "Do you want to load cell-of-interest metadata before creating the mesh?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No,
                    )
                    if ask_load == QMessageBox.Yes:
                        payload = _load_cell_metadata_file()
                        if payload is not None:
                            for k, v in payload.items():
                                params[k] = v
                            params["source_layer"] = str(channel_name)
                            params["voxel_size_um"] = [float(spacing[0]), float(spacing[1]), float(spacing[2])]
                else:
                    params = _default_protein_params(channel_name, spacing)

                while True:
                    if is_cell:
                        edited = _edit_cell_params(params, channel_name)
                    else:
                        edited = _edit_protein_params(params, role_label, channel_name)

                    if edited is None:
                        return None

                    params = dict(edited)

                    try:
                        if is_cell:
                            result = _prepare_cell_role(channel_name, spacing, params)
                        else:
                            result = _prepare_protein_role(role_label, channel_name, spacing, params)
                    except Exception as e:
                        QMessageBox.critical(
                            self,
                            f"{role_label} preview",
                            f"Failed to build preview for {role_label} ({channel_name}):\n{e}"
                        )
                        continue

                    decision = _preview_result_loop(role_label, result, spacing)
                    if decision == "accept":
                        return result
                    if decision == "rerun":
                        if result.get("kind") == "cell":
                            try:
                                del result["clean_mask"]
                            except Exception:
                                pass
                        del result
                        gc.collect()
                        continue

                    if result.get("kind") == "cell":
                        try:
                            del result["clean_mask"]
                        except Exception:
                            pass
                    del result
                    gc.collect()
                    return None

            # ============================================================
            # Initial setup
            # ============================================================
            layer_names = [lyr.name for lyr in self.viewer.layers]
            if not layer_names:
                QMessageBox.warning(self, "No layers", "No layers found in viewer.")
                return

            spacing0 = _suggest_voxel_size()
            if spacing0 is None:
                return
            _warn_if_spacing_looks_swapped(spacing0)

            role_map = _ask_role_assignments(layer_names)
            if role_map is None:
                return

            selected_channels = _get_channel_list_from_role_map(role_map)
            available_channels = []
            bad_channels = []
            for nm in selected_channels:
                arr = _get_3d_array(nm)
                if arr is None or arr.ndim != 3:
                    bad_channels.append(nm)
                else:
                    available_channels.append(nm)

            if not available_channels:
                QMessageBox.warning(
                    self,
                    "Puncta Counter",
                    "No valid 3D channels could be resolved for the assigned roles."
                )
                return

            if bad_channels:
                QMessageBox.information(
                    self,
                    "Puncta Counter",
                    "These assigned channels could not be resolved as 3D and will be skipped:\n\n"
                    + "\n".join(bad_channels)
                )

            accepted_roles = {}
            ordered_roles = ["Cell of interest", "Protein 1", "Protein 2", "Protein 3"]

            for role in ordered_roles:
                ch = role_map.get(role, None)
                if ch is None or ch not in available_channels:
                    continue

                try:
                    res = _run_role_preview(
                        role_label=role,
                        channel_name=ch,
                        spacing=spacing0,
                        is_cell=(role == "Cell of interest"),
                    )
                except Exception as e:
                    QMessageBox.critical(self, "Puncta Counter", f"{role}: preview failed:\n{e}")
                    res = None

                if res is None:
                    skip = QMessageBox.question(
                        self,
                        "Skip role",
                        f"{role} was not accepted.\n\nDo you want to skip this role and continue?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes,
                    )
                    if skip == QMessageBox.No:
                        return
                    continue

                accepted_roles[role] = res

            if not accepted_roles:
                QMessageBox.information(self, "Puncta Counter", "No channels were accepted for the final viewer.")
                return

            # ============================================================
            # Final viewer - surfaces/points only
            # ============================================================
            pc = napari.Viewer(ndisplay=3)
            _configure_viewer_3d(pc, "Puncta Counter")

            cell_result = accepted_roles.get("Cell of interest", None)
            protein_roles = [r for r in ["Protein 1", "Protein 2", "Protein 3"] if r in accepted_roles]
            protein_channels = [accepted_roles[r]["channel"] for r in protein_roles]

            self._puncta_counter_state = {
                "viewer": pc,
                "spacing_um": tuple(spacing0),
                "role_map": dict(role_map),
                "accepted_roles": accepted_roles,
                "surface_layers": {},
                "point_layers": {},
                "analysis_cache": {},
                "cell_channel_name": cell_result["channel"] if cell_result is not None else "",
                "cell_mask": cell_result["clean_mask"] if cell_result is not None else None,
                "cell_bbox": _mask_bbox(cell_result["clean_mask"]) if cell_result is not None else None,
                "cell_mesh_verts_zyx_um": None if cell_result is None else np.asarray(cell_result["surface"]["verts"], dtype=np.float32),
                "cell_mesh_faces": None if cell_result is None else np.asarray(cell_result["surface"]["faces"], dtype=np.int32),
                "cell_surface_tree": None if cell_result is None else cell_result.get("surface_tree", None),
                "cell_mesh_metadata": {} if cell_result is None else dict(cell_result.get("metadata", {})),
                "cell_mesh_metadata_path": "",
                "protein_channels": list(protein_channels),
            }
            st = self._puncta_counter_state

            def _upsert_surface_layer(layer_name, surf_dict, shading="smooth", visible=True):
                if surf_dict is None:
                    return None
                verts = np.asarray(surf_dict["verts"], dtype=np.float32)
                faces = np.asarray(surf_dict["faces"], dtype=np.int32)
                values = np.asarray(surf_dict["values"], dtype=np.float32)
                vertex_colors = np.asarray(surf_dict["vertex_colors"], dtype=np.float32)

                existing = st["surface_layers"].get(layer_name, None)
                if existing is None:
                    lyr = pc.add_surface(
                        (verts, faces, values),
                        name=layer_name,
                        shading=shading,
                        vertex_colors=vertex_colors,
                    )
                    try:
                        lyr.visible = visible
                    except Exception:
                        pass
                    st["surface_layers"][layer_name] = lyr
                    return lyr

                try:
                    existing.data = (verts, faces, values)
                    existing.vertex_colors = vertex_colors
                    existing.visible = visible
                    return existing
                except Exception:
                    try:
                        pc.layers.remove(existing)
                    except Exception:
                        pass
                    lyr = pc.add_surface(
                        (verts, faces, values),
                        name=layer_name,
                        shading=shading,
                        vertex_colors=vertex_colors,
                    )
                    try:
                        lyr.visible = visible
                    except Exception:
                        pass
                    st["surface_layers"][layer_name] = lyr
                    return lyr

        

            for role in ordered_roles:
                if role not in accepted_roles:
                    continue
                res = accepted_roles[role]
                ch = res["channel"]
                _upsert_surface_layer(
                    f"{role} surface - {ch}",
                    res["surface"],
                    shading="smooth",
                    visible=True
                )

            # ============================================================
            # Analysis helpers
            # ============================================================
            def _current_cell_metadata_payload():
                payload = dict(st.get("cell_mesh_metadata", {}) or {})
                payload["source_layer"] = str(st.get("cell_channel_name", ""))
                payload["voxel_size_um"] = [float(x) for x in _current_spacing()]
                return payload

            def _save_cell_metadata_dialog():
                if st.get("cell_mask", None) is None:
                    QMessageBox.warning(self, "Save metadata", "No cell-of-interest mesh exists.")
                    return
                default_name = f"{_safe_stem(st.get('cell_channel_name', 'cell_of_interest'))}_cell_of_interest_mesh.json"
                save_path = _save_cell_metadata_file(_current_cell_metadata_payload(), default_name)
                if save_path:
                    st["cell_mesh_metadata_path"] = save_path
                    _update_status(f"Saved cell metadata: {save_path}")
                    QMessageBox.information(self, "Metadata saved", f"Saved metadata to:\n{save_path}")

            def _export_cell_mesh_obj():
                verts = st.get("cell_mesh_verts_zyx_um", None)
                faces = st.get("cell_mesh_faces", None)
                if verts is None or faces is None or len(verts) == 0:
                    QMessageBox.warning(self, "Export cell mesh", "No cell mesh exists yet.")
                    return

                cell_nm = str(st.get("cell_channel_name", "cell_of_interest"))
                suggested = f"{_safe_stem(cell_nm)}_cell_of_interest.obj"
                obj_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Export cell mesh OBJ",
                    suggested,
                    "OBJ files (*.obj)"
                )
                if not obj_path:
                    return
                if not str(obj_path).lower().endswith(".obj"):
                    obj_path = str(obj_path) + ".obj"

                try:
                    _write_obj(obj_path, verts, faces)
                    meta_path = str(Path(obj_path).with_suffix(".json"))

                    payload = _current_cell_metadata_payload()
                    payload["exported_obj"] = str(obj_path)
                    payload["vertex_count"] = int(len(verts))
                    payload["face_count"] = int(len(faces))

                    with open(meta_path, "w", encoding="utf-8") as f:
                        json.dump(payload, f, indent=2)

                    QMessageBox.information(
                        self,
                        "Export cell mesh",
                        f"Saved OBJ:\n{obj_path}\n\nSaved JSON:\n{meta_path}"
                    )
                    _update_status(f"Exported cell OBJ: {obj_path}")
                except Exception as e:
                    QMessageBox.critical(self, "Export cell mesh", f"Failed to export OBJ:\n{e}")

            def _get_selected_protein_channels():
                out = []
                for i in range(protein_channel_list.count()):
                    it = protein_channel_list.item(i)
                    if it.checkState() == Qt.Checked:
                        ch = it.data(Qt.UserRole)
                        if ch:
                            out.append(str(ch))
                return out

            def _sync_localization_combos():
                selected = _get_selected_protein_channels()
                cur_a = str(loc_ch_a.currentText()) if loc_ch_a.count() else ""
                cur_b = str(loc_ch_b.currentText()) if loc_ch_b.count() else ""

                loc_ch_a.blockSignals(True)
                loc_ch_b.blockSignals(True)

                loc_ch_a.clear()
                loc_ch_b.clear()
                for nm in selected:
                    loc_ch_a.addItem(nm)
                    loc_ch_b.addItem(nm)

                if cur_a in selected:
                    loc_ch_a.setCurrentText(cur_a)
                if cur_b in selected:
                    loc_ch_b.setCurrentText(cur_b)
                elif len(selected) >= 2:
                    loc_ch_b.setCurrentIndex(1)

                loc_ch_a.blockSignals(False)
                loc_ch_b.blockSignals(False)

            def _sync_loc_mode_widgets():
                is_centroid = str(loc_mode.currentText()).startswith("Centroid")
                loc_centroid_tol_um.setEnabled(is_centroid)
                loc_voxel_overlap_pct.setEnabled(not is_centroid)

            def _centroid_inside_mask(mask, centroid_pix):
                shp = mask.shape
                zi = int(np.clip(np.round(centroid_pix[0]), 0, shp[0] - 1))
                yi = int(np.clip(np.round(centroid_pix[1]), 0, shp[1] - 1))
                xi = int(np.clip(np.round(centroid_pix[2]), 0, shp[2] - 1))
                return bool(mask[zi, yi, xi])

            def _analysis_crop_bbox(shell_um):
                cell_mask = st.get("cell_mask", None)
                bbox = st.get("cell_bbox", None)
                if cell_mask is None or bbox is None:
                    return None

                sp = _current_spacing()
                mz = int(np.ceil(float(shell_um) / max(float(sp[0]), 1e-9)))
                my = int(np.ceil(float(shell_um) / max(float(sp[1]), 1e-9)))
                mx = int(np.ceil(float(shell_um) / max(float(sp[2]), 1e-9)))

                margin = (max(1, mz), max(1, my), max(1, mx))
                return _expand_bbox(bbox, cell_mask.shape, margin)

            def _channel_role_name(ch):
                for role, obj in accepted_roles.items():
                    if obj["channel"] == ch:
                        return role
                return ""

            def _prepare_roi_binary(channel_name, threshold_value, shell_um):
                vol = _get_3d_array(channel_name)
                if vol is None:
                    raise RuntimeError(f"Could not resolve 3D data for channel: {channel_name}")

                cell_mask = st.get("cell_mask", None)
                if cell_mask is None:
                    raise RuntimeError("Cell-of-interest mesh must exist first.")

                bbox = _analysis_crop_bbox(shell_um)
                if bbox is None:
                    raise RuntimeError("Cell mask bbox could not be determined.")

                slc = _bbox_to_slices(bbox)

                vol_crop = np.asarray(vol[slc])
                cell_crop = np.asarray(cell_mask[slc], dtype=bool)

                thr = float(threshold_value)
                binary_crop = np.asarray(vol_crop > thr, dtype=bool)

                if float(shell_um) <= 0:
                    roi_mask = cell_crop
                else:
                    spacing = tuple(float(x) for x in _current_spacing())
                    outside_dist_um = ndi.distance_transform_edt(~cell_crop, sampling=spacing)
                    roi_mask = cell_crop | (outside_dist_um <= float(shell_um))
                    del outside_dist_um

                binary_crop &= roi_mask

                del roi_mask
                gc.collect()
                return binary_crop, cell_crop, vol_crop, bbox

            def _analyze_puncta_channel(channel_name, store_coords=False, add_layers=True):
                cell_mask = st.get("cell_mask", None)
                tree = st.get("cell_surface_tree", None)
                if cell_mask is None or tree is None:
                    raise RuntimeError("Cell-of-interest mesh must exist first.")

                shell_um = float(surface_shell_um.value())
                binary_crop, cell_crop, vol_crop, crop_bbox = _prepare_roi_binary(
                    channel_name,
                    threshold_value=float(puncta_voxel_threshold.value()),
                    shell_um=shell_um,
                )

                if binary_crop.ndim != 3:
                    raise RuntimeError("Binary crop is not 3D.")

                displayed_layer_name = f"Displayed puncta centroids - {channel_name}"
                all_layer_name = f"All puncta centroids - {channel_name}"

                if not np.any(binary_crop):
                    if add_layers:
                        _upsert_points_layer(displayed_layer_name, np.zeros((0, 3)), face_color="cyan", size=6, visible=True)
                        _upsert_points_layer(all_layer_name, np.zeros((0, 3)), face_color="yellow", size=5, visible=False)

                    del binary_crop, cell_crop, vol_crop
                    gc.collect()
                    return {
                        "channel": channel_name,
                        "rows": [],
                        "summary": {
                            "total": 0,
                            "inside": 0,
                            "shell_only": 0,
                            "displayed": 0,
                        }
                    }

                lbl, _ = ndi.label(binary_crop)
                objs = ndi.find_objects(lbl)

                rows = []
                displayed_points = []
                all_points = []

                overlap_thr = float(inside_frac.value()) / 100.0
                display_min_px = int(display_min_pixels.value())

                n_total = 0
                n_inside = 0
                n_shell_only = 0
                n_displayed = 0

                base_z, _, base_y, _, base_x, _ = crop_bbox
                sp = np.array(_current_spacing(), dtype=np.float32)

                for puncta_id, slc in enumerate(objs, start=1):
                    if slc is None:
                        continue

                    local_mask = (lbl[slc] == puncta_id)
                    if not np.any(local_mask):
                        continue

                    local_coords = np.argwhere(local_mask).astype(np.int32)
                    z0 = int(slc[0].start)
                    y0 = int(slc[1].start)
                    x0 = int(slc[2].start)

                    crop_coords = local_coords + np.array([[z0, y0, x0]], dtype=np.int32)
                    global_coords = crop_coords + np.array([[base_z, base_y, base_x]], dtype=np.int32)

                    pixel_count = int(global_coords.shape[0])
                    centroid_pix = global_coords.mean(axis=0).astype(np.float32)
                    centroid_phys_zyx = centroid_pix * sp

                    intens_vals = vol_crop[
                        crop_coords[:, 0],
                        crop_coords[:, 1],
                        crop_coords[:, 2]
                    ]

                    mean_intensity = _round_int(np.mean(intens_vals)) if intens_vals.size else 0
                    max_intensity = _round_int(np.max(intens_vals)) if intens_vals.size else 0

                    cell_overlap_vals = cell_mask[
                        global_coords[:, 0],
                        global_coords[:, 1],
                        global_coords[:, 2]
                    ]
                    overlap_frac = float(np.mean(cell_overlap_vals)) if cell_overlap_vals.size else 0.0
                    centroid_in = _centroid_inside_mask(cell_mask, centroid_pix)

                    try:
                        surface_dist_um = float(tree.query(centroid_phys_zyx, k=1)[0])
                    except Exception:
                        surface_dist_um = np.nan

                    inside_flag = bool(centroid_in or (overlap_frac >= overlap_thr))
                    near_surface_flag = bool(np.isfinite(surface_dist_um) and (surface_dist_um <= shell_um))
                    keep_flag = bool(inside_flag or near_surface_flag)
                    display_flag = bool(keep_flag and pixel_count >= display_min_px)

                    if inside_flag:
                        classification = "inside_mesh"
                    elif near_surface_flag:
                        classification = "surface_shell"
                    else:
                        classification = "outside"

                    if keep_flag:
                        n_total += 1
                        if inside_flag:
                            n_inside += 1
                        else:
                            n_shell_only += 1
                        if display_flag:
                            n_displayed += 1

                        row = {
                            "channel": channel_name,
                            "role": _channel_role_name(channel_name),
                            "puncta_id": int(puncta_id),
                            "pixel_count": int(pixel_count),
                            "mean_intensity": int(mean_intensity),
                            "max_intensity": int(max_intensity),
                            "centroid_z_um": _round_int(centroid_phys_zyx[0]),
                            "centroid_y_um": _round_int(centroid_phys_zyx[1]),
                            "centroid_x_um": _round_int(centroid_phys_zyx[2]),
                            "centroid_inside_cell_flag": bool(centroid_in),
                            "overlap_fraction_in_cell_pct": _round_int(overlap_frac * 100.0),
                            "surface_distance_um": "" if not np.isfinite(surface_dist_um) else _round_int(surface_dist_um),
                            "inside_mesh_flag": bool(inside_flag),
                            "within_surface_threshold_flag": bool(near_surface_flag),
                            "display_threshold_pass_flag": bool(pixel_count >= display_min_px),
                            "classification": classification,
                            "_centroid_pix_zyx": np.asarray(centroid_pix, dtype=np.float32),
                            "_centroid_phys_zyx_um": np.asarray(centroid_phys_zyx, dtype=np.float32),
                        }

                        if store_coords:
                            row["_bbox_zyx"] = _bbox_from_coords(global_coords)
                            row["_coords_global"] = np.asarray(global_coords, dtype=np.int32)
                            row["_coords_set"] = set(map(tuple, np.asarray(global_coords, dtype=np.int32)))

                        rows.append(row)
                        all_points.append(centroid_pix)
                        if display_flag:
                            displayed_points.append(centroid_pix)

                    del local_mask, local_coords, crop_coords, global_coords, cell_overlap_vals, intens_vals

                if add_layers:
                    _upsert_points_layer(
                        displayed_layer_name,
                        displayed_points,
                        face_color="cyan",
                        size=6,
                        visible=True
                    )
                    _upsert_points_layer(
                        all_layer_name,
                        all_points,
                        face_color="yellow",
                        size=5,
                        visible=False
                    )

                del binary_crop, cell_crop, vol_crop, lbl, objs
                gc.collect()

                return {
                    "channel": channel_name,
                    "rows": rows,
                    "summary": {
                        "total": int(n_total),
                        "inside": int(n_inside),
                        "shell_only": int(n_shell_only),
                        "displayed": int(n_displayed),
                    }
                }

            def _clear_analysis_cache():
                st["analysis_cache"] = {}
                gc.collect()

            def _remove_named_layer(layer_name, store_key=None):
                lyr = None

                if store_key is not None:
                    try:
                        lyr = st.get(store_key, {}).pop(layer_name, None)
                    except Exception:
                        lyr = None

                if lyr is None:
                    try:
                        lyr = pc.layers[layer_name]
                    except Exception:
                        lyr = None

                if lyr is not None:
                    try:
                        pc.layers.remove(lyr)
                    except Exception:
                        pass

            def _upsert_points_layer(layer_name, points_pix, face_color="yellow", edge_color="black", size=6, visible=True):
                pts = np.asarray(points_pix, dtype=np.float32)
                if pts.size == 0:
                    pts = np.zeros((0, 3), dtype=np.float32)

                existing = st["point_layers"].get(layer_name, None)

                def _style_points_layer(lyr):
                    try:
                        lyr.scale = _current_spacing()
                    except Exception:
                        pass
                    try:
                        lyr.face_color = face_color
                    except Exception:
                        pass
                    try:
                        if hasattr(lyr, "border_color"):
                            lyr.border_color = edge_color
                        elif hasattr(lyr, "edge_color"):
                            lyr.edge_color = edge_color
                    except Exception:
                        pass
                    try:
                        lyr.size = size
                    except Exception:
                        pass
                    try:
                        lyr.visible = visible
                    except Exception:
                        pass

                if existing is None:
                    lyr = pc.add_points(
                        pts,
                        name=layer_name,
                        ndim=3,
                        scale=_current_spacing(),
                        face_color=face_color,
                        size=size,
                        visible=visible,
                    )
                    _style_points_layer(lyr)
                    st["point_layers"][layer_name] = lyr
                    return lyr

                try:
                    existing.data = pts
                    _style_points_layer(existing)
                    return existing
                except Exception:
                    try:
                        pc.layers.remove(existing)
                    except Exception:
                        pass

                    lyr = pc.add_points(
                        pts,
                        name=layer_name,
                        ndim=3,
                        scale=_current_spacing(),
                        face_color=face_color,
                        size=size,
                        visible=visible,
                    )
                    _style_points_layer(lyr)
                    st["point_layers"][layer_name] = lyr
                    return lyr


            def _restrict_selected_puncta():
                selected = _get_selected_protein_channels()
                if not selected:
                    QMessageBox.warning(self, "Restrict puncta", "Select at least one protein channel.")
                    return

                if st.get("cell_mask", None) is None:
                    QMessageBox.warning(self, "Restrict puncta", "No cell-of-interest mesh is available.")
                    return

                for ch in st.get("protein_channels", []):
                    _remove_named_layer(f"Displayed puncta centroids - {ch}", "point_layers")
                    _remove_named_layer(f"All puncta centroids - {ch}", "point_layers")

                summaries = []
                cache = {}

                total = len(selected)
                for i, ch in enumerate(selected, start=1):
                    _update_status(f"Restricting {ch} ({i}/{total})...")
                    QApplication.processEvents()

                    try:
                        res = _analyze_puncta_channel(ch, store_coords=False, add_layers=True)
                        cache[ch] = {
                            "rows": res["rows"],
                            "summary": res["summary"],
                            "coords_stored": False,
                        }
                        s = res["summary"]
                        summaries.append(
                            f"{ch}: kept={s['total']}, inside={s['inside']}, shell_only={s['shell_only']}, displayed={s['displayed']}"
                        )
                    except Exception as e:
                        summaries.append(f"{ch}: failed ({e})")

                    QApplication.processEvents()

                st["analysis_cache"] = cache
                _sync_localization_combos()
                _update_status("Restriction complete")

                QMessageBox.information(
                    self,
                    "Restrict puncta",
                    "Restriction complete.\n\n" + "\n".join(summaries)
                )


            def _prompt_export_threshold_choice():
                reply = QMessageBox.question(
                    self,
                    "Export puncta CSV",
                    "Include puncta below the visualization pixel threshold in the CSV?\n\n"
                    "Yes = include all kept puncta.\n"
                    "No = exclude puncta below the current display min pixel count.\n"
                    "Cancel = do not export.",
                    QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                    QMessageBox.Yes,
                )
                if reply == QMessageBox.Cancel:
                    return None
                return bool(reply == QMessageBox.Yes)

            def _export_puncta_csv():
                selected = _get_selected_protein_channels()
                if not selected:
                    QMessageBox.warning(self, "Export puncta CSV", "Select at least one protein channel.")
                    return

                if st.get("cell_mask", None) is None:
                    QMessageBox.warning(self, "Export puncta CSV", "No cell-of-interest mesh is available.")
                    return

                include_below_threshold = _prompt_export_threshold_choice()
                if include_below_threshold is None:
                    return

                default_name = f"{_safe_stem(str(st.get('cell_channel_name', 'puncta')))}_puncta_measurements.csv"
                save_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Save puncta CSV",
                    default_name,
                    "CSV files (*.csv)"
                )
                if not save_path:
                    return
                if not str(save_path).lower().endswith(".csv"):
                    save_path = str(save_path) + ".csv"

                fieldnames = [
                    "cell_channel",
                    "protein_role",
                    "puncta_channel",
                    "puncta_id",
                    "pixel_count",
                    "mean_intensity",
                    "max_intensity",
                    "centroid_x_um",
                    "centroid_y_um",
                    "centroid_z_um",
                    "centroid_inside_cell_flag",
                    "overlap_fraction_in_cell_pct",
                    "surface_distance_um",
                    "inside_mesh_flag",
                    "within_surface_threshold_flag",
                    "display_threshold_pass_flag",
                    "classification",
                ]

                summary_lines = []
                cache = {}
                pixel_filter = int(display_min_pixels.value())

                try:
                    with open(save_path, "w", newline="", encoding="utf-8") as f:
                        writer = csv.DictWriter(f, fieldnames=fieldnames)
                        writer.writeheader()

                        for ch in selected:
                            res = _analyze_puncta_channel(ch, store_coords=False, add_layers=False)
                            cache[ch] = {
                                "rows": res["rows"],
                                "summary": res["summary"],
                                "coords_stored": False,
                            }

                            exported = 0
                            inside = 0
                            shell_only = 0

                            for r in res["rows"]:
                                if (not include_below_threshold) and (int(r["pixel_count"]) < pixel_filter):
                                    continue

                                writer.writerow({
                                    "cell_channel": str(st.get("cell_channel_name", "")),
                                    "protein_role": str(r["role"]),
                                    "puncta_channel": str(r["channel"]),
                                    "puncta_id": int(r["puncta_id"]),
                                    "pixel_count": int(r["pixel_count"]),
                                    "mean_intensity": int(r["mean_intensity"]),
                                    "max_intensity": int(r["max_intensity"]),
                                    "centroid_x_um": int(r["centroid_x_um"]),
                                    "centroid_y_um": int(r["centroid_y_um"]),
                                    "centroid_z_um": int(r["centroid_z_um"]),
                                    "centroid_inside_cell_flag": bool(r["centroid_inside_cell_flag"]),
                                    "overlap_fraction_in_cell_pct": int(r["overlap_fraction_in_cell_pct"]),
                                    "surface_distance_um": r["surface_distance_um"],
                                    "inside_mesh_flag": bool(r["inside_mesh_flag"]),
                                    "within_surface_threshold_flag": bool(r["within_surface_threshold_flag"]),
                                    "display_threshold_pass_flag": bool(r["display_threshold_pass_flag"]),
                                    "classification": str(r["classification"]),
                                })

                                exported += 1
                                if bool(r["inside_mesh_flag"]):
                                    inside += 1
                                elif bool(r["within_surface_threshold_flag"]):
                                    shell_only += 1

                            summary_lines.append(
                                f"{ch}: exported={exported}, inside={inside}, shell_only={shell_only}"
                            )

                    st["analysis_cache"] = cache
                    QMessageBox.information(
                        self,
                        "Export puncta CSV",
                        f"Saved puncta CSV to:\n{save_path}\n\n" + "\n".join(summary_lines)
                    )
                    _update_status(f"Saved puncta CSV: {save_path}")
                except Exception as e:
                    QMessageBox.critical(self, "Export puncta CSV", f"Failed to save CSV:\n{e}")
                    return

            def _ensure_analysis_for_channel(channel_name, store_coords=False):
                cache = st.get("analysis_cache", {})
                cached = cache.get(channel_name, None)

                if cached is not None:
                    if store_coords and not cached.get("coords_stored", False):
                        res = _analyze_puncta_channel(channel_name, store_coords=True, add_layers=False)
                        cache[channel_name] = {
                            "rows": res["rows"],
                            "summary": res["summary"],
                            "coords_stored": True,
                        }
                        st["analysis_cache"] = cache
                        return cache[channel_name]
                    return cached

                res = _analyze_puncta_channel(channel_name, store_coords=store_coords, add_layers=False)
                cache[channel_name] = {
                    "rows": res["rows"],
                    "summary": res["summary"],
                    "coords_stored": bool(store_coords),
                }
                st["analysis_cache"] = cache
                return cache[channel_name]

            def _localize_selected_channels():
                ch_a = str(loc_ch_a.currentText())
                ch_b = str(loc_ch_b.currentText())

                if not ch_a or not ch_b:
                    QMessageBox.warning(self, "Localization", "Select two protein channels.")
                    return
                if ch_a == ch_b:
                    QMessageBox.warning(self, "Localization", "Choose two different channels.")
                    return
                if st.get("cell_mask", None) is None:
                    QMessageBox.warning(self, "Localization", "No cell-of-interest mesh is available.")
                    return

                mode = str(loc_mode.currentText())
                need_coords = (not mode.startswith("Centroid"))

                try:
                    data_a = _ensure_analysis_for_channel(ch_a, store_coords=need_coords)
                    data_b = _ensure_analysis_for_channel(ch_b, store_coords=need_coords)
                except Exception as e:
                    QMessageBox.critical(self, "Localization", f"Failed during puncta analysis:\n{e}")
                    return

                rows_a = list(data_a["rows"])
                rows_b = list(data_b["rows"])

                if not rows_a or not rows_b:
                    QMessageBox.information(self, "Localization", "One or both channels have no puncta.")
                    return

                matched_rows = []

                if mode.startswith("Centroid"):
                    tol = float(loc_centroid_tol_um.value())

                    pts_b = np.array([r["_centroid_phys_zyx_um"] for r in rows_b], dtype=np.float32)
                    tree_b = cKDTree(pts_b)

                    candidates = []
                    for ia, ra in enumerate(rows_a):
                        center_a = np.asarray(ra["_centroid_phys_zyx_um"], dtype=np.float32)
                        idxs = tree_b.query_ball_point(center_a, r=tol)
                        for ib in idxs:
                            rb = rows_b[int(ib)]
                            d = float(np.linalg.norm(center_a - np.asarray(rb["_centroid_phys_zyx_um"], dtype=np.float32)))
                            candidates.append((d, ia, int(ib)))

                    candidates.sort(key=lambda x: x[0])
                    used_a = set()
                    used_b = set()

                    for dist_um, ia, ib in candidates:
                        if ia in used_a or ib in used_b:
                            continue
                        used_a.add(ia)
                        used_b.add(ib)

                        ra = rows_a[ia]
                        rb = rows_b[ib]

                        matched_rows.append({
                            "channel_a": ch_a,
                            "channel_b": ch_b,
                            "puncta_id_a": int(ra["puncta_id"]),
                            "puncta_id_b": int(rb["puncta_id"]),
                            "match_mode": "centroid_distance",
                            "centroid_distance_um": _round_int(dist_um),
                            "voxel_overlap_fraction_min_object_pct": "",
                            "pixel_count_a": int(ra["pixel_count"]),
                            "pixel_count_b": int(rb["pixel_count"]),
                            "mean_intensity_a": int(ra.get("mean_intensity", 0)),
                            "max_intensity_a": int(ra.get("max_intensity", 0)),
                            "mean_intensity_b": int(rb.get("mean_intensity", 0)),
                            "max_intensity_b": int(rb.get("max_intensity", 0)),
                            "centroid_x_a_um": int(ra["centroid_x_um"]),
                            "centroid_y_a_um": int(ra["centroid_y_um"]),
                            "centroid_z_a_um": int(ra["centroid_z_um"]),
                            "centroid_x_b_um": int(rb["centroid_x_um"]),
                            "centroid_y_b_um": int(rb["centroid_y_um"]),
                            "centroid_z_b_um": int(rb["centroid_z_um"]),
                            "inside_mesh_flag_a": bool(ra["inside_mesh_flag"]),
                            "inside_mesh_flag_b": bool(rb["inside_mesh_flag"]),
                            "within_surface_threshold_flag_a": bool(ra["within_surface_threshold_flag"]),
                            "within_surface_threshold_flag_b": bool(rb["within_surface_threshold_flag"]),
                        })

                    del pts_b, tree_b, candidates

                else:
                    overlap_thr = float(loc_voxel_overlap_pct.value()) / 100.0
                    candidates = []

                    for ia, ra in enumerate(rows_a):
                        bbox_a = ra["_bbox_zyx"]
                        set_a = ra["_coords_set"]
                        len_a = max(1, len(set_a))

                        for ib, rb in enumerate(rows_b):
                            bbox_b = rb["_bbox_zyx"]
                            if not _bboxes_intersect(bbox_a, bbox_b):
                                continue

                            set_b = rb["_coords_set"]
                            len_b = max(1, len(set_b))
                            inter = int(len(set_a.intersection(set_b)))
                            if inter <= 0:
                                continue

                            frac = float(inter) / float(min(len_a, len_b))
                            if frac >= overlap_thr:
                                candidates.append((frac, ia, ib))

                    candidates.sort(key=lambda x: x[0], reverse=True)
                    used_a = set()
                    used_b = set()

                    for frac, ia, ib in candidates:
                        if ia in used_a or ib in used_b:
                            continue
                        used_a.add(ia)
                        used_b.add(ib)

                        ra = rows_a[ia]
                        rb = rows_b[ib]
                        center_a = np.asarray(ra["_centroid_phys_zyx_um"], dtype=np.float32)
                        center_b = np.asarray(rb["_centroid_phys_zyx_um"], dtype=np.float32)
                        dist_um = float(np.linalg.norm(center_a - center_b))

                        matched_rows.append({
                            "channel_a": ch_a,
                            "channel_b": ch_b,
                            "puncta_id_a": int(ra["puncta_id"]),
                            "puncta_id_b": int(rb["puncta_id"]),
                            "match_mode": "voxel_overlap",
                            "centroid_distance_um": _round_int(dist_um),
                            "voxel_overlap_fraction_min_object_pct": _round_int(frac * 100.0),
                            "pixel_count_a": int(ra["pixel_count"]),
                            "pixel_count_b": int(rb["pixel_count"]),
                            "mean_intensity_a": int(ra.get("mean_intensity", 0)),
                            "max_intensity_a": int(ra.get("max_intensity", 0)),
                            "mean_intensity_b": int(rb.get("mean_intensity", 0)),
                            "max_intensity_b": int(rb.get("max_intensity", 0)),
                            "centroid_x_a_um": int(ra["centroid_x_um"]),
                            "centroid_y_a_um": int(ra["centroid_y_um"]),
                            "centroid_z_a_um": int(ra["centroid_z_um"]),
                            "centroid_x_b_um": int(rb["centroid_x_um"]),
                            "centroid_y_b_um": int(rb["centroid_y_um"]),
                            "centroid_z_b_um": int(rb["centroid_z_um"]),
                            "inside_mesh_flag_a": bool(ra["inside_mesh_flag"]),
                            "inside_mesh_flag_b": bool(rb["inside_mesh_flag"]),
                            "within_surface_threshold_flag_a": bool(ra["within_surface_threshold_flag"]),
                            "within_surface_threshold_flag_b": bool(rb["within_surface_threshold_flag"]),
                        })

                    del candidates

                if not matched_rows:
                    old_layer_name = f"Localized puncta pairs - {ch_a} vs {ch_b}"
                    try:
                        lyr = st.get("point_layers", {}).pop(old_layer_name, None)
                        if lyr is not None:
                            pc.layers.remove(lyr)
                    except Exception:
                        pass

                    QMessageBox.information(self, "Localization", "No localized puncta matches were found.")
                    return

                save_name = f"{_safe_stem(ch_a)}_vs_{_safe_stem(ch_b)}_localization.csv"
                save_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Save localization CSV",
                    save_name,
                    "CSV files (*.csv)"
                )
                if not save_path:
                    return
                if not str(save_path).lower().endswith(".csv"):
                    save_path = str(save_path) + ".csv"

                csv_saved = False

                try:
                    with open(save_path, "w", newline="", encoding="utf-8") as f:
                        fieldnames = list(matched_rows[0].keys())
                        writer = csv.DictWriter(f, fieldnames=fieldnames)
                        writer.writeheader()
                        writer.writerows(matched_rows)
                    csv_saved = True
                except Exception as e:
                    QMessageBox.critical(
                        self,
                        "Localization",
                        f"Failed to save localization CSV:\n{e}"
                    )
                    return

                try:
                    rows_a_by_id = {int(r["puncta_id"]): r for r in rows_a}
                    rows_b_by_id = {int(r["puncta_id"]): r for r in rows_b}

                    pts_pix = []
                    for m in matched_rows:
                        ra = rows_a_by_id.get(int(m["puncta_id_a"]))
                        rb = rows_b_by_id.get(int(m["puncta_id_b"]))
                        if ra is None or rb is None:
                            continue

                        za = np.asarray(ra["_centroid_pix_zyx"], dtype=np.float32)
                        zb = np.asarray(rb["_centroid_pix_zyx"], dtype=np.float32)
                        pts_pix.append((za + zb) / 2.0)

                    layer_name = f"Localized puncta pairs - {ch_a} vs {ch_b}"

                    try:
                        old_lyr = st.get("point_layers", {}).pop(layer_name, None)
                        if old_lyr is not None:
                            pc.layers.remove(old_lyr)
                    except Exception:
                        pass

                    _upsert_points_layer(
                        layer_name,
                        pts_pix,
                        face_color="white",
                        size=7,
                        visible=True
                    )

                except Exception as e:
                    QMessageBox.warning(
                        self,
                        "Localization",
                        f"Localization CSV was saved, but the viewer points could not be updated:\n{e}"
                    )
                    _update_status(f"Saved localization CSV: {save_path}")
                    gc.collect()
                    return

                QMessageBox.information(
                    self,
                    "Localization",
                    f"Saved localization CSV to:\n{save_path}\n\n"
                    f"Matched pairs: {len(matched_rows)}"
                )
                _update_status(f"Saved localization CSV: {save_path}")
                gc.collect()



            # ============================================================
            # Dock UI
            # ============================================================
            container = QWidget()
            layout = QVBoxLayout(container)
            layout.setContentsMargins(6, 6, 6, 6)
            layout.setSpacing(6)

            status_label = QLabel("Ready")
            layout.addWidget(status_label)

            assigned_lines = []
            for role in ordered_roles:
                if role in accepted_roles:
                    assigned_lines.append(f"{role}: {accepted_roles[role]['channel']}")
                else:
                    assigned_lines.append(f"{role}: None")

            assigned_group = QGroupBox("Assigned roles")
            assigned_layout = QVBoxLayout(assigned_group)
            for line in assigned_lines:
                assigned_layout.addWidget(QLabel(line))
            layout.addWidget(assigned_group)

            voxel_group = QGroupBox("Voxel size (Z, Y, X) µm")
            voxel_layout = QHBoxLayout(voxel_group)

            z_spacing = QDoubleSpinBox()
            z_spacing.setDecimals(4)
            z_spacing.setRange(1e-6, 1e9)
            z_spacing.setValue(float(spacing0[0]))
            voxel_layout.addWidget(QLabel("Z"))
            voxel_layout.addWidget(z_spacing)

            y_spacing = QDoubleSpinBox()
            y_spacing.setDecimals(4)
            y_spacing.setRange(1e-6, 1e9)
            y_spacing.setValue(float(spacing0[1]))
            voxel_layout.addWidget(QLabel("Y"))
            voxel_layout.addWidget(y_spacing)

            x_spacing = QDoubleSpinBox()
            x_spacing.setDecimals(4)
            x_spacing.setRange(1e-6, 1e9)
            x_spacing.setValue(float(spacing0[2]))
            voxel_layout.addWidget(QLabel("X"))
            voxel_layout.addWidget(x_spacing)

            layout.addWidget(voxel_group)

            def _apply_spacing_from_widgets():
                sp = (
                    float(z_spacing.value()),
                    float(y_spacing.value()),
                    float(x_spacing.value()),
                )
                _set_spacing(sp)
                _update_status(f"Voxel size updated to {sp}")

            z_spacing.valueChanged.connect(lambda _=None: _apply_spacing_from_widgets())
            y_spacing.valueChanged.connect(lambda _=None: _apply_spacing_from_widgets())
            x_spacing.valueChanged.connect(lambda _=None: _apply_spacing_from_widgets())

            # ------------------------------------------------------------
            # Cell export
            # ------------------------------------------------------------
            cell_export_group = QGroupBox("Cell-of-interest export")
            cell_export_layout = QVBoxLayout(cell_export_group)

            save_cell_meta_btn = QPushButton("Save meta data from cell of interest")
            export_cell_obj_btn = QPushButton("Export cell of interest OBJ + JSON")

            cell_export_layout.addWidget(save_cell_meta_btn)
            cell_export_layout.addWidget(export_cell_obj_btn)
            layout.addWidget(cell_export_group)

            # ------------------------------------------------------------
            # Protein channels for analysis
            # ------------------------------------------------------------
            protein_group = QGroupBox("Protein channels for analysis")
            protein_layout = QVBoxLayout(protein_group)

            protein_channel_list = QListWidget()
            protein_channel_list.setSelectionMode(QAbstractItemView.NoSelection)

            for role in ["Protein 1", "Protein 2", "Protein 3"]:
                if role not in accepted_roles:
                    continue
                ch = accepted_roles[role]["channel"]
                it = QListWidgetItem(f"{role} | {ch}")
                it.setData(Qt.UserRole, ch)
                it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
                it.setCheckState(Qt.Checked)
                protein_channel_list.addItem(it)

            protein_layout.addWidget(protein_channel_list)
            layout.addWidget(protein_group)

            # ------------------------------------------------------------
            # Restriction / display
            # ------------------------------------------------------------
            restrict_group = QGroupBox("Puncta restriction / display")
            restrict_layout = QVBoxLayout(restrict_group)

            puncta_thr_row = QHBoxLayout()
            puncta_thr_row.addWidget(QLabel("Puncta voxel threshold (> value)"))
            puncta_voxel_threshold = QDoubleSpinBox()
            puncta_voxel_threshold.setDecimals(0)
            puncta_voxel_threshold.setRange(0, 10**12)
            puncta_voxel_threshold.setValue(0)
            puncta_thr_row.addWidget(puncta_voxel_threshold)
            restrict_layout.addLayout(puncta_thr_row)

            inside_row = QHBoxLayout()
            inside_row.addWidget(QLabel("Inside if overlap fraction >= (%)"))
            inside_frac = QDoubleSpinBox()
            inside_frac.setDecimals(0)
            inside_frac.setRange(0, 100)
            inside_frac.setValue(90)
            inside_row.addWidget(inside_frac)
            restrict_layout.addLayout(inside_row)

            shell_row = QHBoxLayout()
            shell_row.addWidget(QLabel("Surface shell threshold (µm)"))
            surface_shell_um = QDoubleSpinBox()
            surface_shell_um.setDecimals(1)
            surface_shell_um.setRange(0, 10**6)
            surface_shell_um.setValue(1)
            shell_row.addWidget(surface_shell_um)
            restrict_layout.addLayout(shell_row)

            display_row = QHBoxLayout()
            display_row.addWidget(QLabel("Display min pixel count (visualization only)"))
            display_min_pixels = QSpinBox()
            display_min_pixels.setRange(0, 10**9)
            display_min_pixels.setValue(0)
            display_row.addWidget(display_min_pixels)
            restrict_layout.addLayout(display_row)

            restrict_btn = QPushButton("Restrict selected puncta channels to cell / surface shell")
            restrict_layout.addWidget(restrict_btn)

            layout.addWidget(restrict_group)

            # ------------------------------------------------------------
            # CSV export
            # ------------------------------------------------------------
            export_group = QGroupBox("Puncta export")
            export_layout = QVBoxLayout(export_group)

            export_btn = QPushButton("Count puncta + export CSV")
            export_layout.addWidget(export_btn)

            layout.addWidget(export_group)

            # ------------------------------------------------------------
            # Localization
            # ------------------------------------------------------------
            loc_group = QGroupBox("Protein-channel localization")
            loc_layout = QVBoxLayout(loc_group)

            loc_row_a = QHBoxLayout()
            loc_row_a.addWidget(QLabel("Channel A"))
            loc_ch_a = QComboBox()
            loc_row_a.addWidget(loc_ch_a)
            loc_layout.addLayout(loc_row_a)

            loc_row_b = QHBoxLayout()
            loc_row_b.addWidget(QLabel("Channel B"))
            loc_ch_b = QComboBox()
            loc_row_b.addWidget(loc_ch_b)
            loc_layout.addLayout(loc_row_b)

            loc_mode_row = QHBoxLayout()
            loc_mode_row.addWidget(QLabel("Localization mode"))
            loc_mode = QComboBox()
            loc_mode.addItems([
                "Centroid distance threshold",
                "Voxel overlap threshold",
            ])
            loc_mode_row.addWidget(loc_mode)
            loc_layout.addLayout(loc_mode_row)

            loc_tol_row = QHBoxLayout()
            loc_tol_row.addWidget(QLabel("Centroid tolerance (µm)"))
            loc_centroid_tol_um = QDoubleSpinBox()
            loc_centroid_tol_um.setDecimals(1)
            loc_centroid_tol_um.setRange(0, 10**6)
            loc_centroid_tol_um.setValue(1)
            loc_tol_row.addWidget(loc_centroid_tol_um)
            loc_layout.addLayout(loc_tol_row)

            loc_overlap_row = QHBoxLayout()
            loc_overlap_row.addWidget(QLabel("Voxel overlap threshold (%)"))
            loc_voxel_overlap_pct = QDoubleSpinBox()
            loc_voxel_overlap_pct.setDecimals(0)
            loc_voxel_overlap_pct.setRange(0, 100)
            loc_voxel_overlap_pct.setValue(90)
            loc_overlap_row.addWidget(loc_voxel_overlap_pct)
            loc_layout.addLayout(loc_overlap_row)

            localize_btn = QPushButton("Analyze localization + export CSV")
            loc_layout.addWidget(localize_btn)

            layout.addWidget(loc_group)

            # ------------------------------------------------------------
            # Utilities
            # ------------------------------------------------------------
            util_group = QGroupBox("Utilities")
            util_layout = QVBoxLayout(util_group)

            clear_cache_btn = QPushButton("Clear analysis cache")
            screenshot_btn = QPushButton("Save screenshot")

            util_layout.addWidget(clear_cache_btn)
            util_layout.addWidget(screenshot_btn)
            layout.addWidget(util_group)

            layout.addStretch(1)
            container.setLayout(layout)

            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setWidget(container)
            pc.window.add_dock_widget(scroll, area="right")

            # ============================================================
            # Signal wiring
            # ============================================================
            save_cell_meta_btn.clicked.connect(lambda _=False: _save_cell_metadata_dialog())
            export_cell_obj_btn.clicked.connect(lambda _=False: _export_cell_mesh_obj())
            restrict_btn.clicked.connect(lambda _=False: _restrict_selected_puncta())
            export_btn.clicked.connect(lambda _=False: _export_puncta_csv())
            localize_btn.clicked.connect(lambda _=False: _localize_selected_channels())
            clear_cache_btn.clicked.connect(lambda _=False: _clear_analysis_cache())
            screenshot_btn.clicked.connect(
                lambda _=False: self.save_screenshot(
                    viewer=pc,
                    parent=self,
                    default_name="puncta_counter.png",
                    canvas_only=True,
                )
            )

            protein_channel_list.itemChanged.connect(lambda *_: _sync_localization_combos())
            loc_mode.currentIndexChanged.connect(lambda _=None: _sync_loc_mode_widgets())

            # ============================================================
            # Initial state sync
            # ============================================================
            _sync_localization_combos()
            _sync_loc_mode_widgets()
            _apply_spacing_from_widgets()

            summary_lines = []
            for role in ordered_roles:
                if role in accepted_roles:
                    summary_lines.append(f"{role}: {accepted_roles[role]['channel']}")
                else:
                    summary_lines.append(f"{role}: None / N/A")

            QMessageBox.information(
                self,
                "Puncta Counter",
                "Puncta counter opened.\n\n"
                f"Voxel size (Z, Y, X): {tuple(float(x) for x in spacing0)}\n\n"
                "Accepted channels:\n"
                + "\n".join(summary_lines)
                + "\n\nSuggested workflow:\n"
                "1. Inspect accepted surfaces in the 3D viewer.\n"
                "2. Select protein channels to analyze.\n"
                "3. Set puncta threshold, overlap %, shell distance, and display minimum.\n"
                "4. Run restriction first to preview centroid layers.\n"
                "5. Export puncta CSV.\n"
                "6. Run localization export if needed."
            )

            return

        except Exception as e:
            try:
                QMessageBox.critical(self, "Error", f"Failed to open puncta counter:\n{e}")
            except Exception:
                pass
            traceback.print_exc()

    def _cell_counter_resolve_infer_weights(self, model_path: str = "") -> str:
        """
        Returns a string to pass to CellposeModel(pretrained_model=...).
        Either "cpsam" or a path to a trained weights file.
        """
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            return "cpsam"

        mode_combo = st.get("ml_mode_combo", None)
        mode_txt = mode_combo.currentText() if mode_combo is not None else ""

        # Mode A: always use cpsam
        if "Pretrained (cpsam)" in (mode_txt or ""):
            return "cpsam"

        # Mode B: custom weights: try last trained file, then meta.json, else ask user
        model_path = (model_path or st.get("ml_model_path", "") or "").strip()
        model_file = (st.get("ml_last_trained_model_file", "") or "").strip()

        meta_file = os.path.join(model_path, "cellpose_model_meta.json") if model_path else ""
        if (not model_file) and meta_file and os.path.exists(meta_file):
            try:
                with open(meta_file, "r") as f:
                    meta = json.load(f)
                model_file = (meta.get("model_out_path", "") or "").strip()
            except Exception:
                model_file = ""

        if model_file and os.path.exists(model_file):
            return model_file

        # Fall back: prompt user to pick weights file
        cc = st.get("viewer", None)
        parent = cc.window._qt_window if (cc is not None and hasattr(cc.window, "_qt_window")) else self
        picked, _ = QFileDialog.getOpenFileName(
            parent,
            "Select Cellpose model weights file",
            model_path or "",
            "All files (*)",
        )
        if picked:
            st["ml_last_trained_model_file"] = picked
            return picked

        # If user cancels, default back to cpsam (safe)
        return "cpsam"

    def _fmt_hms(self, sec: float) -> str:
        sec = int(max(0, sec))
        h = sec // 3600
        m = (sec % 3600) // 60
        s = sec % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _cell_counter_ml_cancel(self):
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            return
        st["ml_cancel_flag"] = True
        # request abort from napari worker too (works only if worker yields/checks abort_requested)
        wk = st.get("ml_worker", None)
        if wk is not None:
            try:
                wk.quit()
            except Exception:
                pass
        self._ml_ui_set_progress("ML progress: cancel requested...", None)

    def _ml_ui_set_progress(self, text: str = None, pct: int = None):
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            return
        lab = st.get("ml_progress_label", None)
        bar = st.get("ml_progress_bar", None)
        if text is not None and lab is not None:
            try:
                lab.setText(text)
            except Exception:
                pass
        if pct is not None and bar is not None:
            try:
                bar.setValue(int(np.clip(pct, 0, 100)))
            except Exception:
                pass
    
    def _cell_counter_ml_train_threaded(self, backend="Cellpose (recommended)", device="auto", model_path=""):
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            QMessageBox.warning(self, "ML", "No cell-counter session found.")
            return

        if st.get("ml_worker", None) is not None:
            QMessageBox.information(self, "ML", "Training is already running.")
            return

        model_path = (model_path or "").strip()
        if not model_path:
            cc = st.get("viewer", None)
            parent = cc.window._qt_window if hasattr(cc.window, "_qt_window") else None  # avoids qt_viewer deprecation
            QMessageBox.warning(parent or self, "ML", "Please choose a Model path before training.")
            return

        # Collect hyperparameters on UI thread (mac-safe)
        cc = st["viewer"]
        model_name, ok = QInputDialog.getText(
            cc.window._qt_window if hasattr(cc.window, "_qt_window") else None,
            "Cellpose model name",
            "Enter a name for the trained model:",
            text=f"cellpose_custom_{time.strftime('%Y%m%d_%H%M%S')}",
        )
        if not ok:
            return
        model_name = (model_name or "").strip() or f"cellpose_custom_{time.strftime('%Y%m%d_%H%M%S')}"

        n_epochs, ok = QInputDialog.getInt(
            cc.window._qt_window if hasattr(cc.window, "_qt_window") else None,
            "Epochs",
            "Number of training epochs:",
            value=50, min=1, max=5000
        )
        if not ok:
            return

        lr, ok = QInputDialog.getDouble(
            cc.window._qt_window if hasattr(cc.window, "_qt_window") else None,
            "Learning rate",
            "Learning rate:",
            value=1e-5, min=1e-7, max=1e-2, decimals=8
        )
        if not ok:
            return

        wd, ok = QInputDialog.getDouble(
            cc.window._qt_window if hasattr(cc.window, "_qt_window") else None,
            "Weight decay",
            "Weight decay:",
            value=0.1, min=0.0, max=10.0, decimals=6
        )
        if not ok:
            return

        st["ml_cancel_flag"] = False
        self._ml_ui_set_progress("ML progress: starting...", 0)
        self._cell_counter_set_status("Backend: cellpose | Starting training...")

        worker = self._cellpose_train_worker(
            backend=backend,
            device=device,
            model_path=model_path,
            model_name=model_name,
            n_epochs=int(n_epochs),
            lr=float(lr),
            wd=float(wd),
            start_model="cpsam",
        )

        worker.yielded.connect(self._on_cellpose_train_yielded)
        worker.returned.connect(self._on_cellpose_train_returned)
        worker.errored.connect(self._on_cellpose_train_errored)
        worker.finished.connect(self._on_cellpose_train_finished)

        st["ml_worker"] = worker
        worker.start()


    @thread_worker(progress={"desc": "Cellpose training"})  # omit total => indeterminate is fine
    def _cellpose_train_worker(
        self,
        backend="Cellpose (recommended)",
        device="auto",
        model_path="",
        model_name="cellpose_custom",
        n_epochs=50,
        lr=1e-5,
        wd=0.1,
        start_model="cpsam",
    ):
        t0 = time.time()
        yield {"pct": 0, "msg": "Validating session..."}

        st = getattr(self, "_cell_counter_state", None)
        if not st:
            raise RuntimeError("No cell-counter session found.")
        if "Cellpose" not in (backend or ""):
            raise RuntimeError("Only Cellpose backend is implemented right now.")

        model_path = (model_path or "").strip()
        if not model_path:
            raise RuntimeError("Model path is empty. Choose a model folder first.")
        os.makedirs(model_path, exist_ok=True)

        # Device
        dev = self._torch_best_device() if (device == "auto") else self._cellpose_device_from_ui(device)

        # Pull data
        vol = np.asarray(st["vol"], dtype=np.float32)  # (Z,Y,X)
        labels_layer = st["labels"]
        lab3d = np.asarray(labels_layer.data, dtype=np.int32)

        # Manual-first: require labels to exist (recommended)
        if int(lab3d.max()) <= 0:
            raise RuntimeError("No labels available. Generate/Edit 'Detected cells (labels mask)' first, then train.")

        yield {"pct": 10, "msg": "Building 2D training set..."}

        train_imgs, train_masks = self._cell_counter_build_cellpose_training_set(
            vol3d=vol,
            labels3d=lab3d,
            max_slices=64,
            min_masks_per_slice=1,
        )
        if not train_imgs:
            raise RuntimeError("No labeled Z-slices found to train on.")

        # Early cancel point (works)
        if getattr(self, "_cell_counter_state", {}).get("ml_cancel_flag", False):
            raise RuntimeError("Training cancelled before model init.")

        yield {"pct": 20, "msg": f"Init Cellpose model ({start_model}) on {dev}..."}

        try:
            cellpose_io.logger_setup()
        except Exception:
            pass

        try:
            import torch
            device_obj = torch.device(dev)
        except Exception:
            device_obj = None

        gpu_flag = (dev == "cuda" or dev == "mps")
        model = models.CellposeModel(gpu=gpu_flag, device=device_obj, pretrained_model=str(start_model))

        yield {"pct": 30, "msg": f"Training... elapsed {self._fmt_hms(time.time()-t0)} (epochs={int(n_epochs)})"}

        # After this point we cannot cancel cleanly until train_seg returns.
        # IMPORTANT CHANGE: force saving into model_path via save_path (supported by Cellpose train API).
        model_out_path, train_losses, test_losses = train.train_seg(
            model.net,
            train_data=train_imgs,
            train_labels=train_masks,
            test_data=None,
            test_labels=None,
            weight_decay=float(wd),
            learning_rate=float(lr),
            n_epochs=int(n_epochs),
            model_name=str(model_name),
            save_path=str(model_path),
        )

        meta = dict(
            backend="cellpose",
            pretrained_start=str(start_model),
            device=str(dev),
            model_name=str(model_name),
            model_out_path=str(model_out_path),
            n_train_images=len(train_imgs),
            n_epochs=int(n_epochs),
            learning_rate=float(lr),
            weight_decay=float(wd),
            created=time.strftime("%Y-%m-%d %H:%M:%S"),
        )
        with open(os.path.join(model_path, "cellpose_model_meta.json"), "w") as f:
            json.dump(meta, f, indent=2)

        yield {"pct": 100, "msg": f"Done in {self._fmt_hms(time.time()-t0)} | saved: {model_out_path}"}
        return dict(
            model_out_path=str(model_out_path),
            model_name=str(model_name),
            meta_path=os.path.join(model_path, "cellpose_model_meta.json"),
        )


    def _on_cellpose_train_yielded(self, payload):
        # payload is dict: {"pct": int, "msg": str}
        try:
            msg = str(payload.get("msg", ""))
            pct = payload.get("pct", None)
        except Exception:
            msg, pct = str(payload), None

        if msg:
            print(f"[Cellpose] {msg}", flush=True)
            self._cell_counter_set_status(msg)
            self._ml_ui_set_progress("ML progress: ", None)

        if pct is not None:
            self._ml_ui_set_progress(None, int(pct))

    def _on_cellpose_train_returned(self, result):
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            return
        try:
            st["ml_last_trained_model_file"] = result.get("model_out_path", "")
            st["ml_model_path"] = os.path.dirname(result.get("meta_path", "")) if result.get("meta_path", "") else st.get("ml_model_path", "")
            st["ml_backend"] = "cellpose"
        except Exception:
            pass

        self._cell_counter_set_status("Backend: cellpose | Training complete")
        self._ml_ui_set_progress("ML progress: complete", 100)

        try:
            cc = st["viewer"]
            QMessageBox.information(
                cc.window.qt_viewer,
                "ML",
                f"Training complete.\nModel: {st.get('ml_last_trained_model_file','')}",
            )
        except Exception:
            pass

    def _on_cellpose_train_errored(self, err):
        self._cell_counter_set_status("Backend: cellpose | Training failed")
        self._ml_ui_set_progress(f"ML progress: error: {err}", 0)

        st = getattr(self, "_cell_counter_state", None)
        try:
            cc = st["viewer"] if st else None
            parent = cc.window.qt_viewer if cc is not None else self
            QMessageBox.critical(parent, "ML", f"Cellpose training failed:\n{err}")
        except Exception:
            pass

    def _on_cellpose_train_finished(self):
        st = getattr(self, "_cell_counter_state", None)
        if st:
            st["ml_worker"] = None


    def _cell_counter_ml_load(self, backend="Cellpose (recommended)", device="auto", model_path=""):
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            return
        if device == "auto":
            device = self._torch_best_device()
        st["ml_device"] = device
        st["ml_model_path"] = model_path
        st["ml_backend"] = "cellpose" if "Cellpose" in backend else "stardist" if "StarDist" in backend else "custom"
        self._cell_counter_set_status(f"Device: {device} | Backend: {st['ml_backend']} | Model: {model_path or '(none)'}")

    def _cell_counter_ml_save(self, backend="Cellpose (recommended)", device="auto", model_path=""):
        QMessageBox.information(self, "ML", "Save is wired, but training/inference implementation is not added yet.")

    def _cell_counter_ml_train(self, backend="Cellpose (recommended)", device="auto", model_path=""):
        QMessageBox.information(self, "ML", "Train is wired, but training implementation is not added yet.")

    def _cell_counter_ml_infer(self, backend="Cellpose (recommended)", device="auto", model_path=""):
        try:
            st = getattr(self, "_cell_counter_state", None)
            if not st:
                QMessageBox.warning(self, "ML", "No cell-counter session found.")
                return

            if "Cellpose" not in (backend or ""):
                QMessageBox.warning(self, "ML", "Only Cellpose backend is implemented right now.")
                return

            cc = st["viewer"]
            vol = np.asarray(st["vol"], dtype=np.float32)  # (Z,Y,X) expected
            labels_layer = st["labels"]

            dev = self._cellpose_device_from_ui(device)
            st["ml_device"] = dev
            st["ml_backend"] = "cellpose"

            if not model_path:
                model_path = st.get("ml_model_path", "") or ""
            if model_path:
                st["ml_model_path"] = model_path

            # Resolve weights based on Model mode (your helper decides custom path vs "cpsam")
            weights = self._cell_counter_resolve_infer_weights(model_path=model_path)

            # Detect cpsam mode (string sentinel or filename containing it)
            weights_str = str(weights or "")
            is_cpsam = (weights == "cpsam") or ("cpsam" in weights_str.lower())

            self._cell_counter_set_status(
                f"Device: {dev} | Backend: cellpose | Infer ({'cpsam' if is_cpsam else 'custom'})"
            )

            # Build model
            try:
                import torch
                device_obj = torch.device(dev)
            except Exception:
                device_obj = None
            gpu_flag = (dev == "cuda" or dev == "mps")

            model = models.CellposeModel(gpu=gpu_flag, device=device_obj, pretrained_model=weights)

            # Ask for key eval params
            diameter, ok = QInputDialog.getDouble(
                cc.window.qt_viewer,
                "Cellpose diameter",
                "Approx cell diameter in pixels (0 = auto):",
                value=0.0,
                min=0.0,
                max=10000.0,
                decimals=2,
            )
            if not ok:
                return
            diameter_val = None if float(diameter) <= 0 else float(diameter)

            # 2D-per-slice inference
            Z = int(vol.shape[0])
            out = np.zeros_like(vol, dtype=np.int32)
            next_id = 0

            # Classic Cellpose uses channels; Cellpose-SAM (cpsam) does not 
            channels = [0, 0]

            for z in range(Z):
                img2 = np.asarray(vol[z], dtype=np.float32)  # (Y,X)

                # Make sure we provide a shape cpsam can use: first 3 channels are used 
                if is_cpsam:
                    # Convert grayscale (Y,X) -> (Y,X,3)
                    if img2.ndim == 2:
                        img_in = np.stack([img2, img2, img2], axis=-1)
                    elif img2.ndim == 3 and img2.shape[-1] >= 3:
                        img_in = img2[..., :3]
                    elif img2.ndim == 3 and img2.shape[0] >= 3:
                        # if someone stored channels-first, keep first 3 channels
                        img_in = np.moveaxis(img2[:3, ...], 0, -1)
                    else:
                        img_in = img2

                    # cpsam commonly returns 3 outputs (masks, flows, styles) 
                    eval_out = model.eval(img_in, diameter=diameter_val, do_3D=False)
                    if isinstance(eval_out, (tuple, list)) and len(eval_out) == 3:
                        masks, flows, styles = eval_out
                        diams = None
                    else:
                        masks, flows, styles, diams = eval_out
                else:
                    masks, flows, styles, diams = model.eval(
                        img2,
                        channels=channels,
                        diameter=diameter_val,
                        do_3D=False,
                    )

                masks = np.asarray(masks, dtype=np.int32)
                if masks.size == 0 or int(masks.max()) == 0:
                    continue

                unique_ids = np.unique(masks)
                unique_ids = unique_ids[unique_ids != 0]
                if unique_ids.size:
                    remap = {int(uid): (next_id + i + 1) for i, uid in enumerate(unique_ids)}
                    next_id += int(unique_ids.size)
                    m2 = np.zeros_like(masks, dtype=np.int32)
                    for uid, nid in remap.items():
                        m2[masks == uid] = nid
                    out[z] = m2

            labels_layer.data = out
            n_cells = int(out.max())

            # Track last-used weights (for custom mode, selected file; for cpsam it's "cpsam")
            st["ml_last_trained_model_file"] = weights

            self._cell_counter_set_status(f"Device: {dev} | Backend: cellpose | Done ({n_cells} labels)")
            QMessageBox.information(
                cc.window.qt_viewer,
                "ML",
                f"Cellpose inference complete.\nDetected (slice-stacked) objects: {n_cells}",
            )

        except Exception as e:
            try:
                self._cell_counter_set_status("ML inference failed")
            except Exception:
                pass
            QMessageBox.critical(self, "Error", f"Cellpose inference failed:\n{e}")


    def _torch_best_device(self):
        """Return a torch.device string: 'cuda' | 'mps' | 'cpu'."""
        try:
            if torch.cuda.is_available():
                return "cuda"
            mps_ok = hasattr(torch.backends, "mps") and torch.backends.mps.is_available()
            if mps_ok:
                return "mps"
            return "cpu"
        except Exception:
            return "cpu"
    
    def _cellpose_device_from_ui(self, device_choice: str):
        """
        Map UI 'auto|cuda|mps|cpu' to a torch.device-like string for Cellpose.
        Cellpose accepts either gpu=True or an explicit device in newer APIs;
        we keep it simple and pass a device string into CellposeModel where possible.
        """
        dc = (device_choice or "auto").strip().lower()
        if dc == "auto":
            return self._torch_best_device()
        if dc in ("cuda", "mps", "cpu"):
            return dc
        return self._torch_best_device()
    
    def _cell_counter_build_cellpose_training_set(self, vol3d: np.ndarray, labels3d: np.ndarray,
                                             max_slices: int = 64,
                                             min_masks_per_slice: int = 1):
        """
        Build 2D training data for Cellpose from a 3D stack.
        Returns (train_images, train_masks) where each is a list of 2D arrays.

        Strategy:
        - Use Z slices where labels3d has at least `min_masks_per_slice` instances.
        - Subsample up to `max_slices` slices to keep training time reasonable.
        """
        assert vol3d.ndim == 3 and labels3d.ndim == 3
        Z = vol3d.shape[0]

        slice_ids = []
        for z in range(Z):
            lab2 = labels3d[z]
            n_inst = int(lab2.max())
            if n_inst >= int(min_masks_per_slice):
                slice_ids.append(z)

        if not slice_ids:
            return [], []

        # Subsample evenly if too many
        if len(slice_ids) > int(max_slices):
            idx = np.linspace(0, len(slice_ids) - 1, int(max_slices)).round().astype(int)
            slice_ids = [slice_ids[i] for i in idx]

        imgs = []
        masks = []
        for z in slice_ids:
            img2 = np.asarray(vol3d[z], dtype=np.float32)
            m2 = np.asarray(labels3d[z], dtype=np.int32)
            imgs.append(img2)
            masks.append(m2)

        return imgs, masks

    def _cell_counter_ml_train(self, backend="Cellpose (recommended)", device="auto", model_path=""):
        try:
            st = getattr(self, "_cell_counter_state", None)
            if not st:
                QMessageBox.warning(self, "ML", "No cell-counter session found.")
                return

            if "Cellpose" not in (backend or ""):
                QMessageBox.warning(self, "ML", "Only Cellpose backend is implemented right now.")
                return

            cc = st["viewer"]
            vol = np.asarray(st["vol"], dtype=np.float32)  # (Z,Y,X)
            labels_layer = st["labels"]

            # Choose / create model output directory
            if not model_path:
                model_path = QFileDialog.getExistingDirectory(cc.window.qt_viewer, "Select folder to save model")
                if not model_path:
                    return
            os.makedirs(model_path, exist_ok=True)

            # Determine device
            dev = self._cellpose_device_from_ui(device)
            st["ml_device"] = dev
            st["ml_backend"] = "cellpose"
            st["ml_model_path"] = model_path
            self._cell_counter_set_status(f"Device: {dev} | Backend: cellpose | Training...")

            # Ensure we have labels to train on: use existing labels if present,
            # otherwise generate pseudo-labels via classic detection first.
            lab3d = np.asarray(labels_layer.data, dtype=np.int32)
            if int(lab3d.max()) <= 0:
                # Run classic detection with conservative defaults (you can tune these)
                self._cell_counter_run_detection(
                    threshold=0.50,
                    min_voxels=200,
                    max_voxels=0,
                    compactness=0.0,
                    neg_grow=2,
                    gaussian_sigma=1.0,
                    open_radius=1,
                    auto_split=True,
                    min_seed_distance=8,
                )
                lab3d = np.asarray(labels_layer.data, dtype=np.int32)

            if int(lab3d.max()) <= 0:
                QMessageBox.warning(cc.window.qt_viewer, "ML", "No labels available for training (Detected cells is empty).")
                self._cell_counter_set_status(f"Device: {dev} | Backend: cellpose | Training failed")
                return

            # Build 2D training set from labeled slices
            train_imgs, train_masks = self._cell_counter_build_cellpose_training_set(
                vol3d=vol,
                labels3d=lab3d,
                max_slices=64,
                min_masks_per_slice=1,
            )
            if not train_imgs:
                QMessageBox.warning(cc.window.qt_viewer, "ML", "No labeled Z-slices found to train on.")
                self._cell_counter_set_status(f"Device: {dev} | Backend: cellpose | Training failed")
                return

            # Ask for training hyperparameters
            model_name, ok = QInputDialog.getText(
                cc.window.qt_viewer,
                "Cellpose model name",
                "Enter a name for the trained model:",
                text=f"cellpose_custom_{time.strftime('%Y%m%d_%H%M%S')}",
            )
            if not ok:
                return
            model_name = (model_name or "").strip() or f"cellpose_custom_{time.strftime('%Y%m%d_%H%M%S')}"

            n_epochs, ok = QInputDialog.getInt(
                cc.window.qt_viewer, "Epochs", "Number of training epochs:", value=50, min=1, max=5000
            )
            if not ok:
                return

            lr, ok = QInputDialog.getDouble(
                cc.window.qt_viewer, "Learning rate", "Learning rate:", value=1e-5, min=1e-7, max=1e-2, decimals=8
            )
            if not ok:
                return

            wd, ok = QInputDialog.getDouble(
                cc.window.qt_viewer, "Weight decay", "Weight decay:", value=0.1, min=0.0, max=10.0, decimals=6
            )
            if not ok:
                return

            # Cellpose logger setup (optional)
            try:
                cellpose_io.logger_setup()
            except Exception:
                pass

            # Create a Cellpose model starting from built-in pretrained (recommended by Cellpose docs). 
            # We try to pass an explicit device if available; otherwise rely on gpu flag.
            # Cellpose's internal device assignment supports CUDA and checks MPS availability. 
            try:
                import torch
                device_obj = torch.device(dev)
            except Exception:
                device_obj = None

            gpu_flag = (dev == "cuda" or dev == "mps")
            model = models.CellposeModel(gpu=gpu_flag, device=device_obj, pretrained_model="cpsam")

            # Train and save model
            # train.train_seg returns model_path and losses per docs. 
            model_out_path, train_losses, test_losses = train.train_seg(
                model.net,
                train_data=train_imgs,
                train_labels=train_masks,
                test_data=None,
                test_labels=None,
                weight_decay=float(wd),
                learning_rate=float(lr),
                n_epochs=int(n_epochs),
                model_name=str(model_name),
            )

            # Record where it saved + metadata
            meta = dict(
                backend="cellpose",
                pretrained_start="cpsam",
                device=dev,
                model_name=model_name,
                model_out_path=str(model_out_path),
                n_train_images=len(train_imgs),
                n_epochs=int(n_epochs),
                learning_rate=float(lr),
                weight_decay=float(wd),
                created=time.strftime("%Y-%m-%d %H:%M:%S"),
            )
            with open(os.path.join(model_path, "cellpose_model_meta.json"), "w") as f:
                json.dump(meta, f, indent=2)

            st["ml_model_path"] = model_path
            st["ml_backend"] = "cellpose"
            st["ml_last_trained_model_file"] = str(model_out_path)

            self._cell_counter_set_status(f"Device: {dev} | Backend: cellpose | Trained: {model_name}")
            QMessageBox.information(
                cc.window.qt_viewer,
                "ML",
                f"Training complete.\nModel name: {model_name}\nSaved model file:\n{model_out_path}\nMeta in:\n{os.path.join(model_path, 'cellpose_model_meta.json')}",
            )

        except Exception as e:
            try:
                self._cell_counter_set_status("ML training failed")
            except Exception:
                pass
            QMessageBox.critical(self, "Error", f"Cellpose training failed:\n{e}")

    def _cell_counter_ml_save(self, backend="Cellpose (recommended)", device="auto", model_path=""):
        try:
            st = getattr(self, "_cell_counter_state", None)
            if not st:
                return

            cc = st["viewer"]
            if "Cellpose" not in (backend or ""):
                QMessageBox.warning(self, "ML", "Only Cellpose backend is implemented right now.")
                return

            model_file = st.get("ml_last_trained_model_file", "")
            if not model_file or not os.path.exists(model_file):
                QMessageBox.information(cc.window.qt_viewer, "ML", "No trained model file found to save. Train first.")
                return

            if not model_path:
                model_path = QFileDialog.getExistingDirectory(cc.window.qt_viewer, "Select folder to save/copy model")
                if not model_path:
                    return
            os.makedirs(model_path, exist_ok=True)

            # Copy model file into folder
            dst = os.path.join(model_path, os.path.basename(model_file))
            if os.path.abspath(dst) != os.path.abspath(model_file):
                shutil.copy2(model_file, dst)

            # Also copy meta if present from current ml_model_path
            src_meta = ""
            if st.get("ml_model_path", ""):
                cand = os.path.join(st["ml_model_path"], "cellpose_model_meta.json")
                if os.path.exists(cand):
                    src_meta = cand
            if src_meta:
                shutil.copy2(src_meta, os.path.join(model_path, "cellpose_model_meta.json"))

            st["ml_model_path"] = model_path
            self._cell_counter_set_status(f"Device: {st.get('ml_device','?')} | Backend: cellpose | Model saved")
            QMessageBox.information(cc.window.qt_viewer, "ML", f"Saved/copied model to:\n{model_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Save model failed:\n{e}")

    def _cell_counter_set_status(self, text: str):
        st = getattr(self, "_cell_counter_state", None)
        if not st:
            return
        lab = st.get("status_label", None)
        if lab is not None:
            try:
                lab.setText(text)
            except Exception:
                pass

    def _cell_counter_run_detection(
        self,
        threshold=0.5,
        min_voxels=200,
        max_voxels=0,
        compactness=0.0,
        neg_grow=2,
        gaussian_sigma=1.0,
        open_radius=1,
        auto_split=True,
        min_seed_distance=8,
    ):
        """
        Improved seeded segmentation:

        - Foreground from threshold on normalized volume
        - Optional smoothing
        - Optional soma-only cleanup (3D opening + remove_small_objects)
        - Positive ROIs -> markers (seeds)
        - Negative ROIs -> exclusion mask
        - Optional auto-split markers from distance peaks
        - Watershed on distance transform
        - Size filtering
        """
        try:
            st = getattr(self, "_cell_counter_state", None)
            if not st:
                QMessageBox.warning(self, "Cell Counter", "No cell-counter session found.")
                return

            cc = st["viewer"]
            vol = np.asarray(st["vol"], dtype=np.float32)  # (Z,Y,X)
            pos_shapes = st["pos_shapes"]
            neg_shapes = st["neg_shapes"]
            labels_layer = st["labels"]

            # Normalize volume to 0..1 for thresholding
            vmin, vmax = np.percentile(vol, (1, 99))
            if vmax - vmin <= 1e-6:
                norm = np.clip(vol, 0, 1)
            else:
                norm = np.clip((vol - vmin) / (vmax - vmin), 0, 1)

            Z, Y, X = norm.shape

            def _rasterize_shapes_to_masks(shapes_layer):
                out = []
                if shapes_layer is None or len(shapes_layer.data) == 0:
                    return out

                for poly in shapes_layer.data:
                    if poly is None:
                        continue
                    pts = np.asarray(poly)
                    if pts.ndim != 2 or pts.shape[0] < 3:
                        continue

                    if pts.shape[1] == 3:
                        zz = pts[:, 0]
                        yy = pts[:, 1]
                        xx = pts[:, 2]
                        z = int(np.clip(np.round(np.median(zz)), 0, Z - 1))
                    elif pts.shape[1] == 2:
                        try:
                            z = int(cc.dims.point[0]) if hasattr(cc, "dims") and len(cc.dims.point) > 0 else 0
                        except Exception:
                            z = 0
                        z = int(np.clip(z, 0, Z - 1))
                        yy = pts[:, 0]
                        xx = pts[:, 1]
                    else:
                        continue

                    rr, cc_ = sk_polygon(yy, xx, shape=(Y, X))
                    m = np.zeros((Y, X), dtype=bool)
                    m[rr, cc_] = True
                    out.append((z, m))
                return out

            # Seeds from positive ROIs
            seeds = np.zeros((Z, Y, X), dtype=np.int32)
            pos = _rasterize_shapes_to_masks(pos_shapes)
            if len(pos) == 0:
                QMessageBox.information(cc.window.qt_viewer, "Cell Counter", "Draw at least one GREEN cell-body ROI first.")
                return

            seed_id = 0
            for z, m2d in pos:
                seed_id += 1
                seeds[z, m2d] = seed_id

            # Negative mask (with grow)
            neg_mask = np.zeros((Z, Y, X), dtype=bool)
            neg = _rasterize_shapes_to_masks(neg_shapes)
            if len(neg) > 0:
                se2 = disk(int(neg_grow)) if int(neg_grow) > 0 else None
                for z, m2d in neg:
                    if se2 is not None:
                        m2d = dilation(m2d, se2)
                    neg_mask[z, m2d] = True

            # Foreground mask from threshold (optionally after smoothing)
            img = norm
            if float(gaussian_sigma) > 0:
                img = gaussian(img, sigma=float(gaussian_sigma), preserve_range=True)

            fg = (img >= float(threshold))
            if neg_mask.any():
                fg = fg & (~neg_mask)

            # Soma-only cleanup: opening removes thin processes 
            r = int(open_radius)
            if r > 0:
                fg = binary_opening(fg, footprint=ball(r))

            # Remove small junk objects (3D) 
            fg = remove_small_objects(fg, min_size=int(min_voxels), connectivity=1)

            if not fg.any():
                labels_layer.data = np.zeros_like(seeds, dtype=np.int32)
                QMessageBox.information(cc.window.qt_viewer, "Cell Counter", "No foreground after threshold/cleanup.")
                return

            # Watershed on distance transform with markers 
            dist = ndi.distance_transform_edt(fg)
            markers = seeds.copy()

            # Optional auto-split: add distance peaks as extra markers 
            if bool(auto_split):
                coords = peak_local_max(
                    dist,
                    min_distance=int(max(1, min_seed_distance)),
                    labels=fg,
                    exclude_border=False,
                )
                for (zz, yy, xx) in coords:
                    if markers[zz, yy, xx] == 0:
                        seed_id += 1
                        markers[zz, yy, xx] = seed_id

            seg = watershed(-dist, markers=markers, mask=fg, compactness=float(compactness))

            # Size filtering
            counts = np.bincount(seg.ravel())
            out = np.zeros_like(seg, dtype=np.int32)
            new_id = 0
            for old_id in range(1, int(seg.max()) + 1):
                n = int(counts[old_id]) if old_id < len(counts) else 0
                if n < int(min_voxels):
                    continue
                if int(max_voxels) > 0 and n > int(max_voxels):
                    continue
                new_id += 1
                out[seg == old_id] = new_id

            labels_layer.data = out
            n_cells = int(new_id)
            print(f"Cell counter: {n_cells} cells detected")
            QMessageBox.information(cc.window.qt_viewer, "Cell Counter", f"Detected {n_cells} cell bodies.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cell detection failed:\n{e}")

    def calculate_intensities_per_cell(self):
        """
        Export a CSV with per-cell fluorescence intensity stats from a chosen 3D Image layer
        in the Cell Counter viewer or the main viewer (one channel at a time):
        - Cell # (1..N)
        - pixel_count (number of voxels/pixels in the label, including zeros in intensity)
        - min (excluding zeros)
        - max (excluding zeros)
        - average (mean, excluding zeros)
        """
        try:
            st = getattr(self, "_cell_counter_state", None)
            if not st:
                QMessageBox.warning(self, "Cell Counter", "No cell-counter session found.")
                return

            cc = st["viewer"]          # cell counter viewer
            labels_layer = st["labels"]
            lab = np.asarray(labels_layer.data)  # 0=background

            if lab.ndim != 3:
                QMessageBox.warning(
                    cc.window.qt_viewer,
                    "Cell Counter",
                    "Detected cells layer must be 3D.",
                )
                return

            n_cells = int(lab.max())
            if n_cells <= 0:
                QMessageBox.information(
                    cc.window.qt_viewer,
                    "Cell Counter",
                    "No detected cells to export.",
                )
                return

            # ---------------------------------------------------------
            # 0) Choose where to take intensities from
            # ---------------------------------------------------------
            source_choice, ok = QInputDialog.getItem(
                cc.window.qt_viewer,
                "Intensity source",
                "Select where to get intensities:",
                ["Cell Counter viewer", "Main viewer"],
                0,
                False,
            )
            if not ok or not source_choice:
                return

            # Helper to collect candidates from a given viewer
            def _collect_candidates(viewer, labels_shape, labels_layer_to_exclude=None):
                cands = []
                for lyr in viewer.layers:
                    if not hasattr(lyr, "data"):
                        continue
                    if lyr is labels_layer_to_exclude or getattr(lyr, "data", None) is None:
                        continue
                    # skip labels-type layers
                    if lyr.__class__.__name__.lower() == "labels":
                        continue
                    try:
                        arr = np.asarray(lyr.data)
                    except Exception:
                        continue
                    if arr.ndim != 3:
                        continue
                    if arr.shape != labels_shape:
                        continue
                    cands.append((getattr(lyr, "name", "image"), lyr))
                return cands

            # ---------------------------------------------------------
            # 1) Collect candidate intensity layers
            # ---------------------------------------------------------
            if source_choice == "Cell Counter viewer":
                intensity_viewer = cc
            else:
                # assume you stored the main viewer on self, adapt if needed
                main_viewer = getattr(self, "viewer", None)
                if main_viewer is None:
                    QMessageBox.warning(
                        cc.window.qt_viewer,
                        "Cell Counter",
                        "Main viewer not available on this plugin instance.",
                    )
                    return
                intensity_viewer = main_viewer

            candidates = _collect_candidates(
                intensity_viewer,
                labels_shape=lab.shape,
                labels_layer_to_exclude=labels_layer if intensity_viewer is cc else None,
            )

            if not candidates:
                QMessageBox.warning(
                    cc.window.qt_viewer,
                    "Cell Counter",
                    "No valid 3D Image layers found that match the labels shape.\n"
                    f"Labels shape: {lab.shape}\n\n"
                    "Tip: Ensure the fluorescence volume you want to measure is present\n"
                    "as a 3D Image layer with the same (Z,Y,X) shape.",
                )
                return

            names = [nm for nm, _ in candidates]
            default_index = names.index("Detection input") if "Detection input" in names else 0

            chosen_name, ok = QInputDialog.getItem(
                cc.window.qt_viewer,
                "Choose intensity layer",
                "Select the 3D Image layer to measure intensities from:",
                names,
                default_index,
                False,
            )
            if not ok or not chosen_name:
                return

            chosen_layer = dict(candidates)[chosen_name]
            vol = np.asarray(chosen_layer.data, dtype=np.float32)

            if vol.ndim != 3 or vol.shape != lab.shape:
                QMessageBox.warning(
                    cc.window.qt_viewer,
                    "Cell Counter",
                    f"Selected layer does not match labels shape.\n"
                    f"Image shape: {vol.shape}, Labels shape: {lab.shape}",
                )
                return

            # ---------------------------------------------------------
            # 1b) If intensity comes from main viewer, apply mask
            # ---------------------------------------------------------
            if source_choice == "Main viewer":
                # binary mask (0/1) from labels (non-zero voxels are 1)
                mask = (lab > 0).astype(np.float32)
                # multiply mask into the chosen intensity channel
                vol = vol * mask

            # ---------------------------------------------------------
            # 2) Choose output file
            # ---------------------------------------------------------
            base, ok = QInputDialog.getText(
                cc.window.qt_viewer,
                "CSV name",
                "Enter output CSV filename (without extension):",
                text="cell_intensities",
            )
            if not ok:
                return
            base = (base or "").strip() or "cell_intensities"
            default_name = base if base.lower().endswith(".csv") else (base + ".csv")

            save_path, _ = QFileDialog.getSaveFileName(
                cc.window.qt_viewer,
                "Save cell intensities CSV",
                default_name,
                "CSV files (*.csv)",
            )
            if not save_path:
                return
            if not save_path.lower().endswith(".csv"):
                save_path += ".csv"

            # ---------------------------------------------------------
            # 3) Compute per-cell stats + pixel counts
            # ---------------------------------------------------------
            counts = np.bincount(lab.ravel().astype(np.int64), minlength=n_cells + 1)

            rows = []
            for cell_id in range(1, n_cells + 1):
                pixel_count = int(counts[cell_id])  # labeled voxels for this cell
                if pixel_count <= 0:
                    continue

                m = (lab == cell_id)
                vals = vol[m]
                vals = vals[vals > 0]  # exclude 0 background

                if vals.size == 0:
                    min_v = float("nan")
                    max_v = float("nan")
                    mean_v = float("nan")
                else:
                    min_v = float(vals.min())
                    max_v = float(vals.max())
                    mean_v = float(vals.mean())

                rows.append((cell_id, pixel_count, min_v, max_v, mean_v))

            if not rows:
                QMessageBox.information(
                    cc.window.qt_viewer,
                    "Cell Counter",
                    "No cells found to export.",
                )
                return

            with open(save_path, "w", newline="") as f:
                w = csv.writer(f)
                w.writerow(["Cell #", "pixel_count", "min", "max", "average"])
                w.writerows(rows)

            print(
                f"Exported intensities for {len(rows)} cells from layer "
                f"'{chosen_name}' (source: {source_choice}) to: {save_path}"
            )
            QMessageBox.information(
                cc.window.qt_viewer,
                "Cell Counter",
                f"Exported {len(rows)} cells.\n"
                f"Layer: {chosen_name}\n"
                f"Source: {source_choice}\n"
                f"Saved to:\n{save_path}",
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Intensity export failed:\n{e}")

    def open_preprocess(self):
        self_parent = self
        dlg = QDialog(self)
        dlg.setWindowTitle("Pre-process Analysis")

        v = QVBoxLayout(dlg)

        def _update_z_label():
            if not hasattr(self_parent, "z_range_label"):
                return
            end = max(self_parent.n_z - 1, 0)

            # Optional: show crop provenance (original indices)
            if hasattr(self_parent, "last_crop_range") and self_parent.last_crop_range is not None:
                a, b = self_parent.last_crop_range
                self_parent.z_range_label.setText(
                    f"Z: {self_parent.current_z}   (range 0–{end})   |   cropped from {a}–{b}"
                )
            else:
                self_parent.z_range_label.setText(f"Z: {self_parent.current_z}   (range 0–{end})")

        def open_crop_projection_viewer(
                self,
                image_layer_names=None,
                fill_value=0,
                tight_bbox=True,
                visible_only=True,
            ):
            import numpy as np
            import napari
            from napari.layers import Image
            from qtpy.QtCore import QTimer
            from qtpy.QtWidgets import (
                QWidget,
                QVBoxLayout,
                QPushButton,
                QLabel,
                QMessageBox,
            )
            from skimage.draw import polygon2mask

            main_viewer = self.viewer

            def _qt_parent():
                try:
                    return main_viewer.window._qt_window
                except Exception:
                    return None

            def _normalize_names(names):
                if names is None:
                    return None
                if isinstance(names, str):
                    return [names]
                return [str(x) for x in names]

            def _as_nd_tuple(value, ndim, default):
                if value is None:
                    return (default,) * ndim
                try:
                    if np.isscalar(value):
                        return (value,) * ndim
                except Exception:
                    pass
                try:
                    seq = tuple(value)
                except Exception:
                    return (default,) * ndim
                if len(seq) == 0:
                    return (default,) * ndim
                if len(seq) >= ndim:
                    return tuple(seq[-ndim:])
                return (default,) * (ndim - len(seq)) + tuple(seq)

            def _allclose(a, b, atol=1e-9, rtol=1e-7):
                try:
                    return np.allclose(np.asarray(a, dtype=float), np.asarray(b, dtype=float), atol=atol, rtol=rtol)
                except Exception:
                    return False

            def _rotation_is_identity(rot):
                if rot is None:
                    return True
                r = np.asarray(rot, dtype=float)
                if r.size == 0:
                    return True
                if r.ndim == 0:
                    return np.allclose(float(r), 0.0)
                if r.ndim == 1:
                    return np.allclose(r, 0.0)
                if r.ndim == 2 and r.shape[0] == r.shape[1]:
                    return np.allclose(r, np.eye(r.shape[0]), atol=1e-9, rtol=1e-7)
                return False

            def _shear_is_identity(shear):
                if shear is None:
                    return True
                s = np.asarray(shear, dtype=float)
                if s.size == 0:
                    return True
                if s.ndim <= 1:
                    return np.allclose(s, 0.0)
                if s.ndim == 2 and s.shape[0] == s.shape[1]:
                    return np.allclose(s, np.eye(s.shape[0]), atol=1e-9, rtol=1e-7)
                return False

            def _affine_is_identity(layer):
                aff = getattr(layer, "affine", None)
                if aff is None:
                    return True

                try:
                    linear = getattr(aff, "linear_matrix", None)
                except Exception:
                    linear = None

                try:
                    extra_translate = getattr(aff, "translate", None)
                except Exception:
                    extra_translate = None

                if linear is not None:
                    lin = np.asarray(linear, dtype=float)
                    if lin.size:
                        if lin.ndim != 2 or lin.shape[0] != lin.shape[1]:
                            return False
                        if not np.allclose(lin, np.eye(lin.shape[0]), atol=1e-9, rtol=1e-7):
                            return False

                if extra_translate is not None:
                    tr = np.asarray(extra_translate, dtype=float)
                    if tr.size and not np.allclose(tr, 0.0, atol=1e-9, rtol=1e-7):
                        return False

                return True

            def _has_unsupported_transform(layer):
                try:
                    if not _rotation_is_identity(getattr(layer, "rotate", None)):
                        return True, "rotation"
                except Exception:
                    return True, "rotation"

                try:
                    if not _shear_is_identity(getattr(layer, "shear", None)):
                        return True, "shear"
                except Exception:
                    return True, "shear"

                try:
                    if not _affine_is_identity(layer):
                        return True, "non-identity affine"
                except Exception:
                    return True, "non-identity affine"

                return False, None

            def _get_image_layers(names=None, visible_only=True):
                names = _normalize_names(names)
                all_layers = [lyr for lyr in main_viewer.layers if isinstance(lyr, Image)]

                if names is not None:
                    found = {lyr.name for lyr in all_layers}
                    missing = [nm for nm in names if nm not in found]
                    if missing:
                        raise ValueError(f"Requested image layer(s) not found: {', '.join(missing)}")

                layers = list(all_layers)

                if visible_only:
                    layers = [lyr for lyr in layers if getattr(lyr, "visible", True)]

                if names is not None:
                    wanted = set(names)
                    layers = [lyr for lyr in layers if lyr.name in wanted]
                    layers.sort(key=lambda lyr: names.index(lyr.name))

                if not layers:
                    raise ValueError("No matching Image layers found.")

                return layers

            source_layers = _get_image_layers(image_layer_names, visible_only=visible_only)

            validated_layers = []
            rejected = []

            ref_yx = None
            ref_scale_yx = None
            ref_translate_yx = None

            for lyr in source_layers:
                try:
                    if getattr(lyr, "multiscale", False):
                        rejected.append(f"{lyr.name}: multiscale layers are not supported")
                        continue

                    if bool(getattr(lyr, "rgb", False)):
                        rejected.append(f"{lyr.name}: RGB image layers are not supported")
                        continue

                    arr = np.asarray(lyr.data)
                    if arr.ndim < 3:
                        rejected.append(f"{lyr.name}: requires 3D+ data, got shape {arr.shape}")
                        continue

                    unsupported, reason = _has_unsupported_transform(lyr)
                    if unsupported:
                        rejected.append(f"{lyr.name}: unsupported {reason}")
                        continue

                    yx = tuple(int(v) for v in arr.shape[-2:])
                    scale = _as_nd_tuple(getattr(lyr, "scale", None), arr.ndim, 1.0)
                    translate = _as_nd_tuple(getattr(lyr, "translate", None), arr.ndim, 0.0)

                    scale_yx = tuple(float(v) for v in scale[-2:])
                    translate_yx = tuple(float(v) for v in translate[-2:])

                    if ref_yx is None:
                        ref_yx = yx
                        ref_scale_yx = scale_yx
                        ref_translate_yx = translate_yx
                    else:
                        if yx != ref_yx:
                            rejected.append(f"{lyr.name}: YX shape {yx} does not match reference {ref_yx}")
                            continue
                        if not _allclose(scale_yx, ref_scale_yx):
                            rejected.append(f"{lyr.name}: YX scale {scale_yx} does not match reference {ref_scale_yx}")
                            continue
                        if not _allclose(translate_yx, ref_translate_yx):
                            rejected.append(
                                f"{lyr.name}: YX translate {translate_yx} does not match reference {ref_translate_yx}"
                            )
                            continue

                    validated_layers.append(lyr)

                except Exception as e:
                    rejected.append(f"{getattr(lyr, 'name', '<unnamed>')}: {e}")

            if not validated_layers:
                msg = "No compatible 3D+ image layers with shared YX geometry were found."
                if rejected:
                    msg += "\n\nRejected layers:\n" + "\n".join(rejected)
                raise ValueError(msg)

            base_yx = ref_yx
            proj_scale_yx = tuple(ref_scale_yx)
            proj_translate_yx = tuple(ref_translate_yx)

            crop_viewer = napari.Viewer(ndisplay=2, title="Crop ROI - all channels")

            for lyr in validated_layers:
                arr = np.asarray(lyr.data)
                proj = np.max(arr, axis=tuple(range(arr.ndim - 2)))

                scale = _as_nd_tuple(getattr(lyr, "scale", None), arr.ndim, 1.0)
                translate = _as_nd_tuple(getattr(lyr, "translate", None), arr.ndim, 0.0)

                kwargs = dict(
                    name=f"{lyr.name}_maxproj",
                    scale=tuple(scale[-2:]),
                    translate=tuple(translate[-2:]),
                    opacity=float(getattr(lyr, "opacity", 1.0)),
                    blending=str(getattr(lyr, "blending", "additive")),
                    visible=True,
                    rgb=False,
                )

                try:
                    kwargs["colormap"] = lyr.colormap
                except Exception:
                    pass

                try:
                    kwargs["contrast_limits"] = tuple(lyr.contrast_limits)
                except Exception:
                    pass

                try:
                    kwargs["gamma"] = float(lyr.gamma)
                except Exception:
                    pass

                crop_viewer.add_image(proj, **kwargs)

            shapes_layer = crop_viewer.add_shapes(
                name="Crop ROI",
                shape_type="polygon",
                edge_width=2,
                edge_color="yellow",
                face_color=[1.0, 1.0, 0.0, 0.15],
                opacity=1.0,
                scale=proj_scale_yx,
                translate=proj_translate_yx,
            )

            try:
                shapes_layer.mode = "add_polygon"
            except Exception:
                pass

            class CropWidget(QWidget):
                def __init__(self):
                    super().__init__()
                    layout = QVBoxLayout()
                    layout.addWidget(QLabel("Draw one or more polygons on the projection, then click crop."))
                    crop_btn = QPushButton("Crop all channels in place")
                    close_btn = QPushButton("Close")
                    layout.addWidget(crop_btn)
                    layout.addWidget(close_btn)
                    self.setLayout(layout)
                    crop_btn.clicked.connect(self._apply_crop)
                    close_btn.clicked.connect(self._close_viewer)

                def _close_viewer(self):
                    try:
                        crop_viewer.close()
                    except Exception:
                        pass

                def _shape_vertices_to_ref_pixel_yx(self, verts):
                    verts = np.asarray(verts, dtype=float)
                    if verts.ndim != 2 or verts.shape[1] < 2:
                        raise ValueError("ROI vertices are invalid.")

                    yx_world = np.asarray(verts[:, -2:], dtype=float)
                    yx = np.empty_like(yx_world, dtype=float)
                    yx[:, 0] = (yx_world[:, 0] - proj_translate_yx[0]) / proj_scale_yx[0]
                    yx[:, 1] = (yx_world[:, 1] - proj_translate_yx[1]) / proj_scale_yx[1]

                    if not np.all(np.isfinite(yx)):
                        raise ValueError("ROI contains non-finite coordinates.")

                    return yx

                def _build_mask_and_bbox(self):
                    if len(shapes_layer.data) == 0:
                        raise ValueError("Draw at least one polygon first.")

                    try:
                        selected = sorted(int(i) for i in shapes_layer.selected_data)
                    except Exception:
                        selected = []

                    if not selected:
                        selected = list(range(len(shapes_layer.data)))

                    mask2d = np.zeros(base_yx, dtype=bool)

                    for idx in selected:
                        verts = np.asarray(shapes_layer.data[idx], dtype=float)
                        if verts.ndim != 2 or verts.shape[0] < 3 or verts.shape[1] < 2:
                            continue

                        yx = self._shape_vertices_to_ref_pixel_yx(verts)

                        yx[:, 0] = np.clip(yx[:, 0], -0.5, base_yx[0] - 0.5)
                        yx[:, 1] = np.clip(yx[:, 1], -0.5, base_yx[1] - 0.5)

                        poly_mask = polygon2mask(base_yx, yx)
                        if np.any(poly_mask):
                            mask2d |= poly_mask

                    if not np.any(mask2d):
                        raise ValueError("Selected ROI polygon(s) did not produce a valid mask.")

                    if tight_bbox:
                        yy, xx = np.where(mask2d)
                        y0, y1 = int(yy.min()), int(yy.max()) + 1
                        x0, x1 = int(xx.min()), int(xx.max()) + 1
                    else:
                        y0, y1 = 0, base_yx[0]
                        x0, x1 = 0, base_yx[1]

                    if y1 <= y0 or x1 <= x0:
                        raise ValueError("Computed crop bounding box is empty.")

                    return mask2d, (y0, y1, x0, x1)

                def _crop_single_layer_in_place(self, image_layer, mask2d, bbox):
                    y0, y1, x0, x1 = bbox

                    old_data = np.asarray(image_layer.data)
                    ndim = old_data.ndim
                    if ndim < 3:
                        raise ValueError("Layer is no longer 3D+.")

                    slicer = (slice(None),) * (ndim - 2) + (slice(y0, y1), slice(x0, x1))
                    cropped = np.array(old_data[slicer], copy=True, order="C")

                    if cropped.size == 0:
                        raise ValueError("Crop produced an empty array.")

                    mask_crop = np.asarray(mask2d[y0:y1, x0:x1], dtype=bool)
                    mask_view = mask_crop.reshape((1,) * (cropped.ndim - 2) + mask_crop.shape)

                    try:
                        fill_cast = np.asarray(fill_value, dtype=cropped.dtype)
                    except Exception:
                        raise ValueError(f"fill_value={fill_value!r} cannot be cast to dtype {cropped.dtype}")

                    if not np.all(mask_crop):
                        cropped = np.where(mask_view, cropped, fill_cast)

                    # ---- FIX: reset translate to 0 for all axes; the XY offset is
                    # absorbed into channel_list below so world-space stays consistent. ----
                    old_scale = _as_nd_tuple(getattr(image_layer, "scale", None), ndim, 1.0)
                    old_translate = list(_as_nd_tuple(getattr(image_layer, "translate", None), ndim, 0.0))

                    # Only propagate the translate offset if there is a non-trivial
                    # pre-existing world translate on this layer (e.g. from a previous
                    # XY crop).  Otherwise reset to zero so channel_list and the napari
                    # layer stay aligned.
                    old_translate[-2] = float(old_translate[-2]) + float(y0) * float(old_scale[-2])
                    old_translate[-1] = float(old_translate[-1]) + float(x0) * float(old_scale[-1])

                    image_layer.data = cropped

                    try:
                        image_layer.translate = tuple(old_translate)
                    except Exception:
                        pass

                    # ---- FIX: sync channel_list so downstream analysis uses cropped data.
                    # `self` here is CropWidget; use outer `self` (UIWidget) from closure. ----
                    try:
                        ui = self  # CropWidget; access UIWidget via closure var
                        # The outer function parameter named `self` is the UIWidget
                        ch_layers = getattr(self_parent, "channel_layers", None)
                        ch_list   = getattr(self_parent, "channel_list",  None)
                        if ch_layers is not None and ch_list is not None:
                            if image_layer in ch_layers:
                                idx_ch = ch_layers.index(image_layer)
                                if idx_ch < len(ch_list):
                                    ch_list[idx_ch] = cropped
                    except Exception:
                        pass

                    try:
                        image_layer.refresh()
                    except Exception:
                        pass

                    return f"Updated: {image_layer.name} -> {cropped.shape}"

                def _apply_crop(self):
                    try:
                        mask2d, bbox = self._build_mask_and_bbox()
                    except Exception as e:
                        QMessageBox.warning(self, "Crop", str(e))
                        return

                    try:
                        crop_viewer.close()
                    except Exception:
                        pass

                    def _do_crop():
                        summary = []
                        failed = []

                        for image_layer in validated_layers:
                            try:
                                summary.append(self._crop_single_layer_in_place(image_layer, mask2d, bbox))
                            except Exception as e:
                                failed.append(f"{image_layer.name}: {e}")

                        # ---- FIX: after XY crop, invalidate stale derived / mask stacks
                        # whose YX shape no longer matches the newly cropped channel_list. ----
                        if summary:
                            try:
                                new_yx = None
                                for ch in getattr(self_parent, "channel_list", []):
                                    ca = np.asarray(ch)
                                    if ca.ndim == 3:
                                        new_yx = ca.shape[1:]
                                        break

                                if new_yx is not None:
                                    # Invalidate masks with wrong YX
                                    if hasattr(self_parent, "masks_by_channel"):
                                        for k, m in list(self_parent.masks_by_channel.items()):
                                            if m is not None:
                                                ma = np.asarray(m)
                                                if ma.ndim == 3 and ma.shape[1:] != new_yx:
                                                    self_parent.masks_by_channel[k] = None

                                    # Invalidate derived stacks with wrong YX
                                    if hasattr(self_parent, "derived_stacks"):
                                        stale = []
                                        for k, ds in list(self_parent.derived_stacks.items()):
                                            if ds is not None:
                                                dsa = np.asarray(ds)
                                                if dsa.ndim == 3 and dsa.shape[1:] != new_yx:
                                                    stale.append(k)
                                        for k in stale:
                                            del self_parent.derived_stacks[k]
                                            if hasattr(self_parent, "derived_layers") and k in self_parent.derived_layers:
                                                try:
                                                    old_lay = self_parent.derived_layers.pop(k)
                                                    self_parent.viewer.layers.remove(old_lay)
                                                except Exception:
                                                    pass
                            except Exception:
                                pass

                        try:
                            main_viewer.reset_view()
                        except Exception:
                            pass

                        parts = []

                        if summary:
                            parts.append("Crop applied in place.")
                            parts.append("")
                            parts.extend(summary)

                        if failed:
                            if parts:
                                parts.append("")
                            parts.append("Failed:")
                            parts.extend(failed)

                        if rejected:
                            if parts:
                                parts.append("")
                            parts.append("Skipped before cropping:")
                            parts.extend(rejected)

                        text = "\n".join(parts).strip() or "No layers were cropped."

                        if summary:
                            QMessageBox.information(_qt_parent(), "Crop", text)
                        else:
                            QMessageBox.warning(_qt_parent(), "Crop", text)

                    QTimer.singleShot(0, _do_crop)

            crop_widget = CropWidget()
            crop_viewer.window.add_dock_widget(crop_widget, area="right")

            if rejected:
                QMessageBox.information(
                    _qt_parent(),
                    "Crop ROI",
                    "Some layers were skipped because they do not share the same crop geometry.\n\n"
                    + "\n".join(rejected),
                )

            return crop_viewer, shapes_layer

        def _refresh_z_controls():
            """Update n_z-dependent UI + refresh viewer to current slice."""
            end = max(self_parent.n_z - 1, 0)

            self_parent.z_slider.blockSignals(True)
            self_parent.z_slider.setRange(0, end)
            self_parent.current_z = max(0, min(self_parent.current_z, end))
            self_parent.z_slider.setValue(self_parent.current_z)
            self_parent.z_slider.blockSignals(False)

            _update_z_label()
            self_parent.z_changed(self_parent.current_z)

        delete_z = QPushButton("Delete z slices")

        def do_delete():
            z0, ok1 = QInputDialog.getInt(
                self_parent, "Delete Start", "Start Z slice to delete:",
                0, 0, max(0, self_parent.n_z - 1)
            )
            if not ok1:
                return

            z1, ok2 = QInputDialog.getInt(
                self_parent, "Delete End", "End Z slice to delete:",
                z0, 0, max(0, self_parent.n_z - 1)
            )
            if not ok2:
                return

            if z0 > z1:
                QMessageBox.warning(self_parent, "Invalid Range", "Start Z must be <= End Z")
                return

            try:
                for i, ch in enumerate(self_parent.channel_list):
                    arr = np.asarray(ch)
                    if arr.ndim == 3:
                        new = np.concatenate([arr[:z0], arr[z1+1:]], axis=0) if (z1 + 1) < arr.shape[0] else arr[:z0]
                        self_parent.channel_list[i] = new

                first = np.asarray(self_parent.channel_list[0])
                self_parent.n_z = first.shape[0] if first.ndim == 3 else 1

                # After deleting, crop provenance is no longer meaningful
                self_parent.last_crop_range = None

                _refresh_z_controls()
                QMessageBox.information(self_parent, "Deleted", f"Deleted Z {z0}-{z1}")
            except Exception as e:
                QMessageBox.critical(self_parent, "Error", f"Failed to delete slices: {e}")

        delete_z.clicked.connect(do_delete)
        v.addWidget(delete_z)

        crop_z = QPushButton("Crop z slices")

        def do_crop():
            if self_parent.n_z <= 0:
                QMessageBox.warning(self_parent, "No Data", "No Z slices available")
                return

            z0, ok1 = QInputDialog.getInt(
                self_parent, "Crop Start", "Start Z slice to keep:",
                0, 0, max(0, self_parent.n_z - 1)
            )
            if not ok1:
                return

            z1, ok2 = QInputDialog.getInt(
                self_parent, "Crop End", "End Z slice to keep:",
                z0, 0, max(0, self_parent.n_z - 1)
            )
            if not ok2:
                return

            if z0 > z1:
                QMessageBox.warning(self_parent, "Invalid Range", "Start Z must be <= End Z")
                return

            try:
                for i, ch in enumerate(self_parent.channel_list):
                    arr = np.asarray(ch)
                    if arr.ndim == 3:
                        cropped_arr = arr[z0:z1+1].copy()
                        self_parent.channel_list[i] = cropped_arr
                        # ---- FIX: push cropped 3D array into the napari layer ----
                        if hasattr(self_parent, "channel_layers") and i < len(self_parent.channel_layers):
                            try:
                                lay = self_parent.channel_layers[i]
                                lay.data = cropped_arr
                                # Reset translate Z so the layer's world origin stays at 0
                                try:
                                    tr = list(getattr(lay, "translate", None) or ([0.0] * cropped_arr.ndim))
                                    while len(tr) < cropped_arr.ndim:
                                        tr.insert(0, 0.0)
                                    tr[0] = 0.0  # Z translate reset after crop
                                    lay.translate = tuple(tr)
                                except Exception:
                                    pass
                            except Exception:
                                pass

                first = np.asarray(self_parent.channel_list[0])
                self_parent.n_z = first.shape[0] if first.ndim == 3 else 1

                # Store the original crop selection for display only
                self_parent.last_crop_range = (z0, z1)

                # ---- FIX: invalidate derived stacks and masks that have old Z shape ----
                new_nz = self_parent.n_z
                if hasattr(self_parent, "masks_by_channel"):
                    for k, m in list(self_parent.masks_by_channel.items()):
                        if m is not None:
                            ma = np.asarray(m)
                            if ma.ndim == 3 and ma.shape[0] != new_nz:
                                self_parent.masks_by_channel[k] = None
                if hasattr(self_parent, "derived_stacks"):
                    stale_keys = []
                    for k, ds in list(self_parent.derived_stacks.items()):
                        if ds is not None:
                            dsa = np.asarray(ds)
                            if dsa.ndim == 3 and dsa.shape[0] != new_nz:
                                stale_keys.append(k)
                    for k in stale_keys:
                        del self_parent.derived_stacks[k]
                        if hasattr(self_parent, "derived_layers") and k in self_parent.derived_layers:
                            try:
                                old_lay = self_parent.derived_layers.pop(k)
                                self_parent.viewer.layers.remove(old_lay)
                            except Exception:
                                pass

                _refresh_z_controls()
                QMessageBox.information(self_parent, "Cropped", f"Kept Z {z0}-{z1} (now {self_parent.n_z} slices)")
            except Exception as e:
                QMessageBox.critical(self_parent, "Error", f"Failed to crop slices: {e}")

        crop_z.clicked.connect(do_crop)
        v.addWidget(crop_z)

        dlg.setLayout(v)
        dlg.show()

        rot90 = QPushButton('Rotate 90')
        rot90.clicked.connect(lambda: self.rotate_stack(90))
        v.addWidget(rot90)

        rot180 = QPushButton('Rotate 180')
        rot180.clicked.connect(lambda: self.rotate_stack(180))
        v.addWidget(rot180)

        crop_btn = QPushButton("Crop to Selected Shape (All Z)")
        v.addWidget(crop_btn)
        crop_btn.clicked.connect(lambda: open_crop_projection_viewer(self))

        gen_mask_btn = QPushButton("Generate Mask Channel(s) (Otsu)")

        def do_generate_masks():
            txt, ok = QInputDialog.getText(
                self_parent,
                "Channels",
                f'Enter channel numbers (comma-separated) or "all":',
                text="1",
            )
            if not ok or not txt:
                return

            txt = txt.strip()
            n = len(self_parent.channel_list)

            if txt.lower() == "all":
                choices = list(range(1, n + 1))
            else:
                choices = []
                for p in [p.strip() for p in txt.split(",") if p.strip()]:
                    try:
                        v = int(p)
                        if 1 <= v <= n:
                            choices.append(v)
                    except Exception:
                        pass

            if not choices:
                QMessageBox.warning(self_parent, "No Channels", "No valid channels selected")
                return

            # Keep references so dialogs don't get GC'd
            self_parent._mask_tuners = getattr(self_parent, "_mask_tuners", [])

            for chnum in choices:
                idx = chnum - 1
                arr = self_parent.channel_list[idx]

                dlg = MaskTunerDialog(self_parent, idx, arr, file_path=self_parent.file_path)
                dlg.setModal(False)

                self_parent._mask_tuners.append(dlg)

                # remove reference when actually destroyed
                dlg.destroyed.connect(
                    lambda _=None, d=dlg: self_parent._mask_tuners.remove(d)
                    if d in self_parent._mask_tuners else None
                )

                dlg.show()

        gen_mask_btn.clicked.connect(do_generate_masks)
        v.addWidget(gen_mask_btn)

    def threshold_channel_dialog(self, ch_idx):
        """Simple threshold dialog with preview and confirm flow.
        Uses the currently displayed slice for defaults and computes masks
        across the full Z-stack when applying.
        """
        arr = np.asarray(self.channel_list[ch_idx])
        if arr.ndim != 3:
            QMessageBox.information(self, 'Single-slice', 'Channel has no Z dimension; thresholding will operate on full image')

        dialog = QDialog(self)
        dialog.setWindowTitle(f'Threshold Channel {ch_idx+1}')
        layout = QVBoxLayout()

        # Slider for threshold
        th_slider = QSlider(Qt.Horizontal)
        th_slider.setRange(0, 65535)
        # Default based on current Z slice (if available)
        cur_z = int(getattr(self, 'current_z', 0))
        if arr.ndim == 3:
            cur_z = max(0, min(cur_z, arr.shape[0]-1))
        default = int(np.percentile(arr[cur_z], 50)) if arr.size > 0 else 100
        th_slider.setValue(default)
        th_label = QLabel(str(default))
        th_slider.valueChanged.connect(lambda v: th_label.setText(str(v)))

        layout.addWidget(QLabel('Manual threshold'))
        layout.addWidget(th_slider)
        layout.addWidget(th_label)

        # Preview button: shows mask preview across Z-stack
        preview_btn = QPushButton('Preview Mask (full stack)')
        def do_preview():
            th = th_slider.value()
            if arr.ndim == 3:
                masks = np.zeros_like(arr, dtype=np.uint8)
                for z in range(arr.shape[0]):
                    masks[z] = compute_mask(arr[z], threshold=th, blur_enabled=False)
                v = napari.Viewer()
                v.add_image(arr, name=f'Ch{ch_idx+1} stack')
                v.add_image(masks, name='Mask Preview', colormap='gray', opacity=0.6)
            else:
                mask = compute_mask(arr, threshold=th, blur_enabled=False)
                v = napari.Viewer()
                v.add_image(arr, name=f'Ch{ch_idx+1} image')
                v.add_image(mask, name='Mask Preview', colormap='gray', opacity=0.6)
        preview_btn.clicked.connect(do_preview)
        layout.addWidget(preview_btn)

        # Otsu auto-threshold button (uses current slice)
        otsu_btn = QPushButton('Use Otsu (current slice)')
        def do_otsu():
            try:
                if arr.ndim == 3:
                    zidx = int(getattr(self, 'current_z', 0))
                    zidx = max(0, min(zidx, arr.shape[0]-1))
                    t = int(threshold_otsu(arr[zidx]))
                else:
                    t = int(threshold_otsu(arr))
                th_slider.setValue(t)
                th_label.setText(str(t))
            except Exception as e:
                QMessageBox.critical(self, 'Otsu Error', f'Otsu failed: {e}')
        otsu_btn.clicked.connect(do_otsu)
        layout.addWidget(otsu_btn)

        # Confirm/cancel buttons
        btns = QHBoxLayout()
        ok_btn = QPushButton('Apply and Close')
        cancel_btn = QPushButton('Cancel')
        btns.addWidget(ok_btn)
        btns.addWidget(cancel_btn)
        layout.addLayout(btns)

        def do_apply():
            th = th_slider.value()
            # compute full 3D mask or single mask
            if arr.ndim == 3:
                masks = np.zeros_like(arr, dtype=np.uint8)
                for z in range(arr.shape[0]):
                    masks[z] = compute_mask(arr[z], threshold=th)
            else:
                masks = compute_mask(arr, threshold=th)
            # create mask layer and add controls in main UI
            self.create_mask_layer(ch_idx, masks)
            dialog.accept()

        def do_cancel():
            reply = QMessageBox.question(self, 'Still Thresholding?', 'Are you done thresholding this channel?', QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                dialog.reject()

        ok_btn.clicked.connect(do_apply)
        cancel_btn.clicked.connect(do_cancel)

        dialog.setLayout(layout)
        dialog.exec_()

    def create_mask_layer(self, ch_idx, mask_array):
        name = f"Channel {ch_idx+1} mask"

        mask3 = np.asarray(mask_array)
        if mask3.ndim == 2:
            mask3 = mask3[np.newaxis, ...]
        mask3 = (mask3 > 0).astype(np.uint8)

        self.masks_by_channel = getattr(self, "masks_by_channel", {})
        self.masks_by_channel[ch_idx] = mask3

        z = int(getattr(self, "current_z", 0))
        z = max(0, min(z, mask3.shape[0] - 1))
        mask2d = mask3[z]

        # ---- FIX: inherit scale and translate from the matching channel layer
        # so the 2D mask slice renders at exactly the same pixel positions as
        # the channel image in the napari canvas. ----
        ref_scale_yx = None
        ref_translate_yx = None
        try:
            ch_layers = getattr(self, "channel_layers", [])
            if ch_idx < len(ch_layers):
                ref_lay = ch_layers[ch_idx]
                ref_arr = np.asarray(ref_lay.data)
                ndim_ref = ref_arr.ndim

                raw_scale = getattr(ref_lay, "scale", None)
                raw_translate = getattr(ref_lay, "translate", None)

                if raw_scale is not None:
                    s = tuple(float(v) for v in raw_scale)
                    # Take last 2 elements (YX) regardless of whether
                    # the channel layer is 2D or 3D
                    ref_scale_yx = s[-2:] if len(s) >= 2 else (1.0, 1.0)

                if raw_translate is not None:
                    t = tuple(float(v) for v in raw_translate)
                    ref_translate_yx = t[-2:] if len(t) >= 2 else (0.0, 0.0)
        except Exception:
            pass

        add_kwargs = dict(
            name=name,
            colormap="yellow",
            opacity=0.6,
            blending="additive",
            contrast_limits=(0, 1),
        )
        if ref_scale_yx is not None:
            add_kwargs["scale"] = ref_scale_yx
        if ref_translate_yx is not None:
            add_kwargs["translate"] = ref_translate_yx

        try:
            lay = self.viewer.layers.get(name)
        except Exception:
            lay = None

        if lay is None:
            lay = self.viewer.add_image(mask2d, **add_kwargs)
        else:
            lay.data = mask2d
            # Re-apply scale/translate in case the layer existed from before crop
            if ref_scale_yx is not None:
                try:
                    lay.scale = ref_scale_yx
                except Exception:
                    pass
            if ref_translate_yx is not None:
                try:
                    lay.translate = ref_translate_yx
                except Exception:
                    pass
            try:
                lay.contrast_limits = (0, 1)
            except Exception:
                pass

        self.mask_layers = getattr(self, "mask_layers", {})
        self.mask_layers[ch_idx] = lay

    def rotate_stack(self, deg):
        for i, ch in enumerate(self.channel_list):
            arr = np.asarray(ch)
            if arr.ndim == 3:
                for z in range(arr.shape[0]):
                    arr[z] = rotate(arr[z], deg, preserve_range=True)
            else:
                arr = rotate(arr, deg, preserve_range=True)
            self.channel_list[i] = arr
        # refresh current slice
        self.z_changed(self.current_z)

    def load_channels(self):
        """Load more channels and append them to the existing viewer/UI (no overwrite)."""
        try:
            path, _ = QFileDialog.getOpenFileName(
                self,
                "Load LIF/TIFF",
                "",
                "LIF (*.lif);;TIFF (*.tif *.tiff);;All files (*)",
            )
            if not path:
                return

            p = Path(path)
            self.current_file_path = str(p)
            print(f"\n[load_channels] Loading: {p}")

            if p.suffix.lower() == ".lif":
                mmaps, clims, vox, temp = load_lif_memmap(p, max_ch=4)
            else:
                mmaps, clims, vox, temp = load_tiff_memmap(p, max_ch=4)

            # Debug: what did the loader return?
            try:
                print(f"[load_channels] mmaps type={type(mmaps)}, len={len(mmaps) if mmaps is not None else 'None'}")
            except Exception as _:
                print("[load_channels] mmaps: (could not get len)")

            if mmaps is None:
                print("[load_channels] ERROR: mmaps is None")
                QMessageBox.warning(self, "Load", "No channels returned by loader.")
                return

            # Ensure lists exist
            self.channel_list = getattr(self, "channel_list", []) or []
            self.channel_layers = getattr(self, "channel_layers", []) or []
            self.channel_controls = getattr(self, "channel_controls", []) or []
            self.channel_full_stacks = getattr(self, "channel_full_stacks", []) or []

            if not hasattr(self, "contrast_limits") or self.contrast_limits is None:
                self.contrast_limits = {}

            # Keep current z/nz if present
            self.z = int(getattr(self, "z", 0))
            prev_nz = int(getattr(self, "nz", 1) or 1)

            # Debug: current app stacks
            try:
                a0 = np.asarray(getattr(self, "arr3", None))
                print(f"[load_channels] existing arr3: {'None' if a0 is None else (a0.shape, a0.dtype, a0.ndim, a0.size)}")
            except Exception as e:
                print(f"[load_channels] existing arr3: error {e}")

            try:
                pm = np.asarray(getattr(self, "preview_masks", None))
                print(f"[load_channels] existing preview_masks: {'None' if pm is None else (pm.shape, pm.dtype, pm.ndim, pm.size)}")
            except Exception as e:
                print(f"[load_channels] existing preview_masks: error {e}")

            start_idx = len(self.channel_list)
            print(f"[load_channels] start_idx={start_idx}, prev_nz={prev_nz}, current z={self.z}")

            # Add new channels
            for j, ch in enumerate(mmaps):
                new_idx = start_idx + j + 1

                stack = np.asarray(ch)

                # Debug each returned channel
                try:
                    print(
                        f"[load_channels] incoming ch{j+1}: shape={stack.shape}, ndim={stack.ndim}, "
                        f"dtype={stack.dtype}, size={stack.size}, nbytes={stack.nbytes}"
                    )
                except Exception as e:
                    print(f"[load_channels] incoming ch{j+1}: could not print details ({e})")

                self.channel_full_stacks.append(stack)
                self.channel_list.append(ch)

                nz_this = int(stack.shape[0]) if stack.ndim == 3 else 1
                z0 = int(max(0, min(self.z, nz_this - 1)))

                view2d = stack[z0] if stack.ndim == 3 else stack

                # Debug view2d
                try:
                    v2 = np.asarray(view2d)
                    print(
                        f"[load_channels] display view for Channel {new_idx}: shape={v2.shape}, ndim={v2.ndim}, "
                        f"dtype={v2.dtype}, size={v2.size}"
                    )
                except Exception:
                    pass

                channel_name = f"Channel {new_idx}"

                lay = self.viewer.add_image(
                    view2d,
                    name=channel_name,
                    opacity=1.0,
                    metadata={
                        "path": str(p),
                        "file_path": str(p),
                        "source_file": str(p),
                        "filename": str(p.name),
                        "channel_number": int(new_idx),
                        "channel_name": channel_name,
                    },
                )
                self.channel_layers.append(lay)

                # Contrast limits
                guess = clims.get(f"ch{j+1}", None) if isinstance(clims, dict) else None
                lo, hi, slider_max = self._compute_channel_display_range(stack)

                if guess and isinstance(guess, (tuple, list)) and len(guess) == 2:
                    self.contrast_limits[f"ch{new_idx}"] = (float(guess[0]), float(guess[1]))
                else:
                    self.contrast_limits[f"ch{new_idx}"] = (float(lo), float(hi))

                try:
                    lay.contrast_limits = tuple(self.contrast_limits[f"ch{new_idx}"])
                except Exception:
                    pass

                gb, ctrl = self._add_channel_controls(lay, new_idx, max_range=int(slider_max))
                self.channel_controls.append(ctrl)
                self.layout().insertWidget(self.layout().count() - 1, gb)

            # --------------------------
            # Compute nz WITHOUT collapsing it due to 2D loads
            # --------------------------
            nz_target = None

            # 1) main image stack wins
            try:
                a0 = np.asarray(getattr(self, "arr3", None))
                if a0 is not None and a0.ndim == 3:
                    nz_target = int(a0.shape[0])
            except Exception:
                pass

            # 2) else main mask stack
            if nz_target is None:
                try:
                    pm = np.asarray(getattr(self, "preview_masks", None))
                    if pm is not None and pm.ndim == 3:
                        nz_target = int(pm.shape[0])
                except Exception:
                    pass

            # 3) else loaded channel stacks (max Z)
            if nz_target is None:
                z_candidates = []
                for st in self.channel_full_stacks:
                    st = np.asarray(st)
                    if st.ndim == 3:
                        z_candidates.append(int(st.shape[0]))
                if z_candidates:
                    nz_target = int(max(z_candidates))

            # 4) else keep previous nz (do not reduce)
            if nz_target is None:
                nz_target = prev_nz

            self.nz = max(1, int(nz_target))

            print(f"[load_channels] nz_target decided={self.nz}")

            # Update slider range/value
            self.z_slider.setRange(0, max(self.nz - 1, 0))
            self.z = int(max(0, min(self.z, self.nz - 1)))
            self.z_slider.blockSignals(True)
            self.z_slider.setValue(self.z)
            self.z_slider.blockSignals(False)

            print(f"[load_channels] z_slider range set to (0, {max(self.nz - 1, 0)}), value={self.z}")

            # Track temp paths
            self._temp_paths = getattr(self, "_temp_paths", []) + (temp or [])

            QMessageBox.information(
                self,
                "Loaded",
                f"Added {len(mmaps)} channels (now {len(self.channel_list)} total).",
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load channels: {e}")

    def export_selected_layers(self):
        try:
            img_layers = [lyr for lyr in self.viewer.layers if hasattr(lyr, "data")]
            if not img_layers:
                QMessageBox.warning(self, "No layers", "No exportable layers found.")
                return

            names = [lyr.name for lyr in img_layers]
            txt, ok = QInputDialog.getText(
                self,
                "Export",
                f"Enter comma-separated layer numbers to export (1-{len(names)}):\n"
                + "\n".join([f"{i+1}: {n}" for i, n in enumerate(names)]),
            )
            if not ok or not txt:
                return

            nums = [int(x.strip()) for x in txt.split(",") if x.strip().isdigit()]
            out_dir = QFileDialog.getExistingDirectory(self, "Select output folder")
            if not out_dir:
                return

            # Helpers
            chan_pat = re.compile(r"^\s*Channel\s+(\d+)\s*$", re.IGNORECASE)
            mask_pat = re.compile(r"^\s*Channel\s+(\d+)\s+mask\s*$", re.IGNORECASE)

            # Ensure dicts exist
            self.masks_by_channel = getattr(self, "masks_by_channel", {}) or {}
            self.channel_stacks_by_index = getattr(self, "channel_stacks_by_index", {}) or {}

            for n in nums:
                i = n - 1
                if i < 0 or i >= len(img_layers):
                    continue

                lyr = img_layers[i]
                name = str(getattr(lyr, "name", "") or "")
                arr = np.asarray(lyr.data)
                src = "layer.data"

                # 1) Export masks from masks_by_channel using name "Channel N mask"
                m = mask_pat.match(name)
                if m:
                    ch_num = int(m.group(1))          # 1-based
                    ch_idx = ch_num - 1               # 0-based key used by your code
                    backing = self.masks_by_channel.get(ch_idx, None)
                    if backing is not None and np.asarray(backing).ndim == 3:
                        arr = np.asarray(backing)
                        src = f"masks_by_channel[{ch_idx}]"
                    print(f"[export] MASK '{name}': src={src}, shape={arr.shape}, dtype={arr.dtype}")

                # 2) Export channels from stored stacks using name "Channel N"
                else:
                    m2 = chan_pat.match(name)
                    if m2:
                        ch_num = int(m2.group(1))  # 1-based
                        backing = self.channel_stacks_by_index.get(ch_num, None)
                        if backing is not None and np.asarray(backing).ndim == 3:
                            arr = np.asarray(backing)
                            src = f"channel_stacks_by_index[{ch_num}]"

                    # 3) Fallback: if this layer is in channel_layers, export channel_full_stacks[k]
                    if src == "layer.data" and hasattr(self, "channel_layers") and hasattr(self, "channel_full_stacks"):
                        try:
                            k = self.channel_layers.index(lyr)
                        except Exception:
                            k = None
                        if k is not None and 0 <= k < len(self.channel_full_stacks):
                            full = np.asarray(self.channel_full_stacks[k])
                            if full.ndim == 3:
                                arr = full
                                src = f"channel_full_stacks[{k}]"

                    # 4) Main image display layer backing
                    if src == "layer.data" and lyr is getattr(self, "img_layer", None):
                        full = getattr(self, "arr3", None)
                        if full is not None and np.asarray(full).ndim == 3:
                            arr = np.asarray(full)
                            src = "arr3"

                    print(f"[export] '{name}': src={src}, shape={arr.shape}, dtype={arr.dtype}")

                out_arr, out_dtype, _clims = self._suggest_export_dtype_and_scale(arr)

                out_path = Path(out_dir) / f"{Path(self.file_path).stem}_{name}.tif"
                meta = {"axes": "ZYX"} if np.asarray(out_arr).ndim == 3 else None

                print(f"[export] writing '{out_path.name}': out_shape={np.asarray(out_arr).shape}, out_dtype={np.asarray(out_arr).dtype}")

                tifffile.imwrite(
                    str(out_path),
                    out_arr,
                    photometric="minisblack",
                    metadata=meta,
                )

            QMessageBox.information(self, "Exported", "Selected layers exported.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {e}")


    def _compute_channel_display_range(self, arr):
        """Return (lo, hi, slider_max) for UI/contrast based on dtype and data."""
        a = np.asarray(arr)
        if np.issubdtype(a.dtype, np.integer):
            hi = int(np.iinfo(a.dtype).max)
            return 0, hi, hi

        # float: if looks like mask, use 0..1 else use robust max
        finite = a[np.isfinite(a)]
        if finite.size == 0:
            return 0.0, 1.0, 1

        amin = float(finite.min())
        amax = float(finite.max())

        if amin >= -1e-6 and amax <= 1.0 + 1e-6:
            return 0.0, 1.0, 1

        # intensity float: choose slider max as ceil(max) but capped
        # (you can change the cap if you want bigger floats)
        slider_max = int(min(max(1000, np.ceil(amax)), 65535))
        return float(amin), float(amax), slider_max

    def _pick_layer_name(self, title, prompt, layer_names):
        name, ok = QInputDialog.getItem(self, title, prompt, layer_names, 0, False)
        if not ok or not name:
            return None
        return str(name)

    def _is_mask_like(self, name, arr):
        if "mask" in (name or "").lower():
            return True
        a = np.asarray(arr)
        try:
            u = np.unique(a)
            if u.size <= 3 and set(u.tolist()).issubset({0, 1, 255}):
                return True
        except Exception:
            pass
        return False
    
    def _resolve_stack3d(self, layer_name: str):
        """
        Return a full 3D stack (Z,Y,X) for any selectable layer name:
        - "Channel N"          -> self.channel_list[N-1]
        - "Channel N mask"     -> self.masks_by_channel[N-1]
        - derived layer name   -> self.derived_stacks[...] (supports name-keyed or id-keyed)
        - else fallback        -> viewer layer.data if it is already 3D (or 2D->Z=1)
        """
        # ---- Channels ----
        if layer_name.startswith("Channel ") and layer_name[8:].isdigit():
            idx = int(layer_name.split()[1]) - 1
            if 0 <= idx < len(getattr(self, "channel_list", [])):
                arr = np.asarray(self.channel_list[idx])
                return arr if arr.ndim == 3 else arr[np.newaxis, ...]

        # ---- Masks ----
        if layer_name.startswith("Channel ") and layer_name.endswith(" mask"):
            try:
                idx = int(layer_name.split()[1]) - 1
            except Exception:
                idx = None
            if idx is not None:
                m3 = getattr(self, "masks_by_channel", {}).get(idx, None)
                if m3 is not None:
                    m3 = np.asarray(m3)
                    return m3 if m3.ndim == 3 else m3[np.newaxis, ...]

        # ---- Derived stacks (coloc/multiply outputs) ----
        if hasattr(self, "derived_stacks"):
            ds = getattr(self, "derived_stacks", {})

            # name-keyed
            if layer_name in ds:
                arr = np.asarray(ds[layer_name])
                return arr if arr.ndim == 3 else arr[np.newaxis, ...]

            # id(lay)-keyed: find layer object by name
            dl = getattr(self, "derived_layers", {})
            for k, lay in dl.items():
                if getattr(lay, "name", None) == layer_name and k in ds:
                    arr = np.asarray(ds[k])
                    return arr if arr.ndim == 3 else arr[np.newaxis, ...]

        # ---- Fallback to napari layer.data ----
        lay = self._get_layer_by_name(layer_name)
        if lay is None:
            return None
        arr = np.asarray(lay.data)
        return arr if arr.ndim == 3 else arr[np.newaxis, ...]

    def _compute_reasonable_maxrange(self, arr3, fallback=65535):
        """Pick a slider max that makes sense for integer images; clamp to 65535 for UI."""
        try:
            a = np.asarray(arr3)
            if np.issubdtype(a.dtype, np.integer):
                mx = int(np.iinfo(a.dtype).max)
                return int(max(1000, min(mx, 65535)))
        except Exception:
            pass
        return int(fallback)

    def _add_coloc_controls(self, lay, out3, title: str, out_is_mask: bool):
        """
        Create and insert per-coloc UI controls (opacity + contrast min/max + brightness/contrast-mult).
        Stores bookkeeping so you don't create duplicates.
        """
        # init containers
        self.coloc_controls = getattr(self, "coloc_controls", [])
        self.coloc_control_boxes = getattr(self, "coloc_control_boxes", [])

        # don’t duplicate controls for the same layer object
        for c in self.coloc_controls:
            if c.get("layer", None) is lay:
                return

        gb = QGroupBox(title)
        v = QVBoxLayout()

        # ---- Opacity ----
        v.addWidget(QLabel("Opacity"))
        op_row = QHBoxLayout()
        op = QSlider(Qt.Horizontal)
        op.setRange(0, 100)
        op.setValue(int((lay.opacity if lay.opacity is not None else 0.7) * 100))
        op_val = QLineEdit(str(op.value()))
        op_val.setFixedWidth(60)
        op_val.setValidator(QIntValidator(0, 100))
        op_row.addWidget(op)
        op_row.addWidget(op_val)
        v.addLayout(op_row)

        def _apply_op(val=None):
            val = op.value() if val is None else int(val)
            lay.opacity = val / 100.0
            op_val.setText(str(val))

        op.valueChanged.connect(_apply_op)
        op_val.editingFinished.connect(lambda: self.setsliderfromlineedit(op_val, op, 0, 100))

        # ---- Contrast limits ----
        maxrange = 1 if out_is_mask else self._compute_reasonable_maxrange(out3, fallback=65535)

        v.addWidget(QLabel("Contrast"))
        min_row = QHBoxLayout()
        min_row.addWidget(QLabel("Min"))
        smin = QSlider(Qt.Horizontal)
        smin.setRange(0, maxrange)
        min_row.addWidget(smin)
        min_val = QLineEdit("0")
        min_val.setFixedWidth(80)
        min_val.setValidator(QIntValidator(0, maxrange))
        min_row.addWidget(min_val)
        v.addLayout(min_row)

        max_row = QHBoxLayout()
        max_row.addWidget(QLabel("Max"))
        smax = QSlider(Qt.Horizontal)
        smax.setRange(0, maxrange)
        max_row.addWidget(smax)
        max_val = QLineEdit(str(maxrange))
        max_val.setFixedWidth(80)
        max_val.setValidator(QIntValidator(0, maxrange))
        max_row.addWidget(max_val)
        v.addLayout(max_row)

        # initialize from layer if present; else default
        try:
            c0, c1 = lay.contrast_limits
            c0 = int(np.clip(int(c0), 0, maxrange))
            c1 = int(np.clip(int(c1), 0, maxrange))
        except Exception:
            c0, c1 = (0, 1) if out_is_mask else (0, maxrange)

        if c1 < c0:
            c1 = c0

        smin.setValue(c0); smax.setValue(c1)
        min_val.setText(str(c0)); max_val.setText(str(c1))

        def _apply_clims(_=None):
            lo = int(smin.value())
            hi = int(smax.value())
            if hi < lo:
                hi = lo
                smax.setValue(hi)
                max_val.setText(str(hi))
            min_val.setText(str(lo))
            max_val.setText(str(hi))
            try:
                lay.contrast_limits = (lo, hi)
            except Exception:
                pass

        smin.valueChanged.connect(_apply_clims)
        smax.valueChanged.connect(_apply_clims)
        min_val.editingFinished.connect(lambda: self.setsliderfromlineedit(min_val, smin, 0, maxrange))
        max_val.editingFinished.connect(lambda: self.setsliderfromlineedit(max_val, smax, 0, maxrange))

        # ---- Optional display tweaks (brightness + contrast multiplier) ----

        v.addWidget(QLabel("Display (optional)"))
        disp_row = QHBoxLayout()

        bright = QSlider(Qt.Horizontal)
        bright.setRange(-2000, 2000)
        bright.setValue(0)
        bright_val = QLineEdit("0")
        bright_val.setFixedWidth(60)
        bright_val.setValidator(QIntValidator(-2000, 2000))

        mult = QSlider(Qt.Horizontal)
        mult.setRange(10, 300)
        mult.setValue(100)
        mult_val = QLineEdit(str(mult.value()))
        mult_val.setFixedWidth(60)
        mult_val.setValidator(QIntValidator(10, 300))

        # layout
        disp_row.addWidget(QLabel("Brightness"))
        disp_row.addWidget(bright)
        disp_row.addWidget(bright_val)
        disp_row.addWidget(QLabel("Contrast"))
        disp_row.addWidget(mult)
        disp_row.addWidget(mult_val)
        v.addLayout(disp_row)

        def _apply_display(_=None):
            # derive from current min/max sliders 
            try:
                base_lo = float(smin.value())
                base_hi = float(smax.value())
                center = (base_lo + base_hi) / 2.0
                width = max(base_hi - base_lo, 1.0)
                b = float(bright.value())
                m = float(mult.value()) / 100.0
                new_lo = center - (width / 2.0) * m + b
                new_hi = center + (width / 2.0) * m + b
                lay.contrast_limits = (new_lo, new_hi)
            except Exception:
                pass

            bright_val.setText(str(int(bright.value())))
            mult_val.setText(str(int(mult.value())))

        bright.valueChanged.connect(_apply_display)
        mult.valueChanged.connect(_apply_display)
        bright_val.editingFinished.connect(lambda: self.setsliderfromlineedit(bright_val, bright, -2000, 2000))
        mult_val.editingFinished.connect(lambda: self.setsliderfromlineedit(mult_val, mult, 10, 300))

        gb.setLayout(v)

        # insert into UI (just before final stretch)
        self.layout().insertWidget(self.layout().count() - 1, gb)

        # bookkeeping
        self.coloc_controls.append(
            dict(layer=lay, opacity=op, cmin=smin, cmax=smax, bright=bright, mult=mult, groupbox=gb)
        )
        self.coloc_control_boxes.append(gb)

    def generate_coloc_entire_stack(self):
        """
        Multiply (mask x target) where target can be: mask, channel, or an existing derived layer.
        Always operates on full 3D stacks (Z,Y,X) resolved from canonical storage, not the 2D view.
        """
        try:
            layer_names = [lyr.name for lyr in self.viewer.layers]
            if not layer_names:
                QMessageBox.warning(self, "No layers", "No layers found in viewer.")
                return

            mask_name = self._pick_layer_name(
                "Select Mask", "Select MASK layer (binarized as >0):", layer_names
            )
            if mask_name is None:
                return

            target_name = self._pick_layer_name(
                "Select Target", "Select TARGET layer (mask/channel/derived):", layer_names
            )
            if target_name is None:
                return

            mask3 = self._resolve_stack3d(mask_name)
            target3 = self._resolve_stack3d(target_name)

            if mask3 is None or target3 is None:
                QMessageBox.warning(self, "Missing data", "Could not resolve full stack for one of the selections.")
                return

            mask3 = np.asarray(mask3)
            target3 = np.asarray(target3)

            if mask3.ndim != 3 or target3.ndim != 3:
                QMessageBox.warning(self, "Bad dims", f"Expected 3D stacks; got {mask3.ndim}D and {target3.ndim}D.")
                return

            if mask3.shape != target3.shape:
                QMessageBox.warning(
                    self,
                    "Shape mismatch",
                    f"Mask {mask3.shape} != Target {target3.shape}\n\n"
                    "If you recently cropped the stack, please re-generate any masks/coloc "
                    "layers using the cropped data before running colocalization."
                )
                return

            mask_bin = (mask3 > 0)

            target_is_mask = self._is_mask_like(target_name, target3)

            if target_is_mask:
                out3 = (mask_bin & (target3 > 0)).astype(np.uint8)
                out_is_mask = True
            else:
                out3 = target3 * mask_bin.astype(target3.dtype)
                out_is_mask = False

            requested_name = f"Mul({target_name})_xMask({mask_name})"

            z0 = int(getattr(self, "current_z", 0))
            z0 = max(0, min(z0, out3.shape[0] - 1))
            out2 = out3[z0]

            lay = self._get_layer_by_name(requested_name)

            if out_is_mask:
                cmap = "yellow"
            else:
                tlay = self._get_layer_by_name(target_name)
                cmap = getattr(tlay, "colormap", None) if tlay is not None else None
                cmap = cmap or "yellow"

            created_new = False
            if lay is None:
                lay = self.viewer.add_image(
                    out2,
                    name=requested_name,
                    colormap=cmap,
                    blending="additive",
                    opacity=0.6 if out_is_mask else 0.7,
                )
                created_new = True
            else:
                lay.data = out2
                try:
                    lay.colormap = cmap
                except Exception:
                    pass

            if out_is_mask:
                try:
                    lay.contrast_limits = (0, 1)
                except Exception:
                    pass

            # store derived using stable object key
            self.derived_stacks = getattr(self, "derived_stacks", {})
            self.derived_layers = getattr(self, "derived_layers", {})
            k = id(lay)
            self.derived_stacks[k] = out3
            self.derived_layers[k] = lay

            # NEW: create per-coloc sliders/controls (only once per layer)
            # (If you re-run and update same named layer, controls won't duplicate.)
            self._add_coloc_controls(
                lay,
                out3=out3,
                title=f"Coloc: {lay.name}",
                out_is_mask=bool(out_is_mask),
            )

            # force sync right now
            self.z_changed(int(getattr(self, "current_z", 0)))

            QMessageBox.information(self, "Done", f"Created/updated: {lay.name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Multiply failed: {e}")
            
    def generate_3d_surfaces(self):
        """
        3D surface generator (marching cubes) where surface color matches the
        *current* channel color in the main viewer UI, including user-picked hex colors.

        IMPORTANT: napari Surface layers are colored via vertex_values+colormap or vertex_colors,
        not face_color/edge_color. We set vertex_colors for a uniform solid surface color.
        """
        try:
            # ----------------------------
            # 0) Collect candidate layers
            # ----------------------------
            layer_names = [lyr.name for lyr in self.viewer.layers]
            if not layer_names:
                QMessageBox.warning(self, "No layers", "No layers found in viewer.")
                return None

            # ----------------------------
            # 1) Voxel size: always ask user
            #    (use metadata only as suggested defaults)
            # ----------------------------
            suggested = None

            # (A) If you already stored voxel_size somewhere, use it as suggestion
            try:
                vs = getattr(self, "voxel_size", None)
                if isinstance(vs, (list, tuple)) and len(vs) == 3:
                    suggested = (float(vs[0]), float(vs[1]), float(vs[2]))
            except Exception:
                pass

            # (B) Try TIFF metadata as suggestion (still only a suggestion)
            if suggested is None:
                path = getattr(self, "last_loaded_path", None)
                if path:
                    try:
                        from tifffile import TiffFile
                        with TiffFile(str(path)) as tif:
                            ij = getattr(tif, "imagej_metadata", None) or {}
                            z = ij.get("spacing", None)

                            tags = tif.pages[0].tags
                            xr = tags["XResolution"].value if "XResolution" in tags else None
                            yr = tags["YResolution"].value if "YResolution" in tags else None
                            ru = tags["ResolutionUnit"].value if "ResolutionUnit" in tags else None

                            def _res_to_size(res, unit_tag):
                                if res is None:
                                    return None
                                num, den = res
                                if num == 0:
                                    return None
                                pixels_per_unit = num / den
                                if unit_tag == 2:      # inch
                                    microns_per_unit = 25400.0
                                elif unit_tag == 3:    # cm
                                    microns_per_unit = 10000.0
                                else:
                                    microns_per_unit = 1.0
                                return microns_per_unit / pixels_per_unit

                            x_um = _res_to_size(xr, ru)
                            y_um = _res_to_size(yr, ru)

                            if z is not None and x_um is not None and y_um is not None:
                                suggested = (float(z), float(y_um), float(x_um))
                    except Exception:
                        pass

            if suggested is None:
                suggested = (1.0, 1.0, 1.0)

            # Always prompt user (defaulting to suggested values) 
            z_um, ok1 = QInputDialog.getDouble(
                self, "Voxel size", "Z voxel size (µm):", float(suggested[0]), 1e-6, 1e9, 6
            )
            if not ok1:
                return None
            y_um, ok2 = QInputDialog.getDouble(
                self, "Voxel size", "Y voxel size (µm):", float(suggested[1]), 1e-6, 1e9, 6
            )
            if not ok2:
                return None
            x_um, ok3 = QInputDialog.getDouble(
                self, "Voxel size", "X voxel size (µm):", float(suggested[2]), 1e-6, 1e9, 6
            )
            if not ok3:
                return None

            spacing_z, spacing_y, spacing_x = (float(z_um), float(y_um), float(x_um))

            # ----------------------------
            # 2) Checklist dialog
            # ----------------------------
            dlg = QDialog(self)
            dlg.setWindowTitle("Select layers for 3D surfaces")
            dlg.setMinimumWidth(420)
            v = QVBoxLayout(dlg)
            v.addWidget(QLabel("Check layers to convert to 3D surfaces (requires 3D stack):"))

            lst = QListWidget()
            lst.setSelectionMode(QAbstractItemView.NoSelection)
            for nm in layer_names:
                it = QListWidgetItem(nm)
                it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
                it.setCheckState(Qt.Unchecked)
                lst.addItem(it)
            v.addWidget(lst)

            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            v.addWidget(buttons)
            buttons.accepted.connect(dlg.accept)
            buttons.rejected.connect(dlg.reject)

            if dlg.exec() != QDialog.Accepted:
                return None

            selected = [lst.item(i).text() for i in range(lst.count())
                        if lst.item(i).checkState() == Qt.Checked]
            if not selected:
                QMessageBox.information(self, "3D Surfaces", "No layers selected.")
                return None

            # ----------------------------
            # 3) Create 3D viewer
            # ----------------------------
            v3 = napari.Viewer(ndisplay=3)
            v3.title = "3D Surfaces"
            try:
                v3.scale_bar.visible = True
                v3.scale_bar.unit = "um"
                v3.scale_bar.font_size = 12
                v3.scale_bar.color = "white"
                v3.scale_bar.box = True
                v3.scale_bar.position = "bottom_right"
            except Exception:
                pass
            try:
                v3.dims.axis_labels = ["Z", "Y", "X"]
            except Exception:
                pass

            dock = QWidget()
            dock_lay = QVBoxLayout(dock)
            ss_btn = QPushButton("Save Screenshot")
            ss_btn.clicked.connect(
                lambda _=False: self.save_screenshot(
                    viewer=v3, parent=self, default_name="3d_surfaces.png", canvas_only=True
                )
            )
            dock_lay.addWidget(ss_btn)
            dock_lay.addStretch(1)
            v3.window.add_dock_widget(dock, area="right")

            # ----------------------------
            # Helpers: color extraction
            # ----------------------------
            from skimage.measure import marching_cubes

            def _fallback_color(name):
                palette = [
                    (1.0, 0.2, 0.2, 0.85),
                    (0.2, 1.0, 0.2, 0.85),
                    (0.2, 0.5, 1.0, 0.85),
                    (1.0, 1.0, 0.2, 0.85),
                    (1.0, 0.2, 1.0, 0.85),
                    (0.2, 1.0, 1.0, 0.85),
                ]
                return palette[abs(hash(name)) % len(palette)]

            def _hex_to_rgba(hex_str, alpha=0.85):
                s = str(hex_str).strip()
                if not s.startswith("#"):
                    return None
                h = s[1:]
                if len(h) == 6:
                    r = int(h[0:2], 16) / 255.0
                    g = int(h[2:4], 16) / 255.0
                    b = int(h[4:6], 16) / 255.0
                    return (r, g, b, float(alpha))
                if len(h) == 8:
                    a = int(h[0:2], 16) / 255.0
                    r = int(h[2:4], 16) / 255.0
                    g = int(h[4:6], 16) / 255.0
                    b = int(h[6:8], 16) / 255.0
                    return (r, g, b, a)
                return None

            NAMED_RGBA = {
                "red":     (1.0, 0.0, 0.0, 0.85),
                "green":   (0.0, 1.0, 0.0, 0.85),
                "blue":    (0.0, 0.4, 1.0, 0.85),
                "magenta": (1.0, 0.0, 1.0, 0.85),
                "cyan":    (0.0, 1.0, 1.0, 0.85),
                "yellow":  (1.0, 1.0, 0.0, 0.85),
                "gray":    (0.85, 0.85, 0.85, 0.85),
                "grey":    (0.85, 0.85, 0.85, 0.85),
            }

            def _rgba_from_layer_current(layer_obj, fallback_name):
                if layer_obj is None:
                    return _fallback_color(fallback_name)

                cm = getattr(layer_obj, "colormap", None)
                if cm is None:
                    return _fallback_color(fallback_name)

                if isinstance(cm, str):
                    rgba = _hex_to_rgba(cm)
                    if rgba is not None:
                        return rgba
                    if cm.lower() in NAMED_RGBA:
                        return NAMED_RGBA[cm.lower()]
                    return _fallback_color(fallback_name)

                if isinstance(cm, tuple) and len(cm) >= 2:
                    name = cm[0]
                    cmap_obj = cm[1]

                    if isinstance(name, str):
                        rgba = _hex_to_rgba(name)
                        if rgba is not None:
                            return rgba
                        if name.lower() in NAMED_RGBA:
                            return NAMED_RGBA[name.lower()]

                    try:
                        if hasattr(cmap_obj, "map"):
                            rgba = cmap_obj.map([0.95])[0]
                            rgba = tuple(float(x) for x in rgba)
                            if len(rgba) == 4:
                                return (rgba[0], rgba[1], rgba[2], 0.85)
                    except Exception:
                        pass

                    return _fallback_color(fallback_name)

                try:
                    if hasattr(cm, "map"):
                        rgba = cm.map([0.95])[0]
                        rgba = tuple(float(x) for x in rgba)
                        if len(rgba) == 4:
                            return (rgba[0], rgba[1], rgba[2], 0.85)
                except Exception:
                    pass

                return _fallback_color(fallback_name)

            # ----------------------------
            # 4) Marching cubes per layer
            # ----------------------------
            made_any = False
            warnings = []

            for name in selected:
                try:
                    src_layer = self.viewer.layers[name]
                except Exception:
                    src_layer = None

                stack3 = self._resolve_stack3d(name)
                if stack3 is None:
                    warnings.append(f"{name}: could not resolve stack")
                    continue

                vol = np.asarray(stack3)
                if vol.ndim != 3:
                    warnings.append(f"{name}: not 3D (shape {vol.shape})")
                    continue

                volf = vol.astype(np.float32)

                vmin, vmax = np.percentile(volf, (1, 99))
                if vmax - vmin <= 1e-6:
                    norm = np.clip(volf, 0, 1)
                else:
                    norm = np.clip((volf - vmin) / (vmax - vmin), 0.0, 1.0)

                iso_level = 0.5
                if np.nanmax(norm) < iso_level:
                    warnings.append(f"{name}: no voxels above iso-level")
                    continue

                try:
                    verts, faces, normals, values = marching_cubes(
                        norm,
                        level=iso_level,
                        spacing=(spacing_z, spacing_y, spacing_x),
                        step_size=2,
                    )
                except Exception as e:
                    warnings.append(f"{name}: marching cubes failed ({e})")
                    continue

                rgba = _rgba_from_layer_current(src_layer, name)
                vertex_colors = np.tile(np.array(rgba, dtype=np.float32), (verts.shape[0], 1))

                v3.add_surface(
                    (verts, faces, values),
                    name=f"{name} surface",
                    shading="smooth",
                    vertex_colors=vertex_colors,
                )

                made_any = True

            if not made_any:
                QMessageBox.information(self, "3D Surfaces", "No surfaces were created.\n" + "\n".join(warnings[:8]))
                try:
                    v3.close()
                except Exception:
                    pass
                return None

            msg = f"3D viewer opened.\nVoxel size (Z,Y,X) = ({spacing_z}, {spacing_y}, {spacing_x}) µm."
            if warnings:
                msg += "\n\nNotes:\n" + "\n".join(warnings[:10])
            QMessageBox.information(self, "3D Surfaces", msg)

            return v3

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate 3D surfaces:\n{e}")
            return None

    def extract_roi_intensities_for_viewer(self, proj_viewer, projections):
        """
        Extract ROI intensities from a MAX-PROJECTION viewer.

        - Uses Shapes drawn on proj_viewer.
        - Prompts user to choose which *_max image layer to measure.
        - Measures on that chosen 2D max-projection image only (no coloc stack).
        - Writes results to an Excel file (safe-save temp then atomic replace).
        """
        try:
            # ----------------------------
            # 0) Find Shapes layers
            # ----------------------------
            try:
                from napari.layers import Shapes
            except Exception:
                Shapes = None

            shapes_layers = []
            for layer in proj_viewer.layers:
                if Shapes is not None and isinstance(layer, Shapes):
                    shapes_layers.append(layer)
                else:
                    if getattr(layer, "shape_type", None) is not None or layer.__class__.__name__ == "Shapes":
                        shapes_layers.append(layer)

            if len(shapes_layers) == 0:
                QMessageBox.critical(
                    proj_viewer.window.qt_viewer,
                    "No shape layers found to extract pixel intensities.",
                    "No shape layers found to extract pixel intensities.",
                )
                return

            # ----------------------------
            # 1) Ask which max-projection channel to quantify
            # ----------------------------
            # Prefer *_max layers, since generate_max_projection names them that way.
            max_img_layer_names = []
            for lyr in proj_viewer.layers:
                # shapes layers shouldn't be candidates
                if lyr in shapes_layers:
                    continue
                if getattr(lyr, "data", None) is None:
                    continue
                nm = getattr(lyr, "name", "") or ""
                if nm.endswith("_max"):
                    max_img_layer_names.append(nm)

            # Fallback: any non-shapes layer with 2D image data
            if not max_img_layer_names:
                for lyr in proj_viewer.layers:
                    if lyr in shapes_layers:
                        continue
                    try:
                        arr = np.asarray(lyr.data)
                    except Exception:
                        continue
                    if arr.ndim == 2:
                        max_img_layer_names.append(getattr(lyr, "name", "image"))

            if not max_img_layer_names:
                QMessageBox.critical(self, "No projection images", "No max-projection image layers found to quantify.")
                return

            chosen_layer_name, ok = QInputDialog.getItem(
                self,
                "Choose channel",
                "Which max-projection layer should be used for ROI intensity extraction?",
                max_img_layer_names,
                0,
                False,
            )
            if not ok or not chosen_layer_name:
                return

            try:
                img_layer = proj_viewer.layers[chosen_layer_name]  # napari supports indexing by unique name 
            except Exception:
                img_layer = None

            if img_layer is None or getattr(img_layer, "data", None) is None:
                QMessageBox.critical(self, "Missing data", f"Could not access layer data for: {chosen_layer_name}")
                return

            img2d = np.asarray(img_layer.data)
            if img2d.ndim != 2:
                QMessageBox.critical(self, "Unsupported", f"Selected layer is not 2D: {chosen_layer_name} has shape {img2d.shape}")
                return

            ydim, xdim = img2d.shape

            # ----------------------------
            # 2) Ask metadata
            # ----------------------------
            protein, ok = QInputDialog.getText(self, "Protein/labeling", "Enter the name or protein or what has been labeled :")
            if not ok or not protein:
                return

            brain_num, ok = QInputDialog.getText(self, "Image #, or ID", "Enter ID number:")
            if not ok or not brain_num:
                return

            slide_date, ok = QInputDialog.getText(self, "Image Date", "Enter image date:")
            if not ok or not slide_date:
                return

            file_name, ok = QInputDialog.getText(self, "File name", "File will be saved as ROI_intensity_<file name>.xlsx :")
            if not ok or not file_name:
                return

            out_path = Path(f"ROI_intensity_{file_name}.xlsx")

            bg_sub_answer, ok = QInputDialog.getItem(
                self,
                "Background Subtracted?",
                "Is this projection already background subtracted?",
                ["Yes", "No"],
                0,
                False,
            )
            if not ok:
                return

            # ----------------------------
            # 3) Excel setup
            # ----------------------------
            try:
                from openpyxl import Workbook, load_workbook
            except Exception:
                QMessageBox.critical(self, "Missing Dependency", "openpyxl is required to write Excel files. Please install it.")
                return

            region_names = [layer.name for layer in shapes_layers]
            region_bg_cols = [f"region_background_value_{name}" for name in region_names]

            headers = [
                "file_name",
                "protein_name",
                "brain_number",
                "slide_date",
                "background_subtracted",
                "quant_layer",
            ] + region_bg_cols + region_names

            if not out_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.append(headers)
                wb.save(out_path)

            wb = load_workbook(out_path)
            ws = wb.active

            # Map header -> column index
            header_to_col = {str(cell.value): i + 1 for i, cell in enumerate(ws[1]) if cell.value is not None}

            # Append row index (we will append at the end)
            append_row_idx = ws.max_row + 1

            tmp_path = out_path.with_name(out_path.stem + "_temp.xlsx")

            def safe_save():
                wb.save(tmp_path)

            # ----------------------------
            # 4) ROI processing
            # ----------------------------  

            results = {}
            bg_values_for_row = {}

            for layer in shapes_layers:
                name = layer.name or f"Region_{len(results) + 1}"
                bg_val = None

                # Ask per ROI if not already bg-subtracted
                if bg_sub_answer == "No":
                    bg_each, ok_each = QInputDialog.getItem(
                        self,
                        "Background Subtraction",
                        f"Background subtract region '{name}'?",
                        ["Yes", "No"],
                        1,
                        False,
                    )
                    if ok_each and bg_each == "Yes":
                        val, okv = QInputDialog.getDouble(
                            self,
                            "Background Value",
                            f"Enter background value for region '{name}':",
                            0.0,
                            0.0,
                            float(np.nanmax(img2d) if img2d.size else 65535.0),
                            3,
                        )
                        if okv:
                            bg_val = float(val)

                bg_values_for_row[name] = "" if bg_val is None else float(bg_val)

                # Build combined mask from all polygons in this shapes layer
                mask2d = np.zeros((ydim, xdim), dtype=bool)
                for shape in getattr(layer, "data", []):
                    try:
                        poly = np.asarray(shape)
                        if poly.ndim != 2 or poly.shape[0] < 3:
                            continue
                        # polygon2mask expects (row, col) coords 
                        poly_rc = np.column_stack([poly[:, 0], poly[:, 1]])
                        m = polygon2mask((ydim, xdim), poly_rc)
                        mask2d |= m.astype(bool)
                    except Exception:
                        continue

                if not mask2d.any():
                    results[name] = 0.0
                    continue

                roi_pixels = img2d[mask2d].astype(np.float32)

                if bg_val is not None:
                    roi_pixels = roi_pixels - float(bg_val)
                    roi_pixels[roi_pixels < 0] = 0.0

                nz = roi_pixels[roi_pixels > 0]
                avg_val = float(nz.mean()) if nz.size > 0 else 0.0
                results[name] = avg_val

                # Optional verification overlay (shows ROI pixels from chosen layer)
                try:
                    verification2d = np.zeros((ydim, xdim), dtype=np.float32)
                    verification2d[mask2d] = img2d[mask2d].astype(np.float32)
                    proj_viewer.add_image(verification2d, name=f"ROI_{name}", blending="additive", opacity=0.7)
                except Exception:
                    pass

            # ----------------------------
            # 5) Append final row + save atomically
            # ----------------------------
            file_name_val = str(getattr(self, "file_path", ""))  # fallback if you don't have app_state

            row = [
                file_name_val,
                protein,
                brain_num,
                slide_date,
                "Yes" if bg_sub_answer == "Yes" else "No",
                str(chosen_layer_name),
            ]

            for nm in region_names:
                row.append(bg_values_for_row.get(nm, ""))

            for nm in region_names:
                row.append(results.get(nm, 0.0))

            ws.append(row)
            safe_save()

            # Atomic replace
            tmp_path.replace(out_path)

            QMessageBox.information(self, "ROI Extraction Complete", f"ROI intensities written to: {out_path}")

        except Exception as e:
            # Delete failed temp file
            try:
                if "tmp_path" in locals() and tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass

            QMessageBox.critical(self, "Error", f"Failed ROI extraction (viewer):\n{e}")
            import traceback
            traceback.print_exc()

    def generate_max_projection(self):
        try:
            # -------- Build selectable items: use viewer layer names, but resolve stacks via _resolve_stack3d --------
            layer_names = [lyr.name for lyr in self.viewer.layers]
            if not layer_names:
                QMessageBox.warning(self, "No layers", "No layers found in viewer.")
                return

            # -------- Popup dialog with checklist --------
            dlg = QDialog(self)
            dlg.setWindowTitle("Max projection: select layers")
            dlg.setMinimumWidth(420)

            main = QVBoxLayout(dlg)
            main.addWidget(QLabel("Check layers to max-project (channels, masks, derived/coloc):"))

            lst = QListWidget()
            lst.setSelectionMode(QAbstractItemView.NoSelection)

            # Pre-check channels by default (optional)
            for nm in layer_names:
                it = QListWidgetItem(nm)
                it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
                default_checked = nm.startswith("Channel ") and nm[8:].isdigit()
                it.setCheckState(Qt.Checked if default_checked else Qt.Unchecked)
                lst.addItem(it)

            main.addWidget(lst)

            # Convenience buttons
            btn_row = QHBoxLayout()
            btn_all = QPushButton("Select all")
            btn_none = QPushButton("Select none")
            btn_row.addWidget(btn_all)
            btn_row.addWidget(btn_none)
            btn_row.addStretch(1)
            main.addLayout(btn_row)

            def _set_all(state):
                for i in range(lst.count()):
                    lst.item(i).setCheckState(state)

            btn_all.clicked.connect(lambda: _set_all(Qt.Checked))
            btn_none.clicked.connect(lambda: _set_all(Qt.Unchecked))

            # OK/Cancel
            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            main.addWidget(buttons)
            buttons.accepted.connect(dlg.accept)
            buttons.rejected.connect(dlg.reject)

            if dlg.exec() != QDialog.Accepted:
                return

            selected = []
            for i in range(lst.count()):
                it = lst.item(i)
                if it.checkState() == Qt.Checked:
                    selected.append(it.text())

            if not selected:
                QMessageBox.information(self, "MaxProj", "No layers selected.")
                return

            # -------- Create projection viewer --------
            proj = napari.Viewer()
            proj.scale_bar.visible = True
            proj.scale_bar.unit = 'um'
            proj.scale_bar.font_size = 12
            proj.scale_bar.color = 'white'
            projections = {}
            proj_layers = []

            for name in selected:
                stack3 = self._resolve_stack3d(name)
                if stack3 is None:
                    # skip silently or warn; here we warn once per missing
                    QMessageBox.warning(self, "Missing data", f"Could not resolve full stack for: {name}")
                    continue

                arr = np.asarray(stack3)
                if arr.ndim == 3:
                    mp = arr.max(axis=0)
                elif arr.ndim == 2:
                    mp = arr
                else:
                    QMessageBox.warning(self, "Unsupported", f"{name} has unsupported dims: {arr.shape}")
                    continue

                # choose a colormap: reuse existing layer colormap if possible
                cmap = None
                try:
                    src_layer = self._get_layer_by_name(name)
                    cmap = getattr(src_layer, "colormap", None) or getattr(src_layer, "colormap_name", None)
                except Exception:
                    cmap = None

                # reasonable defaults
                if cmap is None:
                    cmap = "gray"

                try:
                    lay = proj.add_image(
                        mp,
                        name=f"{name}_max",
                        colormap=cmap,
                        blending="additive",
                        opacity=0.7,
                    )
                except Exception:
                    lay = proj.add_image(mp, name=f"{name}_max")

                proj_layers.append(lay)
                projections[name] = mp

            if not proj_layers:
                QMessageBox.information(self, "MaxProj", "No projections were created.")
                try:
                    proj.close()
                except Exception:
                    pass
                return

            # -------- Controls dock (screenshot + per-layer opacity + ROI) --------
            widget = QWidget()
            vlay = QVBoxLayout()

            ss_btn = QPushButton("Save Screenshot")
            ss_btn.clicked.connect(
                lambda _=False: self.save_screenshot(
                    viewer=proj,
                    parent=self,
                    default_name="max_projection.png",
                    canvas_only=True,
                )
            )
            vlay.addWidget(ss_btn)

            try:
                for lay in proj_layers:
                    vlay.addWidget(QLabel(lay.name))
                    sl = QSlider(Qt.Horizontal)
                    sl.setRange(0, 100)
                    sl.setValue(int(getattr(lay, "opacity", 0.7) * 100))
                    sl.valueChanged.connect(lambda val, L=lay: setattr(L, "opacity", val / 100.0))
                    vlay.addWidget(sl)
            except Exception:
                pass

            roi_btn = QPushButton("Extract ROI Intensities")
            roi_btn.clicked.connect(lambda: self.extract_roi_intensities_for_viewer(proj, projections))
            vlay.addWidget(roi_btn)

            widget.setLayout(vlay)
            proj.window.add_dock_widget(widget, area="right")

            QMessageBox.information(self, "MaxProj", f"Max projection viewer opened ({len(proj_layers)} layers).")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Max projection failed: {e}")

    def save_video(self):
        try:
            save_path, _ = QFileDialog.getSaveFileName(self, 'Save Video', '', 'MP4 Video (*.mp4)')
            if not save_path:
                return
            if not save_path.lower().endswith('.mp4'):
                save_path += '.mp4'
            frames = []
            for z in range(self.n_z):
                self.z_slider.setValue(z)
                QApplication.processEvents()
                frames.append(self.viewer.screenshot(canvas_only=True))
            writer = imageio.get_writer(save_path, format='ffmpeg', mode='I', fps=10, codec='libx264')
            for f in frames:
                writer.append_data(f)
            writer.close()
            QMessageBox.information(self, 'Saved', f'Video saved to {save_path}')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save video: {e}')

    def save_screenshot(self, viewer=None, parent=None, default_name="screenshot.png", canvas_only=True):
        """
        One screenshot function for any napari Viewer.
        - viewer: napari.Viewer (defaults to self.viewer)
        - parent: QWidget for dialogs/message boxes (defaults to self)
        """
        try:
            viewer = viewer if viewer is not None else self.viewer
            parent = parent if parent is not None else self

            save_path, _ = QFileDialog.getSaveFileName(
                parent,
                "Save Screenshot",
                default_name,
                "PNG (*.png);;JPEG (*.jpg *.jpeg)"
            )
            if not save_path:
                return

            if not any(save_path.lower().endswith(ext) for ext in [".png", ".jpg", ".jpeg"]):
                save_path += ".png"

            img = viewer.screenshot(canvas_only=canvas_only)  # returns RGBA ndarray 
            imageio.imwrite(save_path, img)

            QMessageBox.information(parent, "Saved", f"Screenshot saved to {save_path}")

        except Exception as e:
            QMessageBox.critical(parent if parent is not None else self, "Error", f"Failed to save screenshot: {e}")


###############################################################################
# ------------------------------ MAIN APP ------------------------------------
###############################################################################

def main():

    # Ensure QApplication exists (only one per process).
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)

    # File selection dialog
    file_dialog = QFileDialog()
    file_dialog.setWindowTitle("Select LIF or TIFF file")
    file_dialog.setNameFilters(["LIF files (*.lif)", "TIFF files (*.tif *.tiff)", "All files (*)"])
    file_dialog.setFileMode(QFileDialog.ExistingFile)

    if file_dialog.exec_():
        file_path = Path(file_dialog.selectedFiles()[0])
    else:
        print("No file selected.")
        return

    print(f"Loading: {file_path}")

    try:
        # Load (memmap)
        if file_path.suffix.lower() == ".lif":
            channel_mmaps, contrast_limits, voxel_size, temp_paths = load_lif_memmap(file_path, max_ch=4)
        else:
            channel_mmaps, contrast_limits, voxel_size, temp_paths = load_tiff_memmap(file_path, max_ch=4)

        print(f"Loaded {len(channel_mmaps)} channels (memmaps)")

        # Create napari viewer
        viewer = napari.Viewer()
        viewer.scale_bar.visible = True
        viewer.scale_bar.unit = "um"
        viewer.scale_bar.font_size = 12
        viewer.scale_bar.color = "white"

        # Add channel layers
        channel_layers = []
        default_colormaps = ["blue", "green", "red", "magenta"]
        for idx, ch in enumerate(channel_mmaps, start=1):
            name = f"Channel {idx}"
            cmap = default_colormaps[idx - 1] if idx - 1 < len(default_colormaps) else "gray"
            layer = viewer.add_image(ch, name=name, opacity=0.5, colormap=cmap, blending="additive")
            channel_layers.append(layer)

        # Dock widget
        widget = UIWidget(viewer, channel_mmaps, channel_layers, contrast_limits, file_path)
        widget._temp_paths = list(temp_paths) if temp_paths is not None else []
        try:
            widget.voxel_size = voxel_size
        except Exception:
            pass
        viewer.window.add_dock_widget(widget, area="right")

        # ---- QUIT / CLEANUP (no widget.cleanup) ----
        def _cleanup_and_stop_workers():
            # 1) remove temp memmap files
            for p in getattr(widget, "_temp_paths", []) or []:
                try:
                    os.remove(str(p))
                except Exception:
                    pass

            # 2) ask napari workers to quit; prevents hanging exit 
            try:
                from napari.qt.threading import WorkerBase
                WorkerBase.await_workers(msecs=2000)
            except Exception:
                pass

        try:
            app.aboutToQuit.connect(_cleanup_and_stop_workers)
        except Exception:
            pass

        # Run napari event loop; returns after last viewer closed.
        napari.run()

        # One more short wait for workers right before exiting 
        try:
            from napari.qt.threading import WorkerBase
            WorkerBase.await_workers(msecs=500)
        except Exception:
            pass

        sys.exit(0)
        import threading
        print([t.name for t in threading.enumerate()])

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error loading file: {e}")
        QMessageBox.critical(None, "Error", f"Failed to load file:\n{str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()